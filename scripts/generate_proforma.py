"""
DealGap IQ — Financial Proforma Excel Generator
Generates a multi-sheet Excel workbook with a centralized cost engine
that auto-selects the optimal provider plan as subscriber volume scales.
"""

import math
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ═══════════════════════════════════════════════════════════════════════════════
# COST ENGINE — Provider plan tiers & auto-selection
# ═══════════════════════════════════════════════════════════════════════════════

# Each API provider: calls_per_search and tier list
# Tier: (name, monthly_fee, included_quota, overage_rate | None)
# overage_rate=None means no overage — must upgrade when quota exceeded.
API_PROVIDERS = {
    "RentCast": {
        "calls_per_search": 4,
        "tiers": [
            ("Developer",       0,      50,      0.20),
            ("Foundation",      74,     1000,    0.06),
            ("Growth",          199,    5000,    0.03),
            ("Scale",           449,    25000,   0.015),
        ],
    },
    "Zillow / AXESSO": {
        "calls_per_search": 1.5,
        "tiers": [
            ("Free Trial",      0,      10,      None),
            ("Starter",         16,     10000,   None),
            ("Production",      82,     100000,  None),
            ("Business",        164,    300000,  None),
            ("Ent. Basic",      493,    1000000, None),
            ("Ent. Premium",    877,    2500000, 0.001),
        ],
    },
    "RapidAPI — Redfin": {
        "calls_per_search": 2,
        "tiers": [
            ("Basic",           0,      100,     None),
            ("Pro",             15,     1000,    None),
            ("Ultra",           50,     5000,    None),
            ("Mega",            100,    10000,   None),
            ("Custom",          250,    50000,   0.01),
        ],
    },
    "RapidAPI — Realtor": {
        "calls_per_search": 2,
        "tiers": [
            ("Basic",           0,      100,     None),
            ("Pro",             15,     1000,    None),
            ("Ultra",           50,     5000,    None),
            ("Mega",            100,    10000,   None),
            ("Custom",          250,    50000,   0.01),
        ],
    },
}

# Infrastructure: scales by total paid subscribers (stepped tiers)
# Tier: (name, monthly_cost, max_paid_subs)
INFRA_SERVICES = {
    "Railway": [
        ("Pro",         20,     200),
        ("Pro (scaled)", 50,    500),
        ("Pro (heavy)",  100,   1500),
        ("Pro (high)",   200,   5000),
    ],
    "Vercel": [
        ("Pro",          20,    500),
        ("Pro (scaled)", 40,    2000),
        ("Enterprise",   100,   10000),
    ],
    "Expo (EAS)": [
        ("Starter",      19,    150),
        ("Production",   199,   2500),
        ("Custom",       499,   10000),
    ],
}

# Fixed overhead — does not scale
FIXED_OVERHEAD = {
    "Apple Developer":   8.25,
    "Google Play Dev":   2.08,
    "Domain":            1.50,
    "Sentry":            0.00,
    "Resend":            0.00,
}

# Model constants
BLENDED_NET_ARPU = 34.12
BLENDED_GROSS_PRICE = 36.99
ACTIVITY_RATE = 0.60
SEARCHES_PER_ACTIVE = 15
CACHE_HIT_RATE = 0.30
ANTHROPIC_PER_CALL = 0.01
MOBILE_REVENUE_PCT = 0.20
REVENUECAT_MTR_THRESHOLD = 2500
MAPS_FREE_CREDIT = 200
MAPS_COST_PER_REQ = 0.003
MAPS_REQS_PER_SEARCH = 3


def find_optimal_plan(calls_needed, tiers):
    """Pick cheapest plan for a given call volume. Returns (name, total_cost)."""
    best_plan = tiers[-1][0]
    best_cost = float("inf")

    for name, fee, quota, overage_rate in tiers:
        if overage_rate is not None:
            excess = max(0, calls_needed - quota)
            total = fee + excess * overage_rate
        else:
            if calls_needed <= quota:
                total = fee
            else:
                continue
        if total < best_cost:
            best_cost = total
            best_plan = name

    return best_plan, round(best_cost, 2)


def find_infra_tier(paid_subs, tiers):
    """Pick infrastructure tier by subscriber count. Returns (name, cost)."""
    for name, cost, max_subs in tiers:
        if paid_subs <= max_subs:
            return name, cost
    return tiers[-1][0], tiers[-1][1]


def compute_all_costs(paid_subs):
    """
    Central cost engine.  Returns dict of category → {provider → (plan, cost)}
    and a grand total.
    """
    active = max(1, math.ceil(paid_subs * ACTIVITY_RATE)) if paid_subs > 0 else 0
    cold_searches = active * SEARCHES_PER_ACTIVE * (1 - CACHE_HIT_RATE)

    result = {
        "api": {},
        "infra": {},
        "variable": {},
        "fixed": {},
    }

    # ── API providers ──
    for prov, cfg in API_PROVIDERS.items():
        calls = cold_searches * cfg["calls_per_search"]
        plan, cost = find_optimal_plan(calls, cfg["tiers"])
        result["api"][prov] = (plan, cost, calls)

    # ── Infrastructure ──
    for svc, tiers in INFRA_SERVICES.items():
        plan, cost = find_infra_tier(paid_subs, tiers)
        result["infra"][svc] = (plan, cost)

    # ── Variable / usage-based ──
    anthropic = active * SEARCHES_PER_ACTIVE * ANTHROPIC_PER_CALL
    result["variable"]["Anthropic (Claude)"] = ("Pay-per-use", round(anthropic, 2))

    maps_raw = cold_searches * MAPS_REQS_PER_SEARCH * MAPS_COST_PER_REQ
    maps_cost = max(0, maps_raw - MAPS_FREE_CREDIT)
    result["variable"]["Google Maps"] = ("Pay-as-you-go", round(maps_cost, 2))

    mobile_mtr = paid_subs * BLENDED_GROSS_PRICE * MOBILE_REVENUE_PCT
    rc_fee = max(0, (mobile_mtr - REVENUECAT_MTR_THRESHOLD) * 0.01)
    result["variable"]["RevenueCat"] = (
        "Free" if rc_fee == 0 else "1% MTR", round(rc_fee, 2)
    )

    # ── Fixed overhead ──
    for svc, cost in FIXED_OVERHEAD.items():
        result["fixed"][svc] = ("Fixed", cost)

    # ── Totals ──
    api_total = sum(c for _, c, *_ in result["api"].values())
    infra_total = sum(c for _, c in result["infra"].values())
    var_total = sum(c for _, c in result["variable"].values())
    fixed_total = sum(c for _, c in result["fixed"].values())
    grand_total = api_total + infra_total + var_total + fixed_total

    result["_totals"] = {
        "api": round(api_total, 2),
        "infra": round(infra_total, 2),
        "variable": round(var_total, 2),
        "fixed": round(fixed_total, 2),
        "grand": round(grand_total, 2),
    }

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# STYLING
# ═══════════════════════════════════════════════════════════════════════════════
HEADER_BG = "1A2332"
ACCENT_BLUE = "3B82F6"
ACCENT_GREEN = "22C55E"
ACCENT_RED = "EF4444"
ACCENT_YELLOW = "F59E0B"
LIGHT_GRAY = "F1F5F9"
MED_GRAY = "94A3B8"
ROW_ALT = "F8FAFC"
BORDER_COLOR = "CBD5E1"
WHITE = "FFFFFF"

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)
header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=14, color=ACCENT_BLUE)
section_font = Font(name="Calibri", bold=True, size=12, color=HEADER_BG)
bold_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=11)
small_font = Font(name="Calibri", size=10, color=MED_GRAY)
green_font = Font(name="Calibri", bold=True, size=11, color=ACCENT_GREEN)
red_font = Font(name="Calibri", bold=True, size=11, color=ACCENT_RED)
green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
alt_fill = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
blue_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

pct_fmt = "0.0%"
int_fmt = "#,##0"
acct_fmt = '$#,##0.00'
acct_fmt_neg = '$#,##0.00;[Red]($#,##0.00)'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_row(ws, row, max_col, is_alt=False):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        if is_alt:
            cell.fill = alt_fill


def style_total_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = bold_font
        cell.border = Border(
            top=Side(style="double", color=HEADER_BG),
            bottom=Side(style="double", color=HEADER_BG),
            left=Side(style="thin", color=BORDER_COLOR),
            right=Side(style="thin", color=BORDER_COLOR),
        )
        cell.fill = light_fill


def auto_width(ws, min_width=10, max_width=28):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = min_width
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)) + 3, max_width))
        ws.column_dimensions[col_letter].width = length


# ═══════════════════════════════════════════════════════════════════════════════
# SUB LEVELS used across multiple sheets
# ═══════════════════════════════════════════════════════════════════════════════
SUB_LEVELS = [0, 1, 5, 10, 12, 15, 25, 50, 75, 100, 150, 200, 250, 300,
              400, 500, 750, 1000, 1500, 2000, 3000, 5000]


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = ACCENT_BLUE

    ws.merge_cells("A1:F1")
    ws["A1"] = "DealGap IQ — Financial Proforma"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=ACCENT_BLUE)
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:F2")
    ws["A2"] = "Breakeven & Profitability Analysis  |  April 2026"
    ws["A2"].font = Font(name="Calibri", size=12, color=MED_GRAY)

    costs_0 = compute_all_costs(0)
    base_burn = costs_0["_totals"]["grand"]
    breakeven = math.ceil(base_burn / BLENDED_NET_ARPU)

    costs_100 = compute_all_costs(100)
    net_rev_100 = 100 * BLENDED_NET_ARPU
    margin_100 = (net_rev_100 - costs_100["_totals"]["grand"]) / net_rev_100

    r = 4
    kpis = [
        ("Monthly Burn (0 subs)", f"${base_burn:,.2f}", ACCENT_RED),
        ("Breakeven Subscribers", f"{breakeven} Pro", ACCENT_BLUE),
        ("Blended Net ARPU", f"${BLENDED_NET_ARPU}/mo", ACCENT_GREEN),
        (f"Margin @ 100 Subs", f"{margin_100:.0%}", ACCENT_GREEN),
        ("Primary Cost Driver", "RentCast (51%)", ACCENT_YELLOW),
        ("First Scale Trigger", "~48–80 subs", ACCENT_YELLOW),
    ]
    for i, (label, value, color) in enumerate(kpis):
        col = (i % 3) * 2 + 1
        row = r if i < 3 else r + 3
        ws.cell(row=row, column=col, value=label).font = Font(name="Calibri", size=10, color=MED_GRAY)
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        val_cell = ws.cell(row=row + 1, column=col, value=value)
        val_cell.font = Font(name="Calibri", bold=True, size=16, color=color)
        val_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)

    # ── Pricing ──
    r = 11
    ws.cell(row=r, column=1, value="SUBSCRIPTION PRICING").font = section_font
    r += 1
    headers = ["Tier", "Monthly", "Annual", "Eff. Monthly (Annual)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))
    for row_data in [("Starter (Free)", 0, 0, 0), ("Pro", 39.99, 349.99, 29.17)]:
        r += 1
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c > 1:
                cell.number_format = acct_fmt
        style_data_row(ws, r, len(headers), r % 2 == 0)

    # ── Net Revenue ──
    r += 2
    ws.cell(row=r, column=1, value="NET REVENUE AFTER PROCESSING FEES").font = section_font
    r += 1
    headers = ["Channel", "Fee Structure", "Monthly Net", "Annual Net (/mo)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))
    for row_data in [
        ("Stripe (Web)", "2.9% + $0.30/txn", 38.53, 28.30),
        ("Apple IAP (15% SBP)", "15% commission", 33.99, 24.79),
        ("Google Play (15%)", "15% commission", 33.99, 24.79),
    ]:
        r += 1
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c >= 3:
                cell.number_format = acct_fmt
        style_data_row(ws, r, len(headers), r % 2 == 0)

    # ── Blended ARPU ──
    r += 2
    ws.cell(row=r, column=1, value="BLENDED NET ARPU").font = section_font
    r += 1
    headers = ["Segment", "Mix %", "Net $/mo", "Contribution"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))
    for row_data in [
        ("Web — Monthly", 0.52, 38.53, 20.04),
        ("Web — Annual", 0.28, 28.30, 7.92),
        ("iOS — Monthly", 0.13, 33.99, 4.42),
        ("iOS — Annual", 0.07, 24.79, 1.74),
    ]:
        r += 1
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 2:
                cell.number_format = pct_fmt
            elif c >= 3:
                cell.number_format = acct_fmt
        style_data_row(ws, r, len(headers), r % 2 == 0)
    r += 1
    for c, v in enumerate(["Blended ARPU (net)", 1.0, "", BLENDED_NET_ARPU], 1):
        cell = ws.cell(row=r, column=c, value=v)
        if c == 2:
            cell.number_format = pct_fmt
        elif c == 4:
            cell.number_format = acct_fmt
    style_total_row(ws, r, len(headers))

    auto_width(ws, min_width=12, max_width=30)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Fixed Costs (baseline at 0 subs)
# ═══════════════════════════════════════════════════════════════════════════════
def build_fixed_costs(wb):
    ws = wb.create_sheet("Fixed Costs")
    ws.sheet_properties.tabColor = ACCENT_RED

    ws.merge_cells("A1:G1")
    ws["A1"] = "Baseline Monthly Costs — Current Plans (0 Subscribers)"
    ws["A1"].font = title_font

    baseline = compute_all_costs(0)

    r = 3
    headers = ["#", "Service", "Category", "Plan", "Monthly Cost", "Included Quota", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))

    rows = [
        (1,  "RentCast API",        "Data API",  baseline["api"]["RentCast"][0],             baseline["api"]["RentCast"][1],             "5,000 req/mo",  "$0.03/req overage"),
        (2,  "Zillow / AXESSO",     "Data API",  baseline["api"]["Zillow / AXESSO"][0],      baseline["api"]["Zillow / AXESSO"][1],      "100,000 req/mo", "€75 ≈ $82 USD"),
        (3,  "RapidAPI — Redfin",   "Data API",  baseline["api"]["RapidAPI — Redfin"][0],    baseline["api"]["RapidAPI — Redfin"][1],    "~1,000 req/mo", "Verify quota"),
        (4,  "RapidAPI — Realtor",  "Data API",  baseline["api"]["RapidAPI — Realtor"][0],   baseline["api"]["RapidAPI — Realtor"][1],   "~1,000 req/mo", "Verify quota"),
        (5,  "Railway",             "Infra",     baseline["infra"]["Railway"][0],             baseline["infra"]["Railway"][1],             "$20 usage credit", "Backend hosting"),
        (6,  "Vercel",              "Infra",     baseline["infra"]["Vercel"][0],              baseline["infra"]["Vercel"][1],              "Standard limits",  "Frontend hosting"),
        (7,  "Expo (EAS)",          "Infra",     baseline["infra"]["Expo (EAS)"][0],          baseline["infra"]["Expo (EAS)"][1],          "3K MAUs",          "$45 build credit"),
        (8,  "Apple Developer",     "Fixed",     "Annual",  8.25,  "—", "$99/yr amortized"),
        (9,  "Google Play Dev",     "Fixed",     "One-time", 2.08, "—", "$25 amortized 12 mo"),
        (10, "Domain",              "Fixed",     "Annual",  1.50,  "—", "~$18/yr amortized"),
        (11, "Anthropic (Claude)",  "Variable",  "Pay-per-use", 0.00, "—", "~$0.01/narrative"),
        (12, "Google Maps",         "Variable",  "Pay-as-you-go", 0.00, "$200/mo free", "Autocomplete + Geo"),
        (13, "Sentry",              "Fixed",     "Free",    0.00,  "5K errors/mo", "Error monitoring"),
        (14, "Resend",              "Fixed",     "Free",    0.00,  "3,000 emails/mo", "Transactional email"),
        (15, "RevenueCat",          "Variable",  "Free",    0.00,  "< $2,500 MTR", "1% above threshold"),
    ]

    for i, row_data in enumerate(rows):
        r += 1
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 5 and isinstance(v, (int, float)):
                cell.number_format = acct_fmt
        style_data_row(ws, r, len(headers), i % 2 == 1)

    r += 1
    ws.cell(row=r, column=2, value="TOTAL BASELINE")
    total_cell = ws.cell(row=r, column=5, value=f"=SUM(E4:E{r-1})")
    total_cell.number_format = acct_fmt
    style_total_row(ws, r, len(headers))

    # ── Additional Costs ──
    r += 2
    ws.merge_cells(f"A{r}:G{r}")
    ws.cell(row=r, column=1, value="Additional Costs to Consider (Not in Baseline)").font = section_font
    r += 1
    h2 = ["", "Item", "", "Est. Monthly", "", "", "Notes"]
    for c, h in enumerate(h2, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(h2))
    for i, row_data in enumerate([
        ("", "GitHub (private repos)", "", "0–4", "", "", "Free for individual; Team $4/user/mo"),
        ("", "Cursor IDE", "", "20", "", "", "Development tool subscription"),
        ("", "Accounting / Bookkeeping", "", "0–50", "", "", "Manual or service"),
        ("", "Business Insurance", "", "0–100", "", "", "E&O / general liability"),
        ("", "Legal (entity, ToS)", "", "~50", "", "", "$600/yr amortized"),
        ("", "Marketing / Ads", "", "0+", "", "", "Variable; CAC dependent"),
        ("", "Founder Salary", "", "0+", "", "", "Critical at scale"),
    ]):
        r += 1
        for c, v in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=v)
        style_data_row(ws, r, len(h2), i % 2 == 1)

    auto_width(ws, min_width=10, max_width=32)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Profitability (uses cost engine)
# ═══════════════════════════════════════════════════════════════════════════════
def build_profitability(wb):
    ws = wb.create_sheet("Profitability")
    ws.sheet_properties.tabColor = ACCENT_GREEN

    ws.merge_cells("A1:K1")
    ws["A1"] = "Profitability by Subscriber Count (Auto-Scaled Costs)"
    ws["A1"].font = title_font

    ws["A3"] = "All costs computed by the centralized cost engine. Provider plans"
    ws["A3"].font = small_font
    ws["A4"] = "auto-upgrade when volume exceeds quota. See 'Scaling Detail' sheet."
    ws["A4"].font = small_font

    r = 6
    headers = [
        "Paid Subs", "Active (60%)", "Net Revenue",
        "Data APIs", "Infrastructure", "AI & Variable",
        "Fixed Overhead", "Total Costs", "Profit/Loss", "Margin"
    ]
    ncols = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, ncols)

    for i, subs in enumerate(SUB_LEVELS):
        r += 1
        costs = compute_all_costs(subs)
        t = costs["_totals"]
        net_rev = subs * BLENDED_NET_ARPU
        profit = net_rev - t["grand"]
        margin = profit / net_rev if net_rev > 0 else 0
        active = max(1, math.ceil(subs * ACTIVITY_RATE)) if subs > 0 else 0

        vals = [
            subs, active, net_rev,
            t["api"], t["infra"], t["variable"],
            t["fixed"], t["grand"], profit, margin,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c <= 2:
                cell.number_format = int_fmt
            elif c == ncols:
                cell.number_format = pct_fmt
            elif c == ncols - 1:
                cell.number_format = acct_fmt_neg
            else:
                cell.number_format = acct_fmt
        style_data_row(ws, r, ncols, i % 2 == 1)

        pnl_cell = ws.cell(row=r, column=ncols - 1)
        pnl_cell.font = green_font if profit >= 0 else red_font

    last_row = r

    # ── Profit/Loss chart ──
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Profit/Loss by Subscriber Count"
    chart.y_axis.title = "$ / Month"
    chart.x_axis.title = "Paid Subscribers"
    chart.style = 10
    chart.width = 28
    chart.height = 14
    cats = Reference(ws, min_col=1, min_row=8, max_row=last_row)
    data = Reference(ws, min_col=ncols - 1, min_row=6, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, f"A{last_row + 3}")

    # ── Total cost stacked chart ──
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.title = "Cost Composition by Subscriber Count"
    chart2.y_axis.title = "$ / Month"
    chart2.x_axis.title = "Paid Subscribers"
    chart2.style = 10
    chart2.width = 28
    chart2.height = 14
    for col_idx in [4, 5, 6, 7]:  # Data APIs, Infra, Variable, Fixed
        d = Reference(ws, min_col=col_idx, min_row=6, max_row=last_row)
        chart2.add_data(d, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, f"A{last_row + 20}")

    auto_width(ws, min_width=12, max_width=22)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Scaling Detail (per-provider cost & plan at each sub level)
# ═══════════════════════════════════════════════════════════════════════════════
def build_scaling_detail(wb):
    ws = wb.create_sheet("Scaling Detail")
    ws.sheet_properties.tabColor = ACCENT_YELLOW

    ws.merge_cells("A1:P1")
    ws["A1"] = "Per-Provider Cost & Plan at Each Subscriber Level"
    ws["A1"].font = title_font
    ws["A2"] = "Plans auto-upgrade when call volume exceeds quota. Yellow = upgraded from baseline."
    ws["A2"].font = small_font

    providers_api = list(API_PROVIDERS.keys())
    providers_infra = list(INFRA_SERVICES.keys())
    variable_items = ["Anthropic (Claude)", "Google Maps", "RevenueCat"]

    # Build headers: Subs | Active | then pairs of (Plan, Cost) for each provider
    all_providers = providers_api + providers_infra + variable_items
    r = 4
    headers = ["Paid Subs", "Active"]
    for p in all_providers:
        headers.append(f"{p} Plan")
        headers.append(f"{p} $")
    headers.append("TOTAL COST")
    ncols = len(headers)

    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, ncols)

    # Get baseline plans for highlighting upgrades
    baseline = compute_all_costs(0)
    baseline_plans = {}
    for p in providers_api:
        baseline_plans[p] = baseline["api"][p][0]
    for p in providers_infra:
        baseline_plans[p] = baseline["infra"][p][0]

    for i, subs in enumerate(SUB_LEVELS):
        r += 1
        costs = compute_all_costs(subs)
        active = max(1, math.ceil(subs * ACTIVITY_RATE)) if subs > 0 else 0

        ws.cell(row=r, column=1, value=subs).number_format = int_fmt
        ws.cell(row=r, column=2, value=active).number_format = int_fmt

        col = 3
        for p in providers_api:
            plan, cost, _calls = costs["api"][p]
            plan_cell = ws.cell(row=r, column=col, value=plan)
            cost_cell = ws.cell(row=r, column=col + 1, value=cost)
            cost_cell.number_format = acct_fmt
            if plan != baseline_plans.get(p):
                plan_cell.fill = yellow_fill
                plan_cell.font = bold_font
            col += 2

        for p in providers_infra:
            plan, cost = costs["infra"][p]
            plan_cell = ws.cell(row=r, column=col, value=plan)
            cost_cell = ws.cell(row=r, column=col + 1, value=cost)
            cost_cell.number_format = acct_fmt
            if plan != baseline_plans.get(p):
                plan_cell.fill = yellow_fill
                plan_cell.font = bold_font
            col += 2

        for p in variable_items:
            plan, cost = costs["variable"][p]
            ws.cell(row=r, column=col, value=plan)
            ws.cell(row=r, column=col + 1, value=cost).number_format = acct_fmt
            col += 2

        ws.cell(row=r, column=col, value=costs["_totals"]["grand"]).number_format = acct_fmt
        style_data_row(ws, r, ncols, i % 2 == 1)

    auto_width(ws, min_width=8, max_width=18)
    ws.column_dimensions["A"].width = 10
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — API Capacity
# ═══════════════════════════════════════════════════════════════════════════════
def build_api_capacity(wb):
    ws = wb.create_sheet("API Capacity")
    ws.sheet_properties.tabColor = ACCENT_YELLOW

    ws.merge_cells("A1:H1")
    ws["A1"] = "API Capacity & Plan Transitions"
    ws["A1"].font = title_font

    # ── Per-provider plan tiers ──
    for prov_name, cfg in API_PROVIDERS.items():
        r = ws.max_row + 2
        ws.cell(row=r, column=1, value=f"{prov_name} — PLAN TIERS").font = section_font
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        r += 1
        h = ["Plan", "Monthly Fee", "Included Quota", "Overage Rate",
             "Max Active Subs", "Max Paid Subs (60%)", "Calls/Search"]
        for c, val in enumerate(h, 1):
            ws.cell(row=r, column=c, value=val)
        style_header_row(ws, r, len(h))

        cps = cfg["calls_per_search"]
        for i, (name, fee, quota, ovg) in enumerate(cfg["tiers"]):
            r += 1
            searches_at_cap = quota / cps if cps > 0 else 0
            active_at_cap = searches_at_cap / (SEARCHES_PER_ACTIVE * (1 - CACHE_HIT_RATE))
            paid_at_cap = active_at_cap / ACTIVITY_RATE

            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=fee).number_format = acct_fmt
            ws.cell(row=r, column=3, value=quota).number_format = int_fmt
            ws.cell(row=r, column=4, value=f"${ovg}/req" if ovg else "n/a (must upgrade)")
            ws.cell(row=r, column=5, value=round(active_at_cap)).number_format = int_fmt
            ws.cell(row=r, column=6, value=round(paid_at_cap)).number_format = int_fmt
            ws.cell(row=r, column=7, value=cps)
            style_data_row(ws, r, len(h), i % 2 == 1)

    # ── Infrastructure tiers ──
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="INFRASTRUCTURE SCALING TIERS").font = section_font
    r += 1
    h2 = ["Service", "Plan", "Monthly Cost", "Max Paid Subs"]
    for c, val in enumerate(h2, 1):
        ws.cell(row=r, column=c, value=val)
    style_header_row(ws, r, len(h2))

    for svc, tiers in INFRA_SERVICES.items():
        for i, (name, cost, max_s) in enumerate(tiers):
            r += 1
            ws.cell(row=r, column=1, value=svc)
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=cost).number_format = acct_fmt
            ws.cell(row=r, column=4, value=max_s).number_format = int_fmt
            style_data_row(ws, r, len(h2), i % 2 == 1)

    # ── Plan transition summary ──
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="PLAN UPGRADE TRIGGER POINTS").font = section_font
    r += 1
    h3 = ["Provider", "Current Plan", "Upgrade To", "Trigger (Paid Subs)", "Cost Change"]
    for c, val in enumerate(h3, 1):
        ws.cell(row=r, column=c, value=val)
    style_header_row(ws, r, len(h3))

    prev_plans = {}
    for subs in SUB_LEVELS:
        costs = compute_all_costs(subs)
        for prov in API_PROVIDERS:
            plan = costs["api"][prov][0]
            if prov not in prev_plans:
                prev_plans[prov] = plan
                continue
            if plan != prev_plans[prov]:
                r += 1
                old_cost = compute_all_costs(subs - 1)["api"][prov][1]
                new_cost = costs["api"][prov][1]
                ws.cell(row=r, column=1, value=prov)
                ws.cell(row=r, column=2, value=prev_plans[prov])
                ws.cell(row=r, column=3, value=plan)
                ws.cell(row=r, column=4, value=f"~{subs} subs")
                ws.cell(row=r, column=5, value=new_cost - old_cost).number_format = acct_fmt_neg
                style_data_row(ws, r, len(h3))
                ws.cell(row=r, column=5).fill = yellow_fill
                prev_plans[prov] = plan

        for svc in INFRA_SERVICES:
            plan = costs["infra"][svc][0]
            if svc not in prev_plans:
                prev_plans[svc] = plan
                continue
            if plan != prev_plans[svc]:
                r += 1
                old_cost = compute_all_costs(subs - 1)["infra"][svc][1]
                new_cost = costs["infra"][svc][1]
                ws.cell(row=r, column=1, value=svc)
                ws.cell(row=r, column=2, value=prev_plans[svc])
                ws.cell(row=r, column=3, value=plan)
                ws.cell(row=r, column=4, value=f"~{subs} subs")
                ws.cell(row=r, column=5, value=new_cost - old_cost).number_format = acct_fmt_neg
                style_data_row(ws, r, len(h3))
                ws.cell(row=r, column=5).fill = yellow_fill
                prev_plans[svc] = plan

    auto_width(ws, min_width=12, max_width=30)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Sensitivity Analysis
# ═══════════════════════════════════════════════════════════════════════════════
def build_sensitivity(wb):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = "8B5CF6"

    base_costs = compute_all_costs(0)
    base_fixed = base_costs["_totals"]["grand"]

    ws.merge_cells("A1:G1")
    ws["A1"] = "Breakeven Sensitivity Analysis"
    ws["A1"].font = title_font

    ws["A3"] = "BREAKEVEN BY CHANNEL MIX"
    ws["A3"].font = section_font
    r = 4
    headers = ["Scenario", "Blended Net ARPU", "Base Fixed Costs", "Breakeven Subs"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))

    scenarios = [
        ("100% Web Monthly (best)", 38.53),
        ("100% Web Annual", 28.30),
        ("Blended (base case)", 34.12),
        ("100% iOS Monthly", 33.99),
        ("100% iOS Annual (worst)", 24.79),
        ("Heavy mobile (40%)", 31.95),
        ("Heavy annual (70%)", 30.24),
    ]
    for i, (scenario, arpu) in enumerate(scenarios):
        r += 1
        ws.cell(row=r, column=1, value=scenario)
        ws.cell(row=r, column=2, value=arpu).number_format = acct_fmt
        ws.cell(row=r, column=3, value=base_fixed).number_format = acct_fmt
        ws.cell(row=r, column=4, value=math.ceil(base_fixed / arpu))
        style_data_row(ws, r, len(headers), i % 2 == 1)

    # ── By Price Point ──
    r += 2
    ws.cell(row=r, column=1, value="BREAKEVEN BY PRICE POINT (Web Monthly Only)").font = section_font
    r += 1
    h2 = ["Monthly Price", "Stripe Net", "Breakeven Subs", "Monthly Rev @ 50", "Annual Rev @ 50"]
    for c, h in enumerate(h2, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(h2))

    for i, price in enumerate([19.99, 24.99, 29.99, 34.99, 39.99, 44.99, 49.99, 59.99, 79.99]):
        r += 1
        snet = price * 0.971 - 0.30
        be = math.ceil(base_fixed / snet)
        ws.cell(row=r, column=1, value=price).number_format = acct_fmt
        ws.cell(row=r, column=2, value=snet).number_format = acct_fmt
        ws.cell(row=r, column=3, value=be)
        ws.cell(row=r, column=4, value=snet * 50).number_format = acct_fmt
        ws.cell(row=r, column=5, value=snet * 50 * 12).number_format = acct_fmt
        style_data_row(ws, r, len(h2), i % 2 == 1)
        if price == 39.99:
            for c in range(1, len(h2) + 1):
                ws.cell(row=r, column=c).fill = blue_fill
                ws.cell(row=r, column=c).font = bold_font

    # ── Churn ──
    r += 2
    ws.cell(row=r, column=1, value="SUBSCRIBER RETENTION IMPACT (Starting: 50 subs)").font = section_font
    r += 1
    h3 = ["Monthly Churn", "Avg Lifetime (mo)", "LTV (Net)", "12-Mo Retained", "12-Mo Cumulative Rev"]
    for c, h in enumerate(h3, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(h3))

    for i, churn in enumerate([0.03, 0.05, 0.07, 0.10, 0.12, 0.15, 0.20]):
        r += 1
        lt = 1 / churn
        ltv = lt * BLENDED_NET_ARPU
        retained = 50 * ((1 - churn) ** 12)
        cum_rev = sum(50 * ((1 - churn) ** m) * BLENDED_NET_ARPU for m in range(12))
        ws.cell(row=r, column=1, value=churn).number_format = pct_fmt
        ws.cell(row=r, column=2, value=round(lt, 1)).number_format = "0.0"
        ws.cell(row=r, column=3, value=round(ltv, 2)).number_format = acct_fmt
        ws.cell(row=r, column=4, value=round(retained)).number_format = int_fmt
        ws.cell(row=r, column=5, value=round(cum_rev)).number_format = acct_fmt
        style_data_row(ws, r, len(h3), i % 2 == 1)

    auto_width(ws, min_width=12, max_width=35)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — 12-Month Projection (uses cost engine)
# ═══════════════════════════════════════════════════════════════════════════════
def build_12mo_projection(wb):
    ws = wb.create_sheet("12-Month Projection")
    ws.sheet_properties.tabColor = ACCENT_GREEN

    ws.merge_cells("A1:N1")
    ws["A1"] = "12-Month Financial Projection (Auto-Scaled Costs)"
    ws["A1"].font = title_font

    ws["A3"] = "GROWTH ASSUMPTIONS"
    ws["A3"].font = section_font
    ws.cell(row=4, column=1, value="Starting subscribers (Month 1)").font = small_font
    ws.cell(row=4, column=2, value=5).font = bold_font
    ws.cell(row=5, column=1, value="Monthly subscriber growth rate").font = small_font
    ws.cell(row=5, column=2, value=0.25).font = bold_font
    ws.cell(row=5, column=2).number_format = pct_fmt
    ws.cell(row=6, column=1, value="Monthly churn rate").font = small_font
    ws.cell(row=6, column=2, value=0.07).font = bold_font
    ws.cell(row=6, column=2).number_format = pct_fmt

    r = 8
    months = ["", "Mo 1", "Mo 2", "Mo 3", "Mo 4", "Mo 5", "Mo 6",
              "Mo 7", "Mo 8", "Mo 9", "Mo 10", "Mo 11", "Mo 12", "TOTAL"]
    for c, h in enumerate(months, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(months))

    start_subs = 5
    growth = 0.25
    churn = 0.07
    prev_subs = 0
    cumulative = 0

    row_labels = [
        "New Subscribers", "Churned Subscribers", "Ending Subscribers",
        "",
        "Gross Revenue", "Net Revenue",
        "",
        "Data API Costs", "Infrastructure", "AI & Variable", "Fixed Overhead",
        "Total Costs",
        "",
        "Net Profit/Loss", "Cumulative P/L", "Operating Margin",
    ]
    row_values = {label: [] for label in row_labels}

    for m in range(12):
        if m == 0:
            new = start_subs
            churned = 0
            ending = start_subs
        else:
            new = math.ceil(prev_subs * growth)
            churned = math.ceil(prev_subs * churn)
            ending = prev_subs + new - churned

        prev_subs = ending
        gross_rev = round(ending * BLENDED_GROSS_PRICE, 2)
        net_rev = round(ending * BLENDED_NET_ARPU, 2)

        costs = compute_all_costs(ending)
        t = costs["_totals"]
        total = t["grand"]
        profit = round(net_rev - total, 2)
        cumulative += profit
        margin = profit / net_rev if net_rev > 0 else 0

        row_values["New Subscribers"].append(new)
        row_values["Churned Subscribers"].append(churned)
        row_values["Ending Subscribers"].append(ending)
        row_values[""].append("")
        row_values["Gross Revenue"].append(gross_rev)
        row_values["Net Revenue"].append(net_rev)
        row_values["Data API Costs"].append(t["api"])
        row_values["Infrastructure"].append(t["infra"])
        row_values["AI & Variable"].append(t["variable"])
        row_values["Fixed Overhead"].append(t["fixed"])
        row_values["Total Costs"].append(round(total, 2))
        row_values["Net Profit/Loss"].append(profit)
        row_values["Cumulative P/L"].append(round(cumulative, 2))
        row_values["Operating Margin"].append(margin)

    separator_labels = {"", " "}
    bold_labels = {"Ending Subscribers", "Net Revenue", "Total Costs",
                   "Net Profit/Loss", "Cumulative P/L"}
    count_labels = {"New Subscribers", "Churned Subscribers", "Ending Subscribers"}
    pct_labels = {"Operating Margin"}

    for label in row_labels:
        r += 1
        ws.cell(row=r, column=1, value=label)

        if label == "":
            continue

        values = row_values[label]
        is_money = label not in count_labels and label not in pct_labels

        for c, v in enumerate(values, 2):
            cell = ws.cell(row=r, column=c, value=v)
            if label in pct_labels:
                cell.number_format = pct_fmt
            elif is_money:
                cell.number_format = acct_fmt_neg if "Profit" in label or "Cumulative" in label else acct_fmt
            else:
                cell.number_format = int_fmt

        # Total column
        if label in count_labels:
            val = values[-1] if label == "Ending Subscribers" else sum(values)
            ws.cell(row=r, column=14, value=val).number_format = int_fmt
        elif label in pct_labels:
            tr = sum(row_values["Net Revenue"])
            tp = sum(row_values["Net Profit/Loss"])
            ws.cell(row=r, column=14, value=tp / tr if tr else 0).number_format = pct_fmt
        elif is_money:
            fmt = acct_fmt_neg if "Profit" in label or "Cumulative" in label else acct_fmt
            ws.cell(row=r, column=14, value=round(sum(values), 2)).number_format = fmt

        if label in bold_labels:
            for c in range(1, 15):
                ws.cell(row=r, column=c).font = bold_font
                ws.cell(row=r, column=c).border = thin_border
            if label == "Net Profit/Loss":
                for c in range(2, 15):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell.value, (int, float)):
                        cell.font = green_font if cell.value >= 0 else red_font
        else:
            for c in range(1, 15):
                ws.cell(row=r, column=c).border = thin_border

    last_row = r

    chart = LineChart()
    chart.title = "Subscriber Growth & Profitability"
    chart.y_axis.title = "Count / $"
    chart.style = 10
    chart.width = 28
    chart.height = 14

    sub_row = None
    profit_row = None
    for check_r in range(9, last_row + 1):
        val = ws.cell(row=check_r, column=1).value
        if val == "Ending Subscribers":
            sub_row = check_r
        elif val == "Net Profit/Loss":
            profit_row = check_r

    if sub_row and profit_row:
        cats = Reference(ws, min_col=2, max_col=13, min_row=8)
        data_subs = Reference(ws, min_col=1, max_col=13, min_row=sub_row)
        data_profit = Reference(ws, min_col=1, max_col=13, min_row=profit_row)
        chart.add_data(data_subs, from_rows=True, titles_from_data=True)
        chart.add_data(data_profit, from_rows=True, titles_from_data=True)
        chart.set_categories(cats)
    ws.add_chart(chart, f"A{last_row + 3}")

    # ── Cost composition chart ──
    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "stacked"
    chart2.title = "Monthly Cost Composition"
    chart2.y_axis.title = "$ / Month"
    chart2.style = 10
    chart2.width = 28
    chart2.height = 14

    api_row = infra_row = var_row = fixed_row = None
    for check_r in range(9, last_row + 1):
        val = ws.cell(row=check_r, column=1).value
        if val == "Data API Costs":
            api_row = check_r
        elif val == "Infrastructure":
            infra_row = check_r
        elif val == "AI & Variable":
            var_row = check_r
        elif val == "Fixed Overhead":
            fixed_row = check_r

    if api_row:
        for src_row in [api_row, infra_row, var_row, fixed_row]:
            if src_row:
                d = Reference(ws, min_col=1, max_col=13, min_row=src_row)
                chart2.add_data(d, from_rows=True, titles_from_data=True)
        chart2.set_categories(cats)
    ws.add_chart(chart2, f"A{last_row + 20}")

    auto_width(ws, min_width=10, max_width=16)
    ws.column_dimensions["A"].width = 22
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — 3-Year Projection (computed from cost engine)
# ═══════════════════════════════════════════════════════════════════════════════
def build_3yr_projection(wb):
    ws = wb.create_sheet("3-Year Projection")
    ws.sheet_properties.tabColor = "8B5CF6"

    ws.merge_cells("A1:E1")
    ws["A1"] = "3-Year Annual Projection (Computed)"
    ws["A1"].font = title_font

    year_profiles = [
        ("Year 1", 5, 50, 0.25, 0.07),
        ("Year 2", 50, 250, 0.15, 0.06),
        ("Year 3", 250, 800, 0.10, 0.05),
    ]

    r = 3
    headers = ["Metric", "Year 1 (5→50)", "Year 2 (50→250)", "Year 3 (250→800)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, len(headers))

    yearly_data = []
    for label, start, end, growth, churn_rate in year_profiles:
        subs = start
        annual_gross = 0
        annual_net = 0
        annual_api = 0
        annual_infra = 0
        annual_var = 0
        annual_fixed = 0

        for m in range(12):
            if m > 0:
                new = math.ceil(subs * growth)
                churned = math.ceil(subs * churn_rate)
                subs = subs + new - churned
            subs = min(subs, end)

            gross = subs * BLENDED_GROSS_PRICE
            net = subs * BLENDED_NET_ARPU
            costs = compute_all_costs(subs)
            t = costs["_totals"]

            annual_gross += gross
            annual_net += net
            annual_api += t["api"]
            annual_infra += t["infra"]
            annual_var += t["variable"]
            annual_fixed += t["fixed"]

        annual_total_cost = annual_api + annual_infra + annual_var + annual_fixed
        annual_profit = annual_net - annual_total_cost
        margin = annual_profit / annual_net if annual_net > 0 else 0

        yearly_data.append({
            "end_subs": subs,
            "gross": round(annual_gross),
            "processing": round(annual_gross - annual_net),
            "net": round(annual_net),
            "api": round(annual_api),
            "infra": round(annual_infra),
            "var": round(annual_var),
            "fixed": round(annual_fixed),
            "total_cost": round(annual_total_cost),
            "profit": round(annual_profit),
            "margin": margin,
            "monthly_profit": round(annual_profit / 12),
        })

    rows = [
        ("Ending Paid Subscribers", [d["end_subs"] for d in yearly_data], int_fmt),
        ("", None, None),
        ("Annual Gross Revenue", [d["gross"] for d in yearly_data], acct_fmt),
        ("Payment Processing Fees", [-d["processing"] for d in yearly_data], acct_fmt_neg),
        ("Annual Net Revenue", [d["net"] for d in yearly_data], acct_fmt),
        ("", None, None),
        ("Data API Costs", [-d["api"] for d in yearly_data], acct_fmt_neg),
        ("Infrastructure", [-d["infra"] for d in yearly_data], acct_fmt_neg),
        ("AI & Variable", [-d["var"] for d in yearly_data], acct_fmt_neg),
        ("Fixed Overhead", [-d["fixed"] for d in yearly_data], acct_fmt_neg),
        ("Total Operating Costs", [-d["total_cost"] for d in yearly_data], acct_fmt_neg),
        ("", None, None),
        ("Annual Operating Profit", [d["profit"] for d in yearly_data], acct_fmt_neg),
        ("Operating Margin", [d["margin"] for d in yearly_data], pct_fmt),
        ("", None, None),
        ("Monthly Avg Profit", [d["monthly_profit"] for d in yearly_data], acct_fmt_neg),
    ]

    key_labels = {"Annual Net Revenue", "Total Operating Costs",
                  "Annual Operating Profit", "Monthly Avg Profit"}

    for i, (label, values, fmt) in enumerate(rows):
        r += 1
        ws.cell(row=r, column=1, value=label)
        if values is None:
            continue
        for c, v in enumerate(values, 2):
            ws.cell(row=r, column=c, value=v).number_format = fmt

        if label in key_labels:
            style_total_row(ws, r, len(headers))
        else:
            style_data_row(ws, r, len(headers), i % 2 == 1)

        if label == "Annual Operating Profit":
            for c in range(2, len(headers) + 1):
                ws.cell(row=r, column=c).font = green_font

    auto_width(ws, min_width=14, max_width=28)
    ws.column_dimensions["A"].width = 28
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 9 — Risks & Optimization
# ═══════════════════════════════════════════════════════════════════════════════
def build_risks(wb):
    ws = wb.create_sheet("Risks & Optimization")
    ws.sheet_properties.tabColor = ACCENT_RED

    ws.merge_cells("A1:D1")
    ws["A1"] = "Key Risks & Cost Optimization"
    ws["A1"].font = title_font

    ws["A3"] = "KEY RISKS"
    ws["A3"].font = section_font
    r = 4
    h = ["Risk", "Impact", "Probability", "Mitigation"]
    for c, val in enumerate(h, 1):
        ws.cell(row=r, column=c, value=val)
    style_header_row(ws, r, len(h))

    risks = [
        ("RentCast price increase", "High — 51% of base costs", "Medium", "Optimize caching; negotiate volume"),
        ("AXESSO/Zillow API discontinued", "Critical — lose Zestimate", "Low", "Build fallback to direct Zillow API"),
        ("RapidAPI plan changes", "Medium — lose Redfin/Realtor", "Low", "Evaluate direct API access"),
        ("Apple rejects SBP (15%→30%)", "Medium — mobile margin drops", "Low", "Web-first acquisition strategy"),
        ("Low annual conversion rate", "Medium — lower blended ARPU", "Medium", "Incentivize annual w/ deeper discount"),
        ("High churn (>10%/mo)", "High — never reach scale", "Medium", "Focus on retention, feature value"),
        ("Free tier API abuse", "Medium — consumes paid quota", "Medium", "Enforce rate limits; reduce free cap"),
        ("Competitor with free tier", "High — pricing pressure", "Medium", "Differentiate on IQ Estimate quality"),
    ]
    for i, row_data in enumerate(risks):
        r += 1
        for c, v in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=v)
        style_data_row(ws, r, len(h), i % 2 == 1)
        impact_cell = ws.cell(row=r, column=2)
        if "Critical" in str(impact_cell.value) or "High" in str(impact_cell.value):
            impact_cell.fill = red_fill
        elif "Medium" in str(impact_cell.value):
            impact_cell.fill = yellow_fill

    r += 2
    ws.cell(row=r, column=1, value="COST OPTIMIZATION LEVERS").font = section_font
    r += 1
    h2 = ["Optimization", "Est. Savings", "Effort", "Priority"]
    for c, val in enumerate(h2, 1):
        ws.cell(row=r, column=c, value=val)
    style_header_row(ws, r, len(h2))

    opts = [
        ("Extend cache TTL 24h → 48-72h", "30–50% fewer API calls", "Low", "P0 — Immediate"),
        ("Lazy-load Redfin/Realtor on drill-down", "~40% fewer RapidAPI calls", "Medium", "P1 — Next sprint"),
        ("Annual billing incentives", "Better cash flow, lower churn", "Low", "P1"),
        ("Batch RentCast calls where possible", "~10% fewer RentCast calls", "Medium", "P2"),
        ("Negotiate enterprise pricing @ 500 subs", "10–30% cost reduction", "Low", "P2 — When applicable"),
        ("Reduce free tier from 3 to 2 analyses", "~33% fewer free API calls", "Low", "P3 — If needed"),
    ]
    for i, row_data in enumerate(opts):
        r += 1
        for c, v in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=v)
        style_data_row(ws, r, len(h2), i % 2 == 1)
        pc = ws.cell(row=r, column=4)
        if "P0" in str(pc.value):
            pc.fill = red_fill
        elif "P1" in str(pc.value):
            pc.fill = yellow_fill
        elif "P2" in str(pc.value):
            pc.fill = green_fill

    auto_width(ws, min_width=14, max_width=40)
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    build_dashboard(wb)
    build_fixed_costs(wb)
    build_profitability(wb)
    build_scaling_detail(wb)
    build_api_capacity(wb)
    build_sensitivity(wb)
    build_12mo_projection(wb)
    build_3yr_projection(wb)
    build_risks(wb)

    output = "/Users/bradgeisen/IQ-Data/dealscope/docs/DealGapIQ_Financial_Proforma.xlsx"
    wb.save(output)
    print(f"Proforma saved to: {output}")
    print(f"Sheets: {wb.sheetnames}")

    # Print a quick cost engine verification
    print("\n── Cost Engine Verification ──")
    for subs in [0, 12, 50, 100, 200, 500, 1000, 3000, 5000]:
        c = compute_all_costs(subs)
        net = subs * BLENDED_NET_ARPU
        profit = net - c["_totals"]["grand"]
        plans = []
        for p in API_PROVIDERS:
            plans.append(f"{p}: {c['api'][p][0]}")
        for s in INFRA_SERVICES:
            plans.append(f"{s}: {c['infra'][s][0]}")
        print(f"  {subs:>5} subs → ${c['_totals']['grand']:>9,.2f} total cost"
              f"  |  profit ${profit:>10,.2f}  |  {', '.join(plans)}")


if __name__ == "__main__":
    main()
