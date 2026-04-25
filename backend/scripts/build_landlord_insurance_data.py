#!/usr/bin/env python3
"""
Build JSON artifacts for county-level landlord insurance lookup from
data/DealGapIQ_Landlord_Insurance_Dataset.xlsx (ACS-based).

Output:
  backend/app/data/landlord_insurance/county_rates.json
  backend/app/data/landlord_insurance/state_calibration.json

Usage (from repo root or backend/):
  backend/venv/bin/python backend/scripts/build_landlord_insurance_data.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise e

# U.S. states + DC + PR — two-letter keys match Calibration tab when mapped.
_STATE_FULL_TO_ABBREV: dict[str, str] = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "District of Columbia": "DC",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Puerto Rico": "PR",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
}


def _strip_county_type_prefix(name: str) -> str:
    """Baldwin County -> Baldwin; Juneau City and Borough -> Juneau; Anchorage Municipality -> Anchorage."""
    n = name.strip()
    for suf in (
        " City and Borough",
        " Municipality",
        " County",
        " Parish",
        " Borough",
        " Census Area",
        " Municipio",  # PR
    ):
        if n.endswith(suf):
            n = n[: -len(suf)].strip()
            break
    return n


def _parse_county_label(county_and_state: str, state_full: str) -> str:
    """
    'Palm Beach County, Florida' with state column 'Florida' -> 'Palm Beach County' -> 'Palm Beach'
    'Orleans Parish, Louisiana' -> 'Orleans'
    """
    s = county_and_state.strip()
    if "," in s:
        first = s.rsplit(", ", 1)[0].strip()
    else:
        first = s
    return _strip_county_type_prefix(first)


def _read_calibration(wb) -> dict[str, float]:
    cal = {}
    ws = wb["Calibration"]
    rows = list(ws.iter_rows(values_only=True))
    in_table = False
    for r in rows:
        if r and r[0] == "State" and r[1] == "Multiplier":
            in_table = True
            continue
        if not in_table or not r or r[0] is None:
            continue
        st = str(r[0]).strip()
        if st == "State":
            continue
        mult = r[1]
        if mult is None:
            continue
        ab = _STATE_FULL_TO_ABBREV.get(st)
        if not ab and len(st) == 2 and st.isalpha():
            ab = st.upper()
        if not ab:
            print(f"Warning: skip calibration row for unknown state: {st!r}", file=sys.stderr)
            continue
        try:
            cal[ab] = float(mult)
        except (TypeError, ValueError):
            continue
    return cal


def main() -> int:
    root = Path(__file__).resolve().parents[1]  # backend/
    xlsx = root.parent / "data" / "DealGapIQ_Landlord_Insurance_Dataset.xlsx"
    if not xlsx.is_file():
        print(f"Workbook not found: {xlsx}", file=sys.stderr)
        return 1

    out_dir = root / "app" / "data" / "landlord_insurance"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["County_Data"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    def get(row: tuple, name: str) -> object:
        j = idx.get(name)
        if j is None or j >= len(row):
            return None
        return row[j]

    county_records: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or get(row, "GEOID") is None:
            continue
        geoid = str(get(row, "GEOID")).strip()
        county_raw = str(get(row, "County") or "")
        state_full = str(get(row, "State") or "").strip()
        st = _STATE_FULL_TO_ABBREV.get(state_full)
        if not st:
            print(f"Unknown state: {state_full!r} GEOID {geoid}", file=sys.stderr)
            continue
        county_name = _parse_county_label(county_raw, state_full)
        ratio = float(get(row, "Landlord Insurance % of Value"))
        med_raw = get(row, "Median Home Value")
        med = int(med_raw) if med_raw is not None else None
        ho_raw = get(row, "Avg HO Insurance ($/yr)")
        ho = float(ho_raw) if ho_raw is not None else None

        key = (st, county_name.lower())
        if key in seen:
            print(f"Duplicate county key: {st} {county_name!r} — skipping", file=sys.stderr)
            continue
        seen.add(key)

        county_records.append(
            {
                "geoid": geoid,
                "county": county_name,
                "state": st,
                "state_full": state_full,
                "ratio": ratio,
                "median_home_value": med,
                "avg_ho_insurance_annual": ho,
            }
        )

    cal = _read_calibration(wb)
    if not cal:
        print("No calibration multipliers read", file=sys.stderr)
        return 1

    (out_dir / "county_rates.json").write_text(
        json.dumps(county_records, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )
    (out_dir / "state_calibration.json").write_text(
        json.dumps(cal, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    print(f"Wrote {len(county_records)} county rows and {len(cal)} state calibration multipliers to {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
