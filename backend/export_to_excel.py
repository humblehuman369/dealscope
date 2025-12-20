#!/usr/bin/env python3
"""
Export comprehensive property data from RentCast and Zillow to Excel (.xlsx)

Creates a multi-sheet Excel workbook with:
- Summary sheet with key metrics
- RentCast property data
- Zillow property data
- Valuations comparison
- Rental data
- Tax history
- Comparables
- Market statistics
"""
import asyncio
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.services.unified_property_service import create_unified_service
from app.services.zillow_client import create_zillow_client


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
CURRENCY_FORMAT = '$#,##0'
PERCENT_FORMAT = '0.00%'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def style_header_row(ws, row, cols):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_section_row(ws, row, cols):
    """Apply section styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER


def auto_width(ws):
    """Auto-adjust column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width


def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary."""
    items = []
    if not isinstance(d, dict):
        return [(parent_key, d)]
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            if len(v) > 0 and isinstance(v[0], dict):
                # Skip complex nested lists for now
                items.append((new_key + "_count", len(v)))
            else:
                items.append((new_key, str(v)[:500]))
        else:
            items.append((new_key, v))
    return dict(items)


async def fetch_all_data(address: str):
    """Fetch all data from both APIs."""
    rentcast_key = os.getenv("RENTCAST_API_KEY", "")
    rentcast_url = os.getenv("RENTCAST_URL", "https://api.rentcast.io/v1")
    axesso_key = os.getenv("AXESSO_API_KEY", "")
    axesso_url = os.getenv("AXESSO_URL", "https://api.axesso.de/zil")
    
    service = create_unified_service(
        rentcast_api_key=rentcast_key,
        rentcast_url=rentcast_url,
        axesso_api_key=axesso_key,
        axesso_url=axesso_url
    )
    
    # Get unified data
    result = await service.get_property(address)
    
    # Also get direct Zillow data for more detail
    zillow_client = create_zillow_client(axesso_key, axesso_url)
    zillow_full = await zillow_client.get_complete_property_data(address)
    
    return result, zillow_full


def create_summary_sheet(wb, normalized, metrics, quality, address):
    """Create summary sheet."""
    ws = wb.active
    ws.title = "Summary"
    
    # Title
    ws['A1'] = "INVESTIQ PROPERTY ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Property: {address}"
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"Data Quality Score: {quality.get('score', 0)}%"
    
    # Key Metrics
    row = 6
    ws.cell(row=row, column=1, value="KEY METRICS")
    style_section_row(ws, row, 4)
    row += 1
    
    ws.cell(row=row, column=1, value="Field")
    ws.cell(row=row, column=2, value="Value")
    ws.cell(row=row, column=3, value="Source")
    ws.cell(row=row, column=4, value="Notes")
    style_header_row(ws, row, 4)
    row += 1
    
    key_data = [
        ("Address", normalized.get('formatted_address'), "rentcast", ""),
        ("Property Type", normalized.get('property_type'), "rentcast", ""),
        ("Bedrooms", normalized.get('bedrooms'), "rentcast", ""),
        ("Bathrooms", normalized.get('bathrooms'), "rentcast", ""),
        ("Square Footage", normalized.get('square_footage'), "rentcast", ""),
        ("Year Built", normalized.get('year_built'), "rentcast", ""),
        ("", "", "", ""),
        ("Current AVM Estimate", normalized.get('current_value_avm'), normalized.get('provenance', {}).get('current_value_avm', {}).get('source', ''), "Best estimate"),
        ("RentCast AVM", normalized.get('rentcast_avm'), "rentcast", ""),
        ("Zillow Zestimate", normalized.get('zestimate'), "zillow", ""),
        ("", "", "", ""),
        ("Monthly Rent Estimate", normalized.get('monthly_rent_estimate'), normalized.get('provenance', {}).get('monthly_rent_estimate', {}).get('source', ''), "Best estimate"),
        ("RentCast Rent", normalized.get('rentcast_rent'), "rentcast", ""),
        ("Zillow Rent Zestimate", normalized.get('rent_zestimate'), "zillow", ""),
        ("", "", "", ""),
        ("Annual Property Tax", normalized.get('annual_property_tax'), "rentcast", ""),
        ("Tax Assessed Value", normalized.get('tax_assessed_value'), "rentcast", ""),
        ("Last Sale Price", normalized.get('last_sale_price'), "rentcast", ""),
        ("Last Sale Date", str(normalized.get('last_sale_date', '')), "rentcast", ""),
        ("", "", "", ""),
        ("Walk Score", normalized.get('walk_score'), "zillow", ""),
        ("Transit Score", normalized.get('transit_score'), "zillow", ""),
        ("Bike Score", normalized.get('bike_score'), "zillow", ""),
        ("School Rating Avg", normalized.get('school_rating_avg'), "zillow", ""),
    ]
    
    for field, value, source, notes in key_data:
        ws.cell(row=row, column=1, value=field)
        if isinstance(value, (int, float)) and value and value > 1000:
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        else:
            ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=source)
        ws.cell(row=row, column=4, value=notes)
        row += 1
    
    # Investment Metrics
    row += 1
    ws.cell(row=row, column=1, value="INVESTMENT METRICS")
    style_section_row(ws, row, 4)
    row += 1
    
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    ws.cell(row=row, column=3, value="Notes")
    style_header_row(ws, row, 3)
    row += 1
    
    for key, value in metrics.items():
        if isinstance(value, dict):
            continue
        ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
        if isinstance(value, float):
            if 'pct' in key or 'roi' in key or 'rate' in key:
                ws.cell(row=row, column=2, value=value/100 if value < 1 else value/100)
                ws.cell(row=row, column=2).number_format = PERCENT_FORMAT
            elif value > 100:
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
            else:
                ws.cell(row=row, column=2, value=round(value, 2))
        else:
            ws.cell(row=row, column=2, value=value)
        row += 1
    
    auto_width(ws)


def create_rentcast_sheet(wb, rentcast_data):
    """Create RentCast raw data sheet."""
    ws = wb.create_sheet("RentCast Data")
    
    ws['A1'] = "RENTCAST API RESPONSE DATA"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    if not rentcast_data:
        ws.cell(row=row, column=1, value="No RentCast data available")
        return
    
    endpoints = rentcast_data.get('endpoints', {})
    
    for endpoint_name, endpoint_data in endpoints.items():
        ws.cell(row=row, column=1, value=f"ENDPOINT: {endpoint_name.upper()}")
        style_section_row(ws, row, 3)
        row += 1
        
        status = endpoint_data.get('status_code', 'N/A')
        ws.cell(row=row, column=1, value=f"Status Code: {status}")
        row += 1
        
        data = endpoint_data.get('data')
        if data:
            if isinstance(data, list) and len(data) > 0:
                data = data[0]  # Take first item
            
            if isinstance(data, dict):
                flat = flatten_dict(data)
                
                ws.cell(row=row, column=1, value="Field")
                ws.cell(row=row, column=2, value="Value")
                style_header_row(ws, row, 2)
                row += 1
                
                for key, value in flat.items():
                    ws.cell(row=row, column=1, value=key)
                    if isinstance(value, (int, float)) and value and abs(value) > 1000:
                        ws.cell(row=row, column=2, value=value)
                        ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
                    else:
                        ws.cell(row=row, column=2, value=str(value)[:500] if value else "")
                    row += 1
        
        row += 1
    
    auto_width(ws)


def create_zillow_sheet(wb, zillow_data):
    """Create Zillow raw data sheet."""
    ws = wb.create_sheet("Zillow Data")
    
    ws['A1'] = "ZILLOW (AXESSO) API RESPONSE DATA"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    if not zillow_data:
        ws.cell(row=row, column=1, value="No Zillow data available")
        return
    
    for endpoint_name, endpoint_response in zillow_data.items():
        if endpoint_name == 'zpid':
            ws.cell(row=row, column=1, value=f"ZPID: {endpoint_response}")
            row += 2
            continue
        
        ws.cell(row=row, column=1, value=f"ENDPOINT: {endpoint_name.upper()}")
        style_section_row(ws, row, 3)
        row += 1
        
        # Handle ZillowAPIResponse objects
        if hasattr(endpoint_response, 'success'):
            ws.cell(row=row, column=1, value=f"Success: {endpoint_response.success}")
            row += 1
            if endpoint_response.error:
                ws.cell(row=row, column=1, value=f"Error: {endpoint_response.error}")
                row += 1
            
            data = endpoint_response.data
        else:
            data = endpoint_response
        
        if data and isinstance(data, dict):
            flat = flatten_dict(data)
            
            ws.cell(row=row, column=1, value="Field")
            ws.cell(row=row, column=2, value="Value")
            style_header_row(ws, row, 2)
            row += 1
            
            for key, value in flat.items():
                ws.cell(row=row, column=1, value=key)
                if isinstance(value, (int, float)) and value and abs(value) > 1000:
                    ws.cell(row=row, column=2, value=value)
                    ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
                else:
                    ws.cell(row=row, column=2, value=str(value)[:500] if value else "")
                row += 1
        
        row += 1
    
    auto_width(ws)


def create_comparables_sheet(wb, rentcast_data, zillow_data):
    """Create comparables sheet."""
    ws = wb.create_sheet("Comparables")
    
    ws['A1'] = "COMPARABLE PROPERTIES"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    # RentCast Sale Comps
    avm_data = rentcast_data.get('endpoints', {}).get('avm_value', {}).get('data', {})
    if avm_data:
        comps = avm_data.get('comparables', [])
        if comps:
            ws.cell(row=row, column=1, value="RENTCAST SALE COMPARABLES")
            style_section_row(ws, row, 8)
            row += 1
            
            headers = ["Address", "Price", "Beds", "Baths", "Sq Ft", "Year Built", "Distance (mi)", "Correlation"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            style_header_row(ws, row, len(headers))
            row += 1
            
            for comp in comps[:15]:  # Limit to 15
                ws.cell(row=row, column=1, value=comp.get('formattedAddress', ''))
                ws.cell(row=row, column=2, value=comp.get('price', 0))
                ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
                ws.cell(row=row, column=3, value=comp.get('bedrooms', ''))
                ws.cell(row=row, column=4, value=comp.get('bathrooms', ''))
                ws.cell(row=row, column=5, value=comp.get('squareFootage', ''))
                ws.cell(row=row, column=6, value=comp.get('yearBuilt', ''))
                ws.cell(row=row, column=7, value=round(comp.get('distance', 0), 2))
                ws.cell(row=row, column=8, value=round(comp.get('correlation', 0), 4))
                row += 1
            
            row += 1
    
    # RentCast Rental Comps
    rent_data = rentcast_data.get('endpoints', {}).get('rent_estimate', {}).get('data', {})
    if rent_data:
        comps = rent_data.get('comparables', [])
        if comps:
            ws.cell(row=row, column=1, value="RENTCAST RENTAL COMPARABLES")
            style_section_row(ws, row, 8)
            row += 1
            
            headers = ["Address", "Rent", "Beds", "Baths", "Sq Ft", "Year Built", "Distance (mi)", "Correlation"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            style_header_row(ws, row, len(headers))
            row += 1
            
            for comp in comps[:15]:
                ws.cell(row=row, column=1, value=comp.get('formattedAddress', ''))
                ws.cell(row=row, column=2, value=comp.get('price', 0))
                ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
                ws.cell(row=row, column=3, value=comp.get('bedrooms', ''))
                ws.cell(row=row, column=4, value=comp.get('bathrooms', ''))
                ws.cell(row=row, column=5, value=comp.get('squareFootage', ''))
                ws.cell(row=row, column=6, value=comp.get('yearBuilt', ''))
                ws.cell(row=row, column=7, value=round(comp.get('distance', 0), 2))
                ws.cell(row=row, column=8, value=round(comp.get('correlation', 0), 4))
                row += 1
            
            row += 1
    
    # Zillow Similar Properties
    if zillow_data:
        for comp_type in ['similar_for_sale', 'similar_sold', 'similar_rentals']:
            comp_data = zillow_data.get(comp_type)
            if comp_data and hasattr(comp_data, 'data') and comp_data.data:
                props = comp_data.data if isinstance(comp_data.data, list) else [comp_data.data]
                
                ws.cell(row=row, column=1, value=f"ZILLOW {comp_type.upper().replace('_', ' ')}")
                style_section_row(ws, row, 6)
                row += 1
                
                headers = ["Address", "Price", "Beds", "Baths", "Sq Ft", "Status"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header)
                style_header_row(ws, row, len(headers))
                row += 1
                
                for prop in props[:10]:
                    if isinstance(prop, dict):
                        ws.cell(row=row, column=1, value=prop.get('address', ''))
                        ws.cell(row=row, column=2, value=prop.get('price', 0))
                        ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
                        ws.cell(row=row, column=3, value=prop.get('bedrooms', ''))
                        ws.cell(row=row, column=4, value=prop.get('bathrooms', ''))
                        ws.cell(row=row, column=5, value=prop.get('livingArea', ''))
                        ws.cell(row=row, column=6, value=prop.get('homeStatus', ''))
                        row += 1
                
                row += 1
    
    auto_width(ws)


def create_market_sheet(wb, rentcast_data):
    """Create market statistics sheet."""
    ws = wb.create_sheet("Market Statistics")
    
    ws['A1'] = "MARKET STATISTICS"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    market_data = rentcast_data.get('endpoints', {}).get('market_stats', {}).get('data', {})
    if not market_data:
        ws.cell(row=row, column=1, value="No market data available")
        return
    
    sale_data = market_data.get('saleData', {})
    
    # General Stats
    ws.cell(row=row, column=1, value="SALE STATISTICS")
    style_section_row(ws, row, 3)
    row += 1
    
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    style_header_row(ws, row, 2)
    row += 1
    
    stats = [
        ("Average Price", sale_data.get('averagePrice')),
        ("Median Price", sale_data.get('medianPrice')),
        ("Min Price", sale_data.get('minPrice')),
        ("Max Price", sale_data.get('maxPrice')),
        ("Avg Price/Sq Ft", sale_data.get('averagePricePerSquareFoot')),
        ("Median Price/Sq Ft", sale_data.get('medianPricePerSquareFoot')),
        ("Avg Days on Market", sale_data.get('averageDaysOnMarket')),
        ("Median Days on Market", sale_data.get('medianDaysOnMarket')),
        ("New Listings", sale_data.get('newListings')),
        ("Total Listings", sale_data.get('totalListings')),
    ]
    
    for stat_name, value in stats:
        ws.cell(row=row, column=1, value=stat_name)
        if isinstance(value, (int, float)) and value and value > 100:
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        else:
            ws.cell(row=row, column=2, value=value)
        row += 1
    
    row += 1
    
    # By Property Type
    by_type = sale_data.get('dataByPropertyType', [])
    if by_type:
        ws.cell(row=row, column=1, value="BY PROPERTY TYPE")
        style_section_row(ws, row, 6)
        row += 1
        
        headers = ["Property Type", "Avg Price", "Median Price", "Avg $/SqFt", "Avg DOM", "Total Listings"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        style_header_row(ws, row, len(headers))
        row += 1
        
        for item in by_type:
            ws.cell(row=row, column=1, value=item.get('propertyType', ''))
            ws.cell(row=row, column=2, value=item.get('averagePrice', 0))
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
            ws.cell(row=row, column=3, value=item.get('medianPrice', 0))
            ws.cell(row=row, column=3).number_format = CURRENCY_FORMAT
            ws.cell(row=row, column=4, value=item.get('averagePricePerSquareFoot', 0))
            ws.cell(row=row, column=5, value=item.get('averageDaysOnMarket', 0))
            ws.cell(row=row, column=6, value=item.get('totalListings', 0))
            row += 1
        
        row += 1
    
    # By Bedrooms
    by_beds = sale_data.get('dataByBedrooms', [])
    if by_beds:
        ws.cell(row=row, column=1, value="BY BEDROOMS")
        style_section_row(ws, row, 6)
        row += 1
        
        headers = ["Bedrooms", "Avg Price", "Median Price", "Avg $/SqFt", "Avg DOM", "Total Listings"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        style_header_row(ws, row, len(headers))
        row += 1
        
        for item in by_beds:
            ws.cell(row=row, column=1, value=item.get('bedrooms', ''))
            ws.cell(row=row, column=2, value=item.get('averagePrice', 0))
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
            ws.cell(row=row, column=3, value=item.get('medianPrice', 0))
            ws.cell(row=row, column=3).number_format = CURRENCY_FORMAT
            ws.cell(row=row, column=4, value=item.get('averagePricePerSquareFoot', 0))
            ws.cell(row=row, column=5, value=item.get('averageDaysOnMarket', 0))
            ws.cell(row=row, column=6, value=item.get('totalListings', 0))
            row += 1
    
    auto_width(ws)


def create_tax_history_sheet(wb, rentcast_data, zillow_data):
    """Create tax history sheet."""
    ws = wb.create_sheet("Tax History")
    
    ws['A1'] = "TAX ASSESSMENT HISTORY"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    # RentCast Tax Data
    props = rentcast_data.get('endpoints', {}).get('properties', {}).get('data', [])
    if props and len(props) > 0:
        prop = props[0]
        
        tax_assessments = prop.get('taxAssessments', {})
        property_taxes = prop.get('propertyTaxes', {})
        
        if tax_assessments:
            ws.cell(row=row, column=1, value="TAX ASSESSMENTS (RentCast)")
            style_section_row(ws, row, 4)
            row += 1
            
            headers = ["Year", "Assessed Value", "Land Value", "Improvements"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            style_header_row(ws, row, len(headers))
            row += 1
            
            for year in sorted(tax_assessments.keys(), reverse=True):
                data = tax_assessments[year]
                ws.cell(row=row, column=1, value=year)
                ws.cell(row=row, column=2, value=data.get('value', 0))
                ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
                ws.cell(row=row, column=3, value=data.get('land', 0))
                ws.cell(row=row, column=3).number_format = CURRENCY_FORMAT
                ws.cell(row=row, column=4, value=data.get('improvements', 0))
                ws.cell(row=row, column=4).number_format = CURRENCY_FORMAT
                row += 1
            
            row += 1
        
        if property_taxes:
            ws.cell(row=row, column=1, value="PROPERTY TAXES (RentCast)")
            style_section_row(ws, row, 2)
            row += 1
            
            headers = ["Year", "Tax Amount"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
            style_header_row(ws, row, len(headers))
            row += 1
            
            for year in sorted(property_taxes.keys(), reverse=True):
                data = property_taxes[year]
                ws.cell(row=row, column=1, value=year)
                ws.cell(row=row, column=2, value=data.get('total', 0))
                ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
                row += 1
            
            row += 1
    
    # Zillow Tax History
    if zillow_data:
        tax_hist = zillow_data.get('price_tax_history')
        if tax_hist and hasattr(tax_hist, 'data') and tax_hist.data:
            tax_data = tax_hist.data.get('taxHistory', [])
            if tax_data:
                ws.cell(row=row, column=1, value="TAX HISTORY (Zillow)")
                style_section_row(ws, row, 3)
                row += 1
                
                headers = ["Year", "Tax Paid", "Tax Assessment"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header)
                style_header_row(ws, row, len(headers))
                row += 1
                
                for item in tax_data:
                    ws.cell(row=row, column=1, value=item.get('time', ''))
                    ws.cell(row=row, column=2, value=item.get('taxPaid', 0))
                    ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
                    ws.cell(row=row, column=3, value=item.get('value', 0))
                    ws.cell(row=row, column=3).number_format = CURRENCY_FORMAT
                    row += 1
    
    auto_width(ws)


def create_data_provenance_sheet(wb, normalized):
    """Create data provenance sheet."""
    ws = wb.create_sheet("Data Provenance")
    
    ws['A1'] = "DATA PROVENANCE & SOURCES"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    ws.cell(row=row, column=1, value="Field")
    ws.cell(row=row, column=2, value="Value")
    ws.cell(row=row, column=3, value="Source")
    ws.cell(row=row, column=4, value="Confidence")
    ws.cell(row=row, column=5, value="Conflict")
    style_header_row(ws, row, 5)
    row += 1
    
    provenance = normalized.get('provenance', {})
    for field_name, prov in sorted(provenance.items()):
        value = normalized.get(field_name)
        ws.cell(row=row, column=1, value=field_name)
        if isinstance(value, (int, float)) and value and abs(value) > 1000:
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        else:
            ws.cell(row=row, column=2, value=str(value)[:100] if value else "")
        ws.cell(row=row, column=3, value=prov.get('source', ''))
        ws.cell(row=row, column=4, value=prov.get('confidence', ''))
        ws.cell(row=row, column=5, value="Yes" if prov.get('conflict') else "No")
        row += 1
    
    auto_width(ws)


async def main():
    """Main export function."""
    address = "953 Banyan Dr, Delray Beach, FL 33483"
    
    print("=" * 60)
    print("EXPORTING PROPERTY DATA TO EXCEL")
    print("=" * 60)
    print(f"Property: {address}")
    print()
    
    print("⏳ Fetching data from RentCast and Zillow...")
    result, zillow_full = await fetch_all_data(address)
    print("✅ Data fetched")
    
    normalized = result['normalized']
    metrics = result['investment_metrics']
    quality = result['data_quality']
    rentcast_data = result['raw']['rentcast']
    
    print("📊 Creating Excel workbook...")
    
    wb = Workbook()
    
    # Create all sheets
    create_summary_sheet(wb, normalized, metrics, quality, address)
    create_rentcast_sheet(wb, rentcast_data)
    create_zillow_sheet(wb, zillow_full)
    create_comparables_sheet(wb, rentcast_data, zillow_full)
    create_market_sheet(wb, rentcast_data)
    create_tax_history_sheet(wb, rentcast_data, zillow_full)
    create_data_provenance_sheet(wb, normalized)
    
    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"InvestIQ_953_Banyan_Dr_{timestamp}.xlsx"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    
    wb.save(filepath)
    
    print(f"✅ Excel file saved: {filepath}")
    print()
    print("Sheets created:")
    for sheet in wb.sheetnames:
        print(f"  • {sheet}")
    
    return filepath


if __name__ == "__main__":
    asyncio.run(main())

