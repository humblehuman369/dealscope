"""
Comprehensive Excel export: proforma tabs + financial statements + 6 strategy worksheets.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from app.schemas.proforma import FinancialProforma
from app.services.proforma_exporter import ProformaExcelExporter
from app.services.report_service import ReportService, report_service

STRATEGY_ORDER = ["ltr", "str", "brrrr", "flip", "house_hack", "wholesale"]

STRATEGY_SHEET_TITLES = {
    "ltr": "Long-Term Rental",
    "str": "Short-Term Rental",
    "brrrr": "BRRRR",
    "flip": "Fix & Flip",
    "house_hack": "House Hack",
    "wholesale": "Wholesale",
}


class ComprehensiveExcelExporter:
    """Build a single workbook for Strategy page comprehensive export."""

    def __init__(self, report_svc: ReportService | None = None):
        self.rs = report_svc or report_service

    def generate(
        self,
        property_data: dict[str, Any],
        analytics_data: dict[str, Any],
        proforma: FinancialProforma,
        assumptions: dict[str, Any] | None,
        active_strategy: str = "ltr",
        user_override_keys: set[str] | None = None,
        include_sensitivity: bool = True,
    ) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        self.rs._create_summary_sheet(wb, property_data, analytics_data, active_strategy)

        ProformaExcelExporter(proforma).add_tabs_to_workbook(wb)
        if "Assumptions" in wb.sheetnames:
            wb["Assumptions"].title = "Proforma Assumptions"

        self.rs._create_cash_flow_statement_sheet(wb, property_data, analytics_data, assumptions)
        self.rs._create_dscr_analysis_sheet(wb, property_data, analytics_data, assumptions)
        self.rs._create_pro_forma_sheet(wb, property_data, analytics_data, assumptions)
        self.rs._create_amortization_sheet(wb, property_data, analytics_data, assumptions)
        if include_sensitivity:
            self.rs._create_sensitivity_sheet(wb, property_data, analytics_data, assumptions)

        strategy_order = [active_strategy] + [s for s in STRATEGY_ORDER if s != active_strategy]
        for key in strategy_order:
            data = analytics_data.get(key) or analytics_data.get("str_results" if key == "str" else key)
            if key == "str" and not data:
                data = analytics_data.get("str")
            if data:
                self._create_detailed_strategy_worksheet(wb, key, data)

        if assumptions:
            self.rs._create_assumptions_sheet(wb, assumptions)
            if user_override_keys:
                self._append_user_overrides_note(wb, user_override_keys)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _append_user_overrides_note(self, wb: Workbook, keys: set[str]) -> None:
        sheet_name = "Assumptions" if "Assumptions" in wb.sheetnames else None
        if sheet_name is None:
            return
        ws = wb[sheet_name]
        row = ws.max_row + 2
        ws.cell(row=row, column=1, value="User Adjustments Applied").font = Font(bold=True, size=12)
        row += 1
        for key in sorted(keys):
            ws.cell(row=row, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row, column=2, value="From Strategy worksheet")
            row += 1

    def _create_detailed_strategy_worksheet(self, wb: Workbook, strategy_key: str, data: dict[str, Any]) -> None:
        title = STRATEGY_SHEET_TITLES.get(strategy_key, strategy_key)
        color = self.rs.STRATEGY_COLORS.get(strategy_key, self.rs.BRAND_BLUE)
        ws = wb.create_sheet(title[:31])

        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, size=14, color="FFFFFF")
        section_font = Font(bold=True, size=11, color=self.rs.NAVY)
        section_fill = PatternFill(start_color=self.rs.LIGHT_GRAY, end_color=self.rs.LIGHT_GRAY, fill_type="solid")

        ws.merge_cells("A1:C1")
        ws["A1"] = f"{title} — Financial Worksheet"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws["A2"].font = Font(italic=True, size=10, color="666666")

        sections: list[tuple[str, list[tuple[str, str, str]]]] = self._sections_for_strategy(strategy_key, data)

        row = 4
        for section_name, rows in sections:
            ws.cell(row=row, column=1, value=section_name).font = section_font
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = section_fill
            row += 1
            for label, value, fmt in rows:
                ws.cell(row=row, column=1, value=label)
                cell = ws.cell(row=row, column=2, value=value)
                if fmt == "currency" and isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0'
                elif fmt == "percent" and isinstance(value, (int, float)):
                    cell.number_format = "0.00%"
                    if value > 1:
                        cell.value = value / 100
                elif fmt == "number" and isinstance(value, (int, float)):
                    cell.number_format = "0.00"
                row += 1
            row += 1

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 14

    def _sections_for_strategy(
        self, key: str, d: dict[str, Any]
    ) -> list[tuple[str, list[tuple[str, str, str]]]]:
        """Return (section_title, [(label, value, format)]) per strategy."""

        def cur(label: str, field: str) -> tuple[str, str, str]:
            return (label, d.get(field, 0), "currency")

        def pct(label: str, field: str) -> tuple[str, str, str]:
            return (label, d.get(field, 0), "percent")

        def num(label: str, field: str) -> tuple[str, str, str]:
            return (label, d.get(field, 0), "number")

        def txt(label: str, field: str) -> tuple[str, str, str]:
            return (label, str(d.get(field, "—")), "text")

        if key == "ltr":
            return [
                (
                    "INCOME",
                    [
                        cur("Monthly Rent", "monthly_rent"),
                        cur("Annual Gross Rent", "annual_gross_rent"),
                        cur("Vacancy Loss", "vacancy_loss"),
                        cur("Effective Gross Income", "effective_gross_income"),
                    ],
                ),
                (
                    "OPERATING EXPENSES",
                    [
                        cur("Property Taxes", "property_taxes"),
                        cur("Insurance", "insurance"),
                        cur("Property Management", "property_management"),
                        cur("Maintenance", "maintenance"),
                        cur("Utilities", "utilities"),
                        cur("HOA Fees", "hoa_fees"),
                        cur("Total Operating Expenses", "total_operating_expenses"),
                    ],
                ),
                (
                    "FINANCING",
                    [
                        cur("Loan Amount", "loan_amount"),
                        cur("Down Payment", "down_payment"),
                        cur("Closing Costs", "closing_costs"),
                        cur("Monthly P&I", "monthly_pi"),
                        cur("Annual Debt Service", "annual_debt_service"),
                        cur("Total Cash Required", "total_cash_required"),
                    ],
                ),
                (
                    "SUMMARY",
                    [
                        cur("NOI", "noi"),
                        cur("Monthly Cash Flow", "monthly_cash_flow"),
                        cur("Annual Cash Flow", "annual_cash_flow"),
                        pct("Cap Rate", "cap_rate"),
                        pct("Cash-on-Cash Return", "cash_on_cash_return"),
                        num("DSCR", "dscr"),
                        pct("1% Rule", "one_percent_rule"),
                    ],
                ),
            ]

        if key == "str":
            return [
                (
                    "REVENUE",
                    [
                        cur("Average Daily Rate", "average_daily_rate"),
                        pct("Occupancy Rate", "occupancy_rate"),
                        num("Nights Occupied", "nights_occupied"),
                        cur("Rental Revenue", "rental_revenue"),
                        cur("Total Gross Revenue", "total_gross_revenue"),
                    ],
                ),
                (
                    "EXPENSES",
                    [
                        cur("Property Taxes", "property_taxes"),
                        cur("Insurance", "insurance"),
                        cur("Platform Fees", "platform_fees"),
                        cur("STR Management", "str_management"),
                        cur("Cleaning Costs", "cleaning_costs"),
                        cur("Total Operating Expenses", "total_operating_expenses"),
                    ],
                ),
                (
                    "FINANCING",
                    [
                        cur("Loan Amount", "loan_amount"),
                        cur("Down Payment", "down_payment"),
                        cur("Closing Costs", "closing_costs"),
                        cur("Furniture Setup", "furniture_setup"),
                        cur("Monthly P&I", "monthly_pi"),
                        cur("Total Cash Required", "total_cash_required"),
                    ],
                ),
                (
                    "SUMMARY",
                    [
                        cur("NOI", "noi"),
                        cur("Monthly Cash Flow", "monthly_cash_flow"),
                        cur("Annual Cash Flow", "annual_cash_flow"),
                        pct("Cap Rate", "cap_rate"),
                        pct("Cash-on-Cash Return", "cash_on_cash_return"),
                        pct("Break-Even Occupancy", "break_even_occupancy"),
                        num("DSCR", "dscr"),
                    ],
                ),
            ]

        if key == "brrrr":
            return [
                (
                    "PHASE 1 — BUY",
                    [
                        cur("Purchase Price", "purchase_price"),
                        cur("Down Payment", "down_payment"),
                        cur("Closing Costs", "closing_costs"),
                        cur("Cash Required (Phase 1)", "cash_required_phase1"),
                    ],
                ),
                (
                    "PHASE 2 — REHAB",
                    [
                        cur("Renovation Budget", "renovation_budget"),
                        cur("Contingency", "contingency"),
                        cur("Holding Costs", "holding_costs"),
                        cur("Cash Required (Phase 2)", "cash_required_phase2"),
                    ],
                ),
                (
                    "PHASE 3 — RENT",
                    [
                        cur("ARV", "arv"),
                        cur("Post-Rehab Monthly Rent", "post_rehab_monthly_rent"),
                        pct("Estimated Cap Rate", "estimated_cap_rate"),
                    ],
                ),
                (
                    "PHASE 4 — REFINANCE",
                    [
                        cur("Refinance Loan Amount", "refinance_loan_amount"),
                        cur("Cash Out at Refinance", "cash_out_at_refinance"),
                        cur("New Monthly P&I", "new_monthly_pi"),
                    ],
                ),
                (
                    "PHASE 5 — REPEAT",
                    [
                        cur("Total Cash Invested", "total_cash_invested"),
                        pct("Capital Recycled", "capital_recycled_pct"),
                        cur("Cash Left in Deal", "cash_left_in_deal"),
                        cur("Equity Position", "equity_position"),
                        cur("Post-Refi Monthly Cash Flow", "post_refi_monthly_cash_flow"),
                        pct("Post-Refi Cash-on-Cash", "post_refi_cash_on_cash"),
                        txt("Infinite ROI", "infinite_roi_achieved"),
                    ],
                ),
            ]

        if key == "flip":
            return [
                (
                    "ACQUISITION",
                    [
                        cur("Purchase Price", "purchase_price"),
                        cur("Hard Money Loan", "hard_money_loan"),
                        cur("Down Payment", "down_payment"),
                        cur("Closing Costs", "closing_costs"),
                        cur("Total Acquisition Cash", "total_acquisition_cash"),
                    ],
                ),
                (
                    "RENOVATION & HOLDING",
                    [
                        cur("Renovation Budget", "renovation_budget"),
                        cur("Contingency", "contingency"),
                        cur("Total Holding Costs", "total_holding_costs"),
                    ],
                ),
                (
                    "SALE",
                    [
                        cur("ARV", "arv"),
                        cur("Realtor Commission", "realtor_commission"),
                        cur("Seller Closing Costs", "seller_closing_costs"),
                        cur("Net Sale Proceeds", "net_sale_proceeds"),
                    ],
                ),
                (
                    "PROFIT & METRICS",
                    [
                        cur("Total Project Cost", "total_project_cost"),
                        cur("Gross Profit", "gross_profit"),
                        cur("Net Profit (Pre-Tax)", "net_profit_before_tax"),
                        cur("Net Profit (After-Tax)", "net_profit_after_tax"),
                        pct("ROI", "roi"),
                        pct("Annualized ROI", "annualized_roi"),
                        cur("70% Rule Max Price", "seventy_pct_max_price"),
                        txt("Meets 70% Rule", "meets_70_rule"),
                        cur("Total Cash Required", "total_cash_required"),
                    ],
                ),
            ]

        if key == "house_hack":
            return [
                (
                    "ACQUISITION (FHA)",
                    [
                        cur("Purchase Price", "purchase_price"),
                        cur("Down Payment", "down_payment"),
                        cur("Closing Costs", "closing_costs"),
                        cur("Loan Amount", "loan_amount"),
                        cur("Total Cash Required", "total_cash_required"),
                    ],
                ),
                (
                    "MONTHLY COSTS",
                    [
                        cur("Monthly P&I", "monthly_pi"),
                        cur("Monthly MIP", "monthly_mip"),
                        cur("Monthly PITI", "monthly_piti"),
                    ],
                ),
                (
                    "SCENARIO A — RENT ROOMS",
                    [
                        num("Rooms Rented", "rooms_rented"),
                        cur("Room Rent", "room_rent"),
                        cur("Total Monthly Income", "total_monthly_income"),
                        cur("Net Housing Cost", "net_housing_cost_scenario_a"),
                        cur("Savings vs Renting", "savings_vs_renting_a"),
                    ],
                ),
                (
                    "KEY METRICS",
                    [
                        pct("Housing Cost Offset", "housing_cost_offset_pct"),
                        cur("Live-Free Threshold", "live_free_threshold"),
                        pct("ROI on Savings", "roi_on_savings"),
                    ],
                ),
            ]

        if key == "wholesale":
            return [
                (
                    "DEAL STRUCTURE",
                    [
                        cur("Contract Price", "contract_price"),
                        cur("Earnest Money", "earnest_money"),
                        cur("Assignment Fee", "assignment_fee"),
                        cur("End Buyer Price", "end_buyer_price"),
                    ],
                ),
                (
                    "PROPERTY ANALYSIS",
                    [
                        cur("ARV", "arv"),
                        cur("Estimated Rehab", "estimated_rehab"),
                        cur("70% Max Offer", "seventy_pct_max_offer"),
                        cur("Spread Available", "spread_available"),
                        txt("Deal Viability", "deal_viability"),
                    ],
                ),
                (
                    "PROFIT",
                    [
                        cur("Marketing Costs", "marketing_costs"),
                        cur("Total Cash at Risk", "total_cash_at_risk"),
                        cur("Gross Profit", "gross_profit"),
                        cur("Net Profit", "net_profit"),
                        pct("ROI", "roi"),
                        pct("Annualized ROI", "annualized_roi"),
                    ],
                ),
            ]

        return [("METRICS", [])]


comprehensive_excel_exporter = ComprehensiveExcelExporter()
