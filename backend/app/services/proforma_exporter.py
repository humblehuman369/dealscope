"""
Financial Proforma Excel Exporter
Generates accounting-standard XLSX files using openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from typing import Dict, Any, List, Tuple
from datetime import datetime
from io import BytesIO

from app.schemas.proforma import FinancialProforma


def get_price_label(is_off_market: bool = False, listing_status: str = None) -> str:
    """Get appropriate price label based on property status."""
    if is_off_market:
        return "Est. Market Value"
    if listing_status == "PENDING":
        return "List Price (Pending)"
    return "List Price"


class ProformaExcelExporter:
    """Generate professional Excel proforma workbooks."""
    
    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=10)
    TOTAL_FONT = Font(bold=True, size=10)
    HIGHLIGHT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    CURRENCY_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    CURRENCY_CENTS_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    PERCENT_FORMAT = '0.00%'
    NUMBER_FORMAT = '#,##0.00'
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, proforma: FinancialProforma, is_off_market: bool = False, listing_status: str = None):
        self.data = proforma
        self.wb = Workbook()
        self.price_label = get_price_label(is_off_market, listing_status)
        
    def generate(self) -> BytesIO:
        """Generate complete proforma workbook."""
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Create tabs
        self._create_property_summary_tab()
        self._create_income_expense_tab()
        self._create_cash_flow_projection_tab()
        self._create_loan_amortization_tab()
        self._create_depreciation_tab()
        self._create_exit_analysis_tab()
        self._create_returns_summary_tab()
        self._create_assumptions_tab()
        
        # Save to buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_property_summary_tab(self):
        """Tab 1: Property & Deal Summary"""
        ws = self.wb.create_sheet("Property Summary")
        
        # Property Information Header
        self._write_header(ws, 1, "PROPERTY INFORMATION")
        
        rows = [
            ("Address", self.data.property.address),
            ("City, State ZIP", f"{self.data.property.city}, {self.data.property.state} {self.data.property.zip}"),
            ("Property Type", self.data.property.property_type),
            ("Bedrooms / Bathrooms", f"{self.data.property.bedrooms} / {self.data.property.bathrooms}"),
            ("Square Feet", f"{self.data.property.square_feet:,}"),
            ("Year Built", str(self.data.property.year_built)),
            ("Lot Size (sq ft)", f"{self.data.property.lot_size:,}"),
        ]
        row_num = 2
        for label, value in rows:
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value)
            row_num += 1
        
        row_num += 1
        
        # Acquisition Section
        self._write_header(ws, row_num, "ACQUISITION")
        row_num += 1
        
        acq_rows = [
            ("Purchase Price", self.data.acquisition.purchase_price, "currency"),
            (self.price_label, self.data.acquisition.list_price, "currency"),
            ("Discount from List", self.data.acquisition.discount_from_list / 100, "percent"),
            ("Closing Costs", self.data.acquisition.closing_costs, "currency"),
            ("Rehab/Renovation", self.data.acquisition.rehab_costs, "currency"),
            ("Total Acquisition Cost", self.data.acquisition.total_acquisition_cost, "currency_total"),
        ]
        
        for label, value, fmt in acq_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        
        # Financing Section
        self._write_header(ws, row_num, "FINANCING")
        row_num += 1
        
        fin_rows = [
            ("Down Payment", self.data.financing.down_payment, "currency"),
            ("Down Payment %", self.data.financing.down_payment_percent / 100, "percent"),
            ("Loan Amount", self.data.financing.loan_amount, "currency"),
            ("Interest Rate", self.data.financing.interest_rate / 100, "percent"),
            ("Loan Term (Years)", self.data.financing.loan_term_years, "number"),
            ("Monthly Payment (P&I)", self.data.financing.monthly_payment, "currency"),
            ("Total Interest Over Life", self.data.financing.total_interest_over_life, "currency"),
        ]
        
        for label, value, fmt in fin_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        self._format_column_widths(ws, [35, 25])
        
    def _create_income_expense_tab(self):
        """Tab 2: Year 1 Income Statement"""
        ws = self.wb.create_sheet("Income & Expenses")
        
        row_num = 1
        self._write_header(ws, row_num, "YEAR 1 INCOME STATEMENT")
        row_num += 2
        
        # Income Section
        self._write_subheader(ws, row_num, "INCOME")
        row_num += 1
        
        income_rows = [
            ("Gross Scheduled Rent", self.data.income.annual_gross_rent, "currency"),
            ("Less: Vacancy Allowance", -self.data.income.vacancy_allowance, "currency"),
            ("Other Income", self.data.income.other_income, "currency"),
        ]
        
        for label, value, fmt in income_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        # EGI Total
        ws.cell(row=row_num, column=1, value="EFFECTIVE GROSS INCOME")
        ws.cell(row=row_num, column=1).font = self.TOTAL_FONT
        cell = ws.cell(row=row_num, column=2, value=self.data.income.effective_gross_income)
        self._apply_format(cell, "currency_total")
        row_num += 2
        
        # Expenses Section
        self._write_subheader(ws, row_num, "OPERATING EXPENSES")
        row_num += 1
        
        expense_rows = [
            ("Property Taxes", self.data.expenses.property_taxes, "currency"),
            ("Insurance", self.data.expenses.insurance, "currency"),
            ("HOA Fees", self.data.expenses.hoa_fees, "currency"),
            ("Property Management", self.data.expenses.management, "currency"),
            ("Maintenance & Repairs", self.data.expenses.maintenance, "currency"),
            ("Utilities", self.data.expenses.utilities, "currency"),
            ("Landscaping", self.data.expenses.landscaping, "currency"),
            ("Pest Control", self.data.expenses.pest_control, "currency"),
            ("CapEx Reserve", self.data.expenses.cap_ex_reserve, "currency"),
            ("Other Expenses", self.data.expenses.other_expenses, "currency"),
        ]
        
        for label, value, fmt in expense_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        # Total OpEx
        ws.cell(row=row_num, column=1, value="TOTAL OPERATING EXPENSES")
        ws.cell(row=row_num, column=1).font = self.TOTAL_FONT
        cell = ws.cell(row=row_num, column=2, value=self.data.expenses.total_operating_expenses)
        self._apply_format(cell, "currency_total")
        row_num += 2
        
        # NOI
        ws.cell(row=row_num, column=1, value="NET OPERATING INCOME (NOI)")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
        cell = ws.cell(row=row_num, column=2, value=self.data.metrics.net_operating_income)
        cell.fill = self.HIGHLIGHT_FILL
        self._apply_format(cell, "currency_total")
        row_num += 2
        
        # Debt Service
        self._write_subheader(ws, row_num, "DEBT SERVICE")
        row_num += 1
        
        ws.cell(row=row_num, column=1, value="Annual Mortgage (P&I)")
        cell = ws.cell(row=row_num, column=2, value=self.data.metrics.annual_debt_service)
        self._apply_format(cell, "currency")
        row_num += 2
        
        # Cash Flow
        ws.cell(row=row_num, column=1, value="PRE-TAX CASH FLOW")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=11)
        cell = ws.cell(row=row_num, column=2, value=self.data.metrics.annual_cash_flow)
        cell.fill = self.HIGHLIGHT_FILL
        self._apply_format(cell, "currency_total")
        row_num += 1
        
        ws.cell(row=row_num, column=1, value="Monthly Cash Flow")
        cell = ws.cell(row=row_num, column=2, value=self.data.metrics.monthly_cash_flow)
        self._apply_format(cell, "currency")
        
        self._format_column_widths(ws, [35, 20])
        
    def _create_cash_flow_projection_tab(self):
        """Tab 3: Multi-Year Cash Flow Projection"""
        ws = self.wb.create_sheet("Cash Flow Projection")
        
        years = self.data.projections.hold_period_years
        
        # Headers
        headers = ["Metric"] + [f"Year {i}" for i in range(1, years + 1)]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        
        projections = self.data.projections.annual_projections
        
        # Data rows
        metrics = [
            ("Gross Rental Income", [p.gross_rental_income for p in projections]),
            ("Vacancy Loss", [p.gross_rental_income - p.effective_gross_income for p in projections]),
            ("Effective Gross Income", [p.effective_gross_income for p in projections]),
            ("Operating Expenses", [p.operating_expenses for p in projections]),
            ("Net Operating Income", [p.net_operating_income for p in projections]),
            ("Debt Service", [p.total_debt_service for p in projections]),
            ("Pre-Tax Cash Flow", [p.pre_tax_cash_flow for p in projections]),
            ("Depreciation", [p.depreciation for p in projections]),
            ("Taxable Income", [p.taxable_income for p in projections]),
            ("Tax Liability/(Benefit)", [p.estimated_tax_liability for p in projections]),
            ("After-Tax Cash Flow", [p.after_tax_cash_flow for p in projections]),
        ]
        
        row_num = 2
        for label, values in metrics:
            ws.cell(row=row_num, column=1, value=label)
            for col, value in enumerate(values, 2):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.number_format = self.CURRENCY_FORMAT
            row_num += 1
        
        row_num += 1
        
        # Property value and equity
        ws.cell(row=row_num, column=1, value="Property Value")
        ws.cell(row=row_num, column=1).font = self.SUBHEADER_FONT
        for col, value in enumerate(self.data.projections.property_values, 2):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.number_format = self.CURRENCY_FORMAT
        row_num += 1
        
        ws.cell(row=row_num, column=1, value="Loan Balance")
        for col, value in enumerate(self.data.projections.loan_balances, 2):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.number_format = self.CURRENCY_FORMAT
        row_num += 1
        
        ws.cell(row=row_num, column=1, value="Equity Position")
        ws.cell(row=row_num, column=1).font = self.SUBHEADER_FONT
        for col, value in enumerate(self.data.projections.equity_positions, 2):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.number_format = self.CURRENCY_FORMAT
        row_num += 1
        
        ws.cell(row=row_num, column=1, value="Cumulative Cash Flow")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        for col, value in enumerate(self.data.projections.cumulative_cash_flow, 2):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.number_format = self.CURRENCY_FORMAT
            cell.fill = self.HIGHLIGHT_FILL
        
        self._format_column_widths(ws, [25] + [15] * years)
        
    def _create_loan_amortization_tab(self):
        """Tab 4: Full Loan Amortization Schedule"""
        ws = self.wb.create_sheet("Loan Amortization")
        
        headers = [
            "Month", "Year", "Beginning Balance", "Payment",
            "Principal", "Interest", "Ending Balance",
            "Cumulative Principal", "Cumulative Interest"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        
        for row_idx, amort_row in enumerate(self.data.amortization_schedule, 2):
            ws.cell(row=row_idx, column=1, value=amort_row.month)
            ws.cell(row=row_idx, column=2, value=amort_row.year)
            
            for col, value in enumerate([
                amort_row.beginning_balance,
                amort_row.scheduled_payment,
                amort_row.principal_payment,
                amort_row.interest_payment,
                amort_row.ending_balance,
                amort_row.cumulative_principal,
                amort_row.cumulative_interest,
            ], 3):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.number_format = self.CURRENCY_CENTS_FORMAT
        
        self._format_column_widths(ws, [8, 6, 18, 15, 15, 15, 18, 18, 18])
        
    def _create_depreciation_tab(self):
        """Tab 5: Depreciation Schedule"""
        ws = self.wb.create_sheet("Depreciation")
        
        row_num = 1
        self._write_header(ws, row_num, "DEPRECIATION BASIS CALCULATION")
        row_num += 1
        
        dep = self.data.depreciation
        
        basis_rows = [
            ("Purchase Price", dep.purchase_price, "currency"),
            ("Less: Land Value", -dep.land_value, "currency"),
            (f"  Land Value % (non-depreciable)", dep.land_value_percent, "percent"),
            ("Improvement Value", dep.improvement_value, "currency"),
            ("Add: Capitalized Closing Costs", dep.capitalized_closing_costs, "currency"),
            ("Add: Rehabilitation Costs", dep.rehab_costs, "currency"),
        ]
        
        for label, value, fmt in basis_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        # Total Depreciable Basis
        ws.cell(row=row_num, column=1, value="TOTAL DEPRECIABLE BASIS")
        ws.cell(row=row_num, column=1).font = self.TOTAL_FONT
        cell = ws.cell(row=row_num, column=2, value=dep.total_depreciable_basis)
        self._apply_format(cell, "currency_total")
        row_num += 2
        
        self._write_header(ws, row_num, "DEPRECIATION SCHEDULE")
        row_num += 1
        
        schedule_rows = [
            ("Depreciation Method", dep.depreciation_method.value, "text"),
            ("Recovery Period (Years)", dep.depreciation_years, "number"),
            ("Annual Depreciation", dep.annual_depreciation, "currency"),
            ("Monthly Depreciation", dep.monthly_depreciation, "currency"),
        ]
        
        for label, value, fmt in schedule_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        
        # Year-by-year depreciation schedule
        dep_headers = ["Year", "Beginning Book Value", "Depreciation", "Ending Book Value", "Accumulated Depreciation"]
        for col, header in enumerate(dep_headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        row_num += 1
        
        book_value = dep.total_depreciable_basis
        accumulated = 0
        for year in range(1, int(dep.depreciation_years) + 2):
            annual_dep = min(dep.annual_depreciation, book_value)
            if annual_dep <= 0:
                break
            accumulated += annual_dep
            ending = book_value - annual_dep
            
            ws.cell(row=row_num, column=1, value=year)
            for col, value in enumerate([book_value, annual_dep, ending, accumulated], 2):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.number_format = self.CURRENCY_FORMAT
            
            book_value = ending
            row_num += 1
        
        self._format_column_widths(ws, [8, 22, 18, 22, 24])
        
    def _create_exit_analysis_tab(self):
        """Tab 6: Exit/Disposition Analysis"""
        ws = self.wb.create_sheet("Exit Analysis")
        
        row_num = 1
        self._write_header(ws, row_num, "EXIT ANALYSIS")
        row_num += 1
        
        exit_data = self.data.exit
        
        ws.cell(row=row_num, column=1, value="Hold Period (Years)")
        ws.cell(row=row_num, column=2, value=exit_data.hold_period_years)
        row_num += 2
        
        self._write_subheader(ws, row_num, "SALE PROCEEDS")
        row_num += 1
        
        sale_rows = [
            ("Projected Sale Price", exit_data.projected_sale_price, "currency"),
            ("Less: Broker Commission", -exit_data.broker_commission, "currency"),
            ("Less: Closing Costs", -exit_data.closing_costs, "currency"),
            ("Less: Loan Payoff", -exit_data.remaining_loan_balance, "currency"),
        ]
        
        for label, value, fmt in sale_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        ws.cell(row=row_num, column=1, value="NET SALE PROCEEDS (Before Tax)")
        ws.cell(row=row_num, column=1).font = self.TOTAL_FONT
        cell = ws.cell(row=row_num, column=2, value=exit_data.net_sale_proceeds)
        self._apply_format(cell, "currency_total")
        row_num += 2
        
        self._write_subheader(ws, row_num, "CAPITAL GAINS CALCULATION")
        row_num += 1
        
        cg_rows = [
            ("Original Cost Basis", exit_data.adjusted_cost_basis + exit_data.accumulated_depreciation, "currency"),
            ("Less: Accumulated Depreciation", -exit_data.accumulated_depreciation, "currency"),
            ("Adjusted Cost Basis", exit_data.adjusted_cost_basis, "currency"),
            ("Total Gain on Sale", exit_data.total_gain, "currency"),
        ]
        
        for label, value, fmt in cg_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        self._write_subheader(ws, row_num, "TAX ON SALE")
        row_num += 1
        
        tax_rows = [
            ("Depreciation Recapture (25%)", exit_data.depreciation_recapture, "currency"),
            ("  Tax on Recapture", exit_data.depreciation_recapture_tax, "currency"),
            ("Capital Gain", exit_data.capital_gain, "currency"),
            (f"  Tax on Capital Gain ({exit_data.capital_gains_tax_rate:.0%})", exit_data.capital_gains_tax, "currency"),
        ]
        
        for label, value, fmt in tax_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        ws.cell(row=row_num, column=1, value="TOTAL TAX ON SALE")
        ws.cell(row=row_num, column=1).font = self.TOTAL_FONT
        cell = ws.cell(row=row_num, column=2, value=exit_data.total_tax_on_sale)
        self._apply_format(cell, "currency_total")
        row_num += 2
        
        ws.cell(row=row_num, column=1, value="AFTER-TAX PROCEEDS")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        cell = ws.cell(row=row_num, column=2, value=exit_data.after_tax_proceeds)
        cell.fill = self.HIGHLIGHT_FILL
        self._apply_format(cell, "currency_total")
        
        self._format_column_widths(ws, [35, 22])
        
    def _create_returns_summary_tab(self):
        """Tab 7: Investment Returns Summary"""
        ws = self.wb.create_sheet("Returns Summary")
        
        row_num = 1
        self._write_header(ws, row_num, "INVESTMENT RETURNS SUMMARY")
        row_num += 2
        
        self._write_subheader(ws, row_num, "OPERATING METRICS (Year 1)")
        row_num += 1
        
        metrics = self.data.metrics
        returns = self.data.returns
        
        op_rows = [
            ("Cap Rate", metrics.cap_rate / 100, "percent"),
            ("Cash-on-Cash Return", metrics.cash_on_cash_return / 100, "percent"),
            ("DSCR", metrics.dscr, "number"),
            ("Gross Rent Multiplier", metrics.gross_rent_multiplier, "number"),
            ("1% Rule", metrics.one_percent_rule / 100, "percent"),
            ("Break-Even Occupancy", metrics.break_even_occupancy / 100, "percent"),
        ]
        
        for label, value, fmt in op_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        self._write_subheader(ws, row_num, "TOTAL INVESTMENT RETURNS")
        row_num += 1
        
        return_rows = [
            ("Internal Rate of Return (IRR)", returns.irr / 100, "percent"),
            ("Equity Multiple", f"{returns.equity_multiple:.2f}x", "text"),
            ("Total Cash Flows (Over Hold)", returns.total_cash_flows, "currency"),
            ("Total Distributions (incl. Sale)", returns.total_distributions, "currency"),
            ("Average Annual Return", returns.average_annual_return / 100, "percent"),
            ("CAGR", returns.cagr / 100, "percent"),
            ("Payback Period", f"{returns.payback_period_months or 'N/A'} months" if returns.payback_period_months else "N/A", "text"),
        ]
        
        for label, value, fmt in return_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        self._write_subheader(ws, row_num, "DEAL SCORE")
        row_num += 1
        
        score_rows = [
            ("Score", f"{self.data.deal_score.score}/100", "text"),
            ("Grade", self.data.deal_score.grade or "N/A", "text"),
            ("Verdict", self.data.deal_score.verdict or "N/A", "text"),
            ("Income Value", self.data.deal_score.income_value, "currency"),
        ]
        
        for label, value, fmt in score_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        self._format_column_widths(ws, [35, 22])
        
    def _create_assumptions_tab(self):
        """Tab 8: All Assumptions & Data Sources"""
        ws = self.wb.create_sheet("Assumptions")
        
        row_num = 1
        self._write_header(ws, row_num, "ANALYSIS ASSUMPTIONS")
        row_num += 1
        
        meta_rows = [
            ("Generated", self.data.generated_at),
            ("Property ID", self.data.property_id),
            ("Strategy", self.data.strategy_type.upper()),
        ]
        
        for label, value in meta_rows:
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value)
            row_num += 1
        
        row_num += 1
        self._write_subheader(ws, row_num, "FINANCING ASSUMPTIONS")
        row_num += 1
        
        fin_rows = [
            ("Down Payment %", self.data.financing.down_payment_percent / 100, "percent"),
            ("Interest Rate", self.data.financing.interest_rate / 100, "percent"),
            ("Loan Term", f"{self.data.financing.loan_term_years} years", "text"),
            ("Loan Type", self.data.financing.loan_type, "text"),
        ]
        
        for label, value, fmt in fin_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        self._write_subheader(ws, row_num, "OPERATING ASSUMPTIONS")
        row_num += 1
        
        op_rows = [
            ("Vacancy Rate", self.data.income.vacancy_percent / 100, "percent"),
            ("Management Fee", self.data.expenses.management_percent / 100, "percent"),
            ("Maintenance Reserve", self.data.expenses.maintenance_percent / 100, "percent"),
            ("CapEx Reserve", self.data.expenses.cap_ex_reserve_percent / 100, "percent"),
        ]
        
        for label, value, fmt in op_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        self._write_subheader(ws, row_num, "GROWTH ASSUMPTIONS")
        row_num += 1
        
        growth_rows = [
            ("Appreciation Rate", self.data.projections.appreciation_rate / 100, "percent"),
            ("Rent Growth Rate", self.data.projections.rent_growth_rate / 100, "percent"),
            ("Expense Growth Rate", self.data.projections.expense_growth_rate / 100, "percent"),
        ]
        
        for label, value, fmt in growth_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        self._write_subheader(ws, row_num, "TAX ASSUMPTIONS")
        row_num += 1
        
        tax_rows = [
            ("Land Value %", self.data.depreciation.land_value_percent, "percent"),
            ("Depreciation Years", self.data.depreciation.depreciation_years, "number"),
            ("Marginal Tax Rate", self.data.projections.annual_projections[0].marginal_tax_rate if self.data.projections.annual_projections else 0.24, "percent"),
            ("Capital Gains Rate", self.data.exit.capital_gains_tax_rate, "percent"),
        ]
        
        for label, value, fmt in tax_rows:
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)
            self._apply_format(cell, fmt)
            row_num += 1
        
        row_num += 1
        self._write_subheader(ws, row_num, "DATA SOURCES")
        row_num += 1
        
        source_rows = [
            ("Rent Estimate", self.data.sources.rent_estimate_source),
            ("Property Value", self.data.sources.property_value_source),
            ("Tax Data", self.data.sources.tax_data_source),
            ("Market Data", self.data.sources.market_data_source),
            ("Data Freshness", self.data.sources.data_freshness),
        ]
        
        for label, value in source_rows:
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value)
            row_num += 1
        
        self._format_column_widths(ws, [25, 25])
    
    # Helper methods
    def _write_header(self, ws, row_num: int, text: str):
        """Write a header row with styling."""
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    
    def _write_subheader(self, ws, row_num: int, text: str):
        """Write a subheader row with styling."""
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = self.SUBHEADER_FONT
        cell.fill = self.SUBHEADER_FILL
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    
    def _apply_format(self, cell, fmt: str):
        """Apply formatting to a cell."""
        if fmt == "currency":
            cell.number_format = self.CURRENCY_FORMAT
        elif fmt == "currency_total":
            cell.number_format = self.CURRENCY_FORMAT
            cell.font = self.TOTAL_FONT
        elif fmt == "percent":
            cell.number_format = self.PERCENT_FORMAT
        elif fmt == "number":
            cell.number_format = self.NUMBER_FORMAT
        # "text" format needs no special handling
    
    def _format_column_widths(self, ws, widths: List[int]):
        """Set column widths."""
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
