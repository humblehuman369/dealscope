"""
House Hack Proforma — Excel Exporter  (v1.0)

Generates a professional proforma for owner-occupied house hacking
strategies (FHA financing, room rentals, duplex conversion):

  1. Property Summary     - property details, FHA terms, overview
  2. Monthly Budget       - PITI + MIP breakdown
  3. Scenario A           - rent rooms, net housing cost, savings
  4. Scenario B           - duplex conversion (if applicable)
  5. Savings Analysis     - side-by-side comparison, live-free analysis
  6. Assumptions          - FHA terms, rental assumptions, data sources
"""

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from app.schemas.proforma import FinancialProforma

# ── Style tokens ─────────────────────────────────────────────────────────────

_BRAND = "0A84FF"
_HEADER_BG = "1F4E79"
_GREEN = "22C55E"
_RED = "EF4444"
_WHITE = "FFFFFF"

_HDR_FILL = PatternFill("solid", fgColor=_HEADER_BG)
_HDR_FONT = Font(bold=True, color=_WHITE, size=11, name="Calibri")
_SUB_FILL = PatternFill("solid", fgColor="D6DCE4")
_SUB_FONT = Font(bold=True, size=10, name="Calibri")
_TOTAL_FNT = Font(bold=True, size=10, name="Calibri")
_BODY_FONT = Font(size=10, name="Calibri")
_GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAD_FILL = PatternFill("solid", fgColor="FCE4EC")

_CUR = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
_PCT = "0.00%"
_INT = "#,##0"


class HouseHackExcelExporter:
    """Generate a house-hack-specific proforma workbook."""

    def __init__(self, proforma: FinancialProforma):
        self.d = proforma
        self.wb = Workbook()
        self.bd: dict[str, Any] = proforma.strategy_breakdown or {}

    # ── public ────────────────────────────────────────────────────────────

    def generate(self) -> BytesIO:
        self.wb.remove(self.wb.active)

        self._tab_property_summary()
        self._tab_monthly_budget()
        self._tab_scenario_a()
        self._tab_scenario_b()
        self._tab_savings_analysis()
        self._tab_assumptions()

        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    # ── Tab 1: Property Summary ───────────────────────────────────────────

    def _tab_property_summary(self):
        ws = self.wb.create_sheet("Property Summary")
        r = 1

        r = self._section_header(ws, r, "HOUSE HACK SUMMARY")

        prop = self.d.property
        rows = [
            ("Property Address", prop.address),
            ("City / State / Zip", f"{prop.city}, {prop.state} {prop.zip}"),
            ("Property Type", prop.property_type),
            ("Bedrooms / Bathrooms", f"{prop.bedrooms} bd / {prop.bathrooms} ba"),
            ("Square Feet", prop.square_feet),
            ("Year Built", prop.year_built),
            ("Lot Size (sqft)", prop.lot_size),
        ]
        for label, val in rows:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "STRATEGY OVERVIEW")
        rows2 = [
            ("Strategy", "House Hack (Owner-Occupied)"),
            ("Objective", "Reduce or eliminate personal housing costs"),
            ("Financing", "FHA Loan (3.5% down)"),
            ("Risk Profile", "Low — owner-occupied, government-backed loan"),
            ("Tax Benefit", "Mortgage interest deduction + rental income offsets"),
        ]
        for label, val in rows2:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "KEY NUMBERS AT A GLANCE")
        net_a = self.bd.get("net_housing_cost_scenario_a", 0)
        offset = self.bd.get("housing_cost_offset_pct", 0)
        savings = self.bd.get("savings_vs_renting_a", 0)
        live_free = net_a <= 0

        glance = [
            ("Purchase Price", self.bd.get("purchase_price", 0), _CUR),
            ("Down Payment (3.5% FHA)", self.bd.get("down_payment", 0), _CUR),
            ("Total Cash Required", self.bd.get("total_cash_required", 0), _CUR),
            ("Monthly Rental Income", self.bd.get("total_monthly_income", 0), _CUR),
            ("Net Monthly Housing Cost", net_a, _CUR),
            ("Housing Cost Offset", offset, _PCT),
            ("Monthly Savings vs. Renting", savings, _CUR),
            ("Living for Free?", "YES" if live_free else "NO", None),
        ]
        for label, val, fmt in glance:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        self._auto_width(ws)

    # ── Tab 2: Monthly Budget ─────────────────────────────────────────────

    def _tab_monthly_budget(self):
        ws = self.wb.create_sheet("Monthly Budget")
        r = 1

        r = self._section_header(ws, r, "MONTHLY HOUSING COSTS (PITI + MIP)")
        pi = self.bd.get("monthly_pi", 0)
        mip = self.bd.get("monthly_mip", 0)
        piti = self.bd.get("monthly_piti", 0)
        purchase = self.bd.get("purchase_price", 0)
        loan = self.bd.get("loan_amount", 0)

        # Derive taxes & insurance from PITI
        piti - pi - mip - (self.d.expenses.insurance / 12 if self.d.expenses else 0)
        self.d.expenses.insurance / 12 if self.d.expenses else 0
        # Simpler: calculate from known values
        taxes_annual = self.d.expenses.property_taxes if self.d.expenses else purchase * 0.012
        insurance_annual = self.d.expenses.insurance if self.d.expenses else purchase * 0.005

        rows = [
            ("Loan Amount", loan, _CUR),
            ("Principal & Interest (P&I)", pi, _CUR),
            ("FHA Mortgage Insurance (MIP)", mip, _CUR),
            ("Property Taxes", taxes_annual / 12, _CUR),
            ("Homeowner's Insurance", insurance_annual / 12, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Monthly PITI + MIP").font = _TOTAL_FNT
        c = ws.cell(r, 2, piti)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "ADDITIONAL MONTHLY COSTS")
        utilities = self.bd.get("utilities_shared", 0)
        maintenance = self.bd.get("maintenance", 0)
        total_expenses = self.bd.get("total_monthly_expenses", 0)

        rows2 = [
            ("Shared Utilities", utilities, _CUR),
            ("Maintenance", maintenance, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Monthly Expenses").font = _TOTAL_FNT
        c = ws.cell(r, 2, total_expenses)
        c.number_format = _CUR
        c.font = _TOTAL_FNT
        r += 2

        r = self._section_header(ws, r, "ANNUAL COST SUMMARY")
        rows3 = [
            ("Annual Housing Costs", total_expenses * 12, _CUR),
            ("Annual PITI + MIP Only", piti * 12, _CUR),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        self._auto_width(ws)

    # ── Tab 3: Scenario A — Rent Rooms ────────────────────────────────────

    def _tab_scenario_a(self):
        ws = self.wb.create_sheet("Scenario A - Rent Rooms")
        r = 1

        rooms = self.bd.get("rooms_rented", 0)
        room_rent = self.bd.get("room_rent", 0)
        income = self.bd.get("total_monthly_income", 0)
        expenses = self.bd.get("total_monthly_expenses", 0)
        net_a = self.bd.get("net_housing_cost_scenario_a", 0)
        savings = self.bd.get("savings_vs_renting_a", 0)

        r = self._section_header(ws, r, "SCENARIO A: RENT ROOMS / UNITS")
        rows = [
            ("Rooms / Units Rented", rooms, _INT),
            ("Rent per Room / Unit", room_rent, _CUR),
            ("Total Monthly Rental Income", income, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        r += 1
        r = self._section_header(ws, r, "NET HOUSING COST")
        rows2 = [
            ("Total Monthly Expenses", expenses, _CUR),
            ("- Rental Income", income, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Your Net Monthly Housing Cost").font = _TOTAL_FNT
        c = ws.cell(r, 2, net_a)
        c.number_format = _CUR
        c.font = Font(bold=True, size=12, color=_GREEN if net_a <= 0 else _RED, name="Calibri")
        r += 1

        if net_a <= 0:
            ws.cell(r, 1, "You are living for FREE (or getting paid to live!)").font = Font(
                bold=True, size=11, color=_GREEN, name="Calibri"
            )
            ws.cell(r, 1).fill = _GOOD_FILL
            r += 1
        r += 1

        r = self._section_header(ws, r, "SAVINGS vs. RENTING")
        rows3 = [
            ("Market Rent (comparable)", self.bd.get("owner_unit_market_rent", 1500), _CUR),
            ("Your Net Housing Cost", net_a, _CUR),
            ("Monthly Savings", savings, _CUR),
            ("Annual Savings", self.bd.get("annual_savings_a", 0), _CUR),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=(label == "Annual Savings"))

        self._auto_width(ws)

    # ── Tab 4: Scenario B — Duplex Conversion ─────────────────────────────

    def _tab_scenario_b(self):
        ws = self.wb.create_sheet("Scenario B - Duplex")
        r = 1

        conversion = self.bd.get("conversion_cost")
        unit2_rent = self.bd.get("unit2_rent")
        heloc = self.bd.get("heloc_payment")
        net_b = self.bd.get("net_housing_cost_scenario_b")
        savings_b = self.bd.get("savings_vs_renting_b")

        if conversion and unit2_rent:
            r = self._section_header(ws, r, "SCENARIO B: DUPLEX CONVERSION")
            rows = [
                ("Conversion Cost", conversion, _CUR),
                ("HELOC Payment (8%, 10yr)", heloc, _CUR),
                ("Unit 2 Monthly Rent", unit2_rent, _CUR),
            ]
            for label, val, fmt in rows:
                r = self._label_value(ws, r, label, val, fmt=fmt)

            r += 1
            r = self._section_header(ws, r, "NET HOUSING COST (DUPLEX)")
            ws.cell(r, 1, "= Your Net Monthly Housing Cost").font = _TOTAL_FNT
            c = ws.cell(r, 2, net_b)
            c.number_format = _CUR
            c.font = Font(bold=True, size=12, color=_GREEN if (net_b or 0) <= 0 else _RED, name="Calibri")
            r += 2

            r = self._section_header(ws, r, "SAVINGS vs. RENTING (DUPLEX)")
            rows2 = [
                ("Monthly Savings", savings_b, _CUR),
                ("Annual Savings", (savings_b or 0) * 12, _CUR),
            ]
            for label, val, fmt in rows2:
                r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)
        else:
            r = self._section_header(ws, r, "SCENARIO B: DUPLEX CONVERSION")
            r += 1
            ws.cell(r, 1, "Duplex conversion analysis not applicable for this property.").font = Font(
                italic=True, size=10, color="6B7280", name="Calibri"
            )
            r += 1
            ws.cell(r, 1, "This scenario requires a conversion cost estimate and Unit 2 rent projection.").font = Font(
                italic=True, size=10, color="6B7280", name="Calibri"
            )

        self._auto_width(ws)

    # ── Tab 5: Savings Analysis ───────────────────────────────────────────

    def _tab_savings_analysis(self):
        ws = self.wb.create_sheet("Savings Analysis")
        r = 1

        r = self._section_header(ws, r, "HOUSE HACK SAVINGS OVERVIEW")
        offset = self.bd.get("housing_cost_offset_pct", 0)
        threshold = self.bd.get("live_free_threshold", 0)
        roi = self.bd.get("roi_on_savings", 0)
        net_a = self.bd.get("net_housing_cost_scenario_a", 0)

        rows = [
            ("Housing Cost Offset %", offset, _PCT),
            ("Live-Free Threshold (rent needed)", threshold, _CUR),
            ("Your Monthly Rental Income", self.bd.get("total_monthly_income", 0), _CUR),
            ("Gap to Live Free", max(0, threshold - self.bd.get("total_monthly_income", 0)), _CUR),
            ("ROI on Savings", roi, _PCT),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        r += 1
        r = self._section_header(ws, r, "SCENARIO COMPARISON")

        # Table header
        headers = ["Metric", "Scenario A (Rooms)", "Scenario B (Duplex)"]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(r, col_idx, h)
            c.font = _SUB_FONT
            c.fill = _SUB_FILL
        r += 1

        net_b = self.bd.get("net_housing_cost_scenario_b")
        savings_a = self.bd.get("savings_vs_renting_a", 0)
        savings_b = self.bd.get("savings_vs_renting_b")

        comparisons = [
            ("Net Monthly Housing Cost", net_a, net_b),
            ("Monthly Savings vs. Renting", savings_a, savings_b),
            ("Annual Savings", self.bd.get("annual_savings_a", 0), (savings_b or 0) * 12 if savings_b else None),
        ]
        for label, val_a, val_b in comparisons:
            ws.cell(r, 1, label).font = _BODY_FONT
            ca = ws.cell(r, 2, val_a if val_a is not None else "N/A")
            if isinstance(val_a, (int, float)):
                ca.number_format = _CUR
            ca.font = _BODY_FONT
            cb = ws.cell(r, 3, val_b if val_b is not None else "N/A")
            if isinstance(val_b, (int, float)):
                cb.number_format = _CUR
            cb.font = _BODY_FONT
            r += 1

        r += 1
        r = self._section_header(ws, r, "5-YEAR SAVINGS PROJECTION")
        annual_a = self.bd.get("annual_savings_a", 0)
        headers2 = ["Year", "Cumulative Savings (A)"]
        for col_idx, h in enumerate(headers2, 1):
            c = ws.cell(r, col_idx, h)
            c.font = _SUB_FONT
            c.fill = _SUB_FILL
        r += 1

        for yr in range(1, 6):
            ws.cell(r, 1, f"Year {yr}").font = _BODY_FONT
            c = ws.cell(r, 2, annual_a * yr)
            c.number_format = _CUR
            c.font = _BODY_FONT
            r += 1

        self._auto_width(ws)

    # ── Tab 6: Assumptions ────────────────────────────────────────────────

    def _tab_assumptions(self):
        ws = self.wb.create_sheet("Assumptions")
        r = 1

        r = self._section_header(ws, r, "HOUSE HACK ASSUMPTIONS")
        assumptions = [
            ("Down Payment", "3.5% (FHA)"),
            ("FHA MIP Rate", "0.55% annually"),
            ("Loan Term", "30 years"),
            ("Rooms Rented", str(self.bd.get("rooms_rented", 0))),
            ("Rent per Room", f"${self.bd.get('room_rent', 0):,.0f}/mo"),
            ("Owner Unit Market Rent", f"${self.bd.get('owner_unit_market_rent', 1500):,.0f}/mo (comparison)"),
            ("Shared Utilities", f"${self.bd.get('utilities_shared', 150):,.0f}/mo"),
            ("Maintenance", f"${self.bd.get('maintenance', 200):,.0f}/mo"),
        ]
        for label, val in assumptions:
            r = self._label_value(ws, r, label, val)

        conversion = self.bd.get("conversion_cost")
        if conversion:
            r += 1
            r = self._section_header(ws, r, "DUPLEX CONVERSION ASSUMPTIONS")
            dup_assumptions = [
                ("Conversion Cost", f"${conversion:,.0f}"),
                ("HELOC Rate", "8%"),
                ("HELOC Term", "10 years"),
                ("Unit 2 Rent", f"${self.bd.get('unit2_rent', 0):,.0f}/mo"),
                ("Maintenance Increase", "25% (for duplex)"),
            ]
            for label, val in dup_assumptions:
                r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "DATA SOURCES")
        src = self.d.sources
        sources = [
            ("RentCast Estimate", src.rent_estimate_source),
            ("Property Value", src.property_value_source),
            ("Tax Data", src.tax_data_source),
            ("Market Data", src.market_data_source),
            ("Data Freshness", src.data_freshness),
        ]
        for label, val in sources:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "REPORT METADATA")
        r = self._label_value(ws, r, "Generated", datetime.now().strftime("%B %d, %Y at %I:%M %p"))
        r = self._label_value(ws, r, "Strategy", "House Hack (Owner-Occupied)")
        r = self._label_value(ws, r, "Engine", "DealGapIQ StrategyIQ")

        self._auto_width(ws)

    # ── Helpers ───────────────────────────────────────────────────────────

    def _section_header(self, ws, row: int, title: str) -> int:
        c = ws.cell(row, 1, title)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="left")
        for col in range(2, 4):
            ws.cell(row, col).fill = _HDR_FILL
        return row + 1

    def _label_value(
        self,
        ws,
        row: int,
        label: str,
        value: Any,
        *,
        fmt: str | None = None,
        highlight: bool = False,
    ) -> int:
        lc = ws.cell(row, 1, label)
        lc.font = _BODY_FONT
        lc.alignment = Alignment(horizontal="left")

        vc = ws.cell(row, 2, value)
        vc.font = _BODY_FONT
        vc.alignment = Alignment(horizontal="right")
        if fmt:
            vc.number_format = fmt
        if highlight:
            vc.fill = _GOOD_FILL
            vc.font = Font(bold=True, size=10, name="Calibri")
        return row + 1

    @staticmethod
    def _auto_width(ws, min_w: int = 12, max_w: int = 55):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)
