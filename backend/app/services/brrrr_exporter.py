"""
BRRRR Deal Proforma — Excel Exporter  (v1.0)

Generates a professional proforma structured around the five BRRRR phases:

  1. Deal Summary        – property + strategy overview + key numbers
  2. Phase 1: Buy        – purchase, discount, financing, cash needed
  3. Phase 2: Rehab      – renovation budget, contingency, holding costs
  4. Phase 3: Rent       – post-rehab rental income, NOI, cap rate
  5. Phase 4: Refinance  – refi terms, cash-out, new payment
  6. Phase 5: Repeat     – capital recycled, equity, infinite ROI check
  7. Assumptions         – all inputs + data sources
"""

from io import BytesIO
from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill,
)
from openpyxl.utils import get_column_letter

from app.schemas.proforma import FinancialProforma


# ── Style tokens ─────────────────────────────────────────────────────────────

_BRAND     = "0A84FF"
_HEADER_BG = "1F4E79"
_GREEN     = "22C55E"
_RED       = "EF4444"
_WHITE     = "FFFFFF"

_HDR_FILL  = PatternFill("solid", fgColor=_HEADER_BG)
_HDR_FONT  = Font(bold=True, color=_WHITE, size=11, name="Calibri")
_SUB_FILL  = PatternFill("solid", fgColor="D6DCE4")
_SUB_FONT  = Font(bold=True, size=10, name="Calibri")
_TOTAL_FNT = Font(bold=True, size=10, name="Calibri")
_BODY_FONT = Font(size=10, name="Calibri")
_GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAD_FILL  = PatternFill("solid", fgColor="FCE4EC")

_CUR = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
_PCT = "0.00%"
_INT = "#,##0"


class BRRRRExcelExporter:
    """Generate a BRRRR-specific proforma workbook."""

    def __init__(self, proforma: FinancialProforma):
        self.d = proforma
        self.wb = Workbook()
        self.bd: Dict[str, Any] = proforma.strategy_breakdown or {}

    # ── public ────────────────────────────────────────────────────────────

    def generate(self) -> BytesIO:
        self.wb.remove(self.wb.active)

        self._tab_deal_summary()
        self._tab_phase1_buy()
        self._tab_phase2_rehab()
        self._tab_phase3_rent()
        self._tab_phase4_refinance()
        self._tab_phase5_repeat()
        self._tab_assumptions()

        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    # ── Tab 1: Deal Summary ───────────────────────────────────────────────

    def _tab_deal_summary(self):
        ws = self.wb.create_sheet("Deal Summary")
        r = 1

        r = self._section_header(ws, r, "BRRRR DEAL SUMMARY")

        prop = self.d.property
        rows = [
            ("Property Address", prop.address),
            ("City / State / Zip", f"{prop.city}, {prop.state} {prop.zip}"),
            ("Property Type", prop.property_type),
            ("Bedrooms / Bathrooms", f"{prop.bedrooms} bd / {prop.bathrooms} ba"),
            ("Square Feet", prop.square_feet),
            ("Year Built", prop.year_built),
            ("Lot Size (sqft)", prop.lot_size),
        ]
        for label, val in rows:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "STRATEGY OVERVIEW")
        total_months = self.bd.get("total_months_to_repeat", 9)
        rows2 = [
            ("Strategy", "BRRRR (Buy, Rehab, Rent, Refinance, Repeat)"),
            ("Objective", "Recycle capital through forced appreciation & refinance"),
            ("Timeline", f"~{total_months} months to complete cycle"),
            ("Financing", "Initial loan + cash-out refinance"),
            ("Risk Profile", "Moderate — renovation & appraisal risk"),
        ]
        for label, val in rows2:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "KEY NUMBERS AT A GLANCE")
        infinite = self.bd.get("infinite_roi_achieved", False)
        glance = [
            ("Purchase Price", self.bd.get("purchase_price", 0), _CUR),
            ("After-Repair Value (ARV)", self.bd.get("arv", 0), _CUR),
            ("Total Renovation", self.bd.get("renovation_budget", 0) + self.bd.get("contingency", 0), _CUR),
            ("Total Cash Invested", self.bd.get("total_cash_invested", 0), _CUR),
            ("Cash Recovered at Refi", self.bd.get("cash_out_at_refinance", 0), _CUR),
            ("Cash Left in Deal", self.bd.get("cash_left_in_deal", 0), _CUR),
            ("Capital Recycled %", self.bd.get("capital_recycled_pct", 0), _PCT),
            ("Infinite ROI?", "YES" if infinite else "NO", None),
        ]
        for label, val, fmt in glance:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        self._auto_width(ws)

    # ── Tab 2: Phase 1 — Buy ─────────────────────────────────────────────

    def _tab_phase1_buy(self):
        ws = self.wb.create_sheet("Phase 1 - Buy")
        r = 1

        r = self._section_header(ws, r, "PHASE 1: BUY AT A DISCOUNT")
        purchase = self.bd.get("purchase_price", 0)
        arv = self.bd.get("arv", 0)
        discount = ((arv - purchase) / arv) if arv else 0

        rows = [
            ("Market Value / ARV", arv, _CUR),
            ("Purchase Discount", discount, _PCT),
            ("Purchase Price", purchase, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        r += 1
        r = self._section_header(ws, r, "INITIAL FINANCING")
        down = self.bd.get("down_payment", 0)
        closing = self.bd.get("closing_costs", 0)
        loan = self.bd.get("initial_loan_amount", 0)
        cash_p1 = self.bd.get("cash_required_phase1", 0)

        rows2 = [
            ("Down Payment", down, _CUR),
            ("Closing Costs", closing, _CUR),
            ("Initial Loan Amount", loan, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Cash Required (Phase 1)").font = _TOTAL_FNT
        c = ws.cell(r, 2, cash_p1)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")

        self._auto_width(ws)

    # ── Tab 3: Phase 2 — Rehab ───────────────────────────────────────────

    def _tab_phase2_rehab(self):
        ws = self.wb.create_sheet("Phase 2 - Rehab")
        r = 1

        reno = self.bd.get("renovation_budget", 0)
        contingency = self.bd.get("contingency", 0)
        holding = self.bd.get("holding_costs", 0)
        cash_p2 = self.bd.get("cash_required_phase2", 0)

        r = self._section_header(ws, r, "PHASE 2: RENOVATE & STABILIZE")
        rows = [
            ("Renovation Budget", reno, _CUR),
            ("Contingency (10%)", contingency, _CUR),
            ("Holding Costs During Rehab", holding, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Cash Required (Phase 2)").font = _TOTAL_FNT
        c = ws.cell(r, 2, cash_p2)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "TOTAL CASH INVESTED (PHASES 1 + 2)")
        cash_p1 = self.bd.get("cash_required_phase1", 0)
        total = self.bd.get("total_cash_invested", 0)
        rows2 = [
            ("Phase 1 (Buy)", cash_p1, _CUR),
            ("Phase 2 (Rehab)", cash_p2, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Cash Invested").font = _TOTAL_FNT
        c = ws.cell(r, 2, total)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")

        self._auto_width(ws)

    # ── Tab 4: Phase 3 — Rent ────────────────────────────────────────────

    def _tab_phase3_rent(self):
        ws = self.wb.create_sheet("Phase 3 - Rent")
        r = 1

        r = self._section_header(ws, r, "PHASE 3: RENT — STABILIZED INCOME")
        rent = self.bd.get("post_rehab_monthly_rent", 0)
        annual_rent = self.bd.get("annual_gross_rent", 0)
        cap = self.bd.get("estimated_cap_rate", 0)
        arv = self.bd.get("arv", 0)

        rows = [
            ("Post-Rehab Monthly Rent", rent, _CUR),
            ("Annual Gross Rent", annual_rent, _CUR),
            ("Estimated Cap Rate", cap, _PCT),
            ("Stabilized Property Value (ARV)", arv, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=(label == "Estimated Cap Rate"))

        r += 1
        r = self._section_header(ws, r, "RENTAL READINESS CHECKLIST")
        checklist = [
            "Renovation complete and inspected",
            "Property listed for rent at market rate",
            "Tenant screened and lease signed",
            "2+ months of stabilized occupancy (for lender seasoning)",
        ]
        for item in checklist:
            ws.cell(r, 1, f"  {item}").font = _BODY_FONT
            r += 1

        self._auto_width(ws)

    # ── Tab 5: Phase 4 — Refinance ───────────────────────────────────────

    def _tab_phase4_refinance(self):
        ws = self.wb.create_sheet("Phase 4 - Refinance")
        r = 1

        r = self._section_header(ws, r, "PHASE 4: CASH-OUT REFINANCE")
        arv = self.bd.get("arv", 0)
        refi_loan = self.bd.get("refinance_loan_amount", 0)
        refi_costs = self.bd.get("refinance_costs", 0)
        payoff = self.bd.get("original_loan_payoff", 0)
        cash_out = self.bd.get("cash_out_at_refinance", 0)
        new_pi = self.bd.get("new_monthly_pi", 0)

        rows = [
            ("Appraised Value (ARV)", arv, _CUR),
            ("New Loan (75% LTV)", refi_loan, _CUR),
            ("− Refinance Closing Costs", refi_costs, _CUR),
            ("− Original Loan Payoff", payoff, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Cash Out at Refinance").font = _TOTAL_FNT
        c = ws.cell(r, 2, cash_out)
        c.number_format = _CUR
        c.font = Font(bold=True, size=12, color=_GREEN if cash_out > 0 else _RED, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "NEW LOAN TERMS")
        rows2 = [
            ("New Loan Amount", refi_loan, _CUR),
            ("New Monthly P&I", new_pi, _CUR),
            ("New Annual Debt Service", new_pi * 12, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        self._auto_width(ws)

    # ── Tab 6: Phase 5 — Repeat ──────────────────────────────────────────

    def _tab_phase5_repeat(self):
        ws = self.wb.create_sheet("Phase 5 - Repeat")
        r = 1

        total_invested = self.bd.get("total_cash_invested", 0)
        cash_out = self.bd.get("cash_out_at_refinance", 0)
        cash_left = self.bd.get("cash_left_in_deal", 0)
        recycled = self.bd.get("capital_recycled_pct", 0)
        equity = self.bd.get("equity_position", 0)
        equity_pct = self.bd.get("equity_pct", 0)
        infinite = self.bd.get("infinite_roi_achieved", False)

        r = self._section_header(ws, r, "PHASE 5: CAPITAL RECOVERY & REPEAT")
        rows = [
            ("Total Cash Invested", total_invested, _CUR),
            ("Cash Recovered at Refinance", cash_out, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Cash Left in Deal").font = _TOTAL_FNT
        c = ws.cell(r, 2, cash_left)
        c.number_format = _CUR
        c.font = Font(bold=True, size=12, color=_GREEN if cash_left <= 0 else _RED, name="Calibri")
        r += 1

        r = self._label_value(ws, r, "Capital Recycled %", recycled, fmt=_PCT, highlight=True)
        r += 1

        # Infinite ROI badge
        ws.cell(r, 1, "Infinite ROI Achieved?").font = _SUB_FONT
        vc = ws.cell(r, 2, "YES — All capital recovered!" if infinite else "NO — Cash still in deal")
        vc.font = Font(bold=True, size=12, name="Calibri")
        vc.fill = _GOOD_FILL if infinite else _WARN_FILL
        r += 2

        r = self._section_header(ws, r, "EQUITY POSITION")
        rows2 = [
            ("Property Value (ARV)", self.bd.get("arv", 0), _CUR),
            ("− Refinance Loan", self.bd.get("refinance_loan_amount", 0), _CUR),
            ("= Equity Position", equity, _CUR),
            ("Equity %", equity_pct, _PCT),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        r += 1
        r = self._section_header(ws, r, "POST-REFINANCE CASH FLOW")
        annual_cf = self.bd.get("post_refi_annual_cash_flow", 0)
        monthly_cf = self.bd.get("post_refi_monthly_cash_flow", 0)
        coc = self.bd.get("post_refi_cash_on_cash", 0)

        rows3 = [
            ("Monthly Cash Flow", monthly_cf, _CUR),
            ("Annual Cash Flow", annual_cf, _CUR),
            ("Cash-on-Cash Return", coc if coc != float("inf") else "Infinite", _PCT if coc != float("inf") else None),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        self._auto_width(ws)

    # ── Tab 7: Assumptions ────────────────────────────────────────────────

    def _tab_assumptions(self):
        ws = self.wb.create_sheet("Assumptions")
        r = 1

        r = self._section_header(ws, r, "BRRRR ASSUMPTIONS")
        assumptions = [
            ("Purchase Discount", "20% off market value"),
            ("Down Payment (Initial)", "20%"),
            ("Initial Interest Rate", "7%"),
            ("Renovation Contingency", "10%"),
            ("Refinance LTV", "75% of ARV"),
            ("Refinance Interest Rate", "7%"),
            ("Refinance Term", "30 years"),
            ("Vacancy Rate", "5%"),
            ("Operating Expense Ratio", "~12% of rent (maintenance + management)"),
        ]
        for label, val in assumptions:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "DATA SOURCES")
        src = self.d.sources
        sources = [
            ("Rent Estimate", src.rent_estimate_source),
            ("Property Value", src.property_value_source),
            ("Tax Data", src.tax_data_source),
            ("Market Data", src.market_data_source),
            ("Data Freshness", src.data_freshness),
        ]
        for label, val in sources:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "REPORT METADATA")
        r = self._label_value(ws, r, "Generated", datetime.now().strftime("%B %d, %Y at %I:%M %p"))
        r = self._label_value(ws, r, "Strategy", "BRRRR (Buy, Rehab, Rent, Refinance, Repeat)")
        r = self._label_value(ws, r, "Engine", "InvestIQ StrategyIQ")

        self._auto_width(ws)

    # ── Helpers ───────────────────────────────────────────────────────────

    def _section_header(self, ws, row: int, title: str) -> int:
        c = ws.cell(row, 1, title)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="left")
        ws.cell(row, 2).fill = _HDR_FILL
        return row + 1

    def _label_value(
        self, ws, row: int, label: str, value: Any,
        *, fmt: str | None = None, highlight: bool = False,
    ) -> int:
        lc = ws.cell(row, 1, label)
        lc.font = _BODY_FONT
        lc.alignment = Alignment(horizontal="left")

        vc = ws.cell(row, 2, value)
        vc.font = _BODY_FONT
        vc.alignment = Alignment(horizontal="right")
        if fmt:
            vc.number_format = fmt
        if highlight:
            vc.fill = _GOOD_FILL
            vc.font = Font(bold=True, size=10, name="Calibri")
        return row + 1

    @staticmethod
    def _auto_width(ws, min_w: int = 12, max_w: int = 55):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)
