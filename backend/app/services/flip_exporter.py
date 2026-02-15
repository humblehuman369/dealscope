"""
Fix & Flip Deal Proforma — Excel Exporter  (v1.0)

Generates a professional, single-transaction financial proforma tailored
for fix-and-flip real-estate deals.  Focuses on:

  1. Deal Summary        – property + strategy overview + key numbers
  2. Acquisition         – 70% Rule, hard money financing, cash needed
  3. Renovation & Holding– rehab budget, holding costs by month
  4. Sale & Profit       – ARV, selling costs, gross/net profit, tax
  5. Deal Metrics        – ROI, annualized ROI, breakeven, 70% rule
  6. Assumptions         – all inputs + data sources
"""

from io import BytesIO
from datetime import datetime
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from app.schemas.proforma import FinancialProforma


# ── Style tokens ─────────────────────────────────────────────────────────────

_BRAND     = "0A84FF"
_HEADER_BG = "1F4E79"
_GREEN     = "22C55E"
_RED       = "EF4444"
_WHITE     = "FFFFFF"

_HDR_FILL  = PatternFill("solid", fgColor=_HEADER_BG)
_HDR_FONT  = Font(bold=True, color=_WHITE, size=11, name="Calibri")
_SUB_FILL  = PatternFill("solid", fgColor="D6DCE4")
_SUB_FONT  = Font(bold=True, size=10, name="Calibri")
_TOTAL_FNT = Font(bold=True, size=10, name="Calibri")
_BODY_FONT = Font(size=10, name="Calibri")
_GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAD_FILL  = PatternFill("solid", fgColor="FCE4EC")

_CUR = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
_PCT = "0.00%"
_NUM = "#,##0.00"
_INT = "#,##0"


class FlipExcelExporter:
    """Generate a fix-and-flip-specific proforma workbook."""

    def __init__(self, proforma: FinancialProforma):
        self.d = proforma
        self.wb = Workbook()
        self.bd: Dict[str, Any] = proforma.strategy_breakdown or {}

    # ── public ────────────────────────────────────────────────────────────

    def generate(self) -> BytesIO:
        self.wb.remove(self.wb.active)

        self._tab_deal_summary()
        self._tab_acquisition()
        self._tab_renovation_holding()
        self._tab_sale_profit()
        self._tab_deal_metrics()
        self._tab_assumptions()

        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    # ── Tab 1: Deal Summary ───────────────────────────────────────────────

    def _tab_deal_summary(self):
        ws = self.wb.create_sheet("Deal Summary")
        r = 1

        r = self._section_header(ws, r, "FIX & FLIP DEAL SUMMARY")

        prop = self.d.property
        rows = [
            ("Property Address", prop.address),
            ("City / State / Zip", f"{prop.city}, {prop.state} {prop.zip}"),
            ("Property Type", prop.property_type),
            ("Bedrooms / Bathrooms", f"{prop.bedrooms} bd / {prop.bathrooms} ba"),
            ("Square Feet", prop.square_feet),
            ("Year Built", prop.year_built),
            ("Lot Size (sqft)", prop.lot_size),
        ]
        for label, val in rows:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "STRATEGY OVERVIEW")
        rows2 = [
            ("Strategy", "Fix & Flip"),
            ("Objective", "Purchase, renovate, and sell for profit"),
            ("Hold Period", f"{self.bd.get('holding_period_months', 6)} months (estimated)"),
            ("Financing", "Hard money / private loan"),
            ("Risk Profile", "Moderate — renovation & market timing risk"),
        ]
        for label, val in rows2:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "KEY NUMBERS AT A GLANCE")
        glance = [
            ("Purchase Price", self.bd.get("purchase_price", 0), _CUR),
            ("After-Repair Value (ARV)", self.bd.get("arv", 0), _CUR),
            ("Total Renovation", self.bd.get("total_renovation", 0), _CUR),
            ("Total Project Cost", self.bd.get("total_project_cost", 0), _CUR),
            ("Net Profit (Before Tax)", self.bd.get("net_profit_before_tax", 0), _CUR),
            ("ROI", self.bd.get("roi", 0), _PCT),
            ("Annualized ROI", self.bd.get("annualized_roi", 0), _PCT),
        ]
        for label, val, fmt in glance:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        self._auto_width(ws)

    # ── Tab 2: Acquisition ────────────────────────────────────────────────

    def _tab_acquisition(self):
        ws = self.wb.create_sheet("Acquisition")
        r = 1

        arv = self.bd.get("arv", 0)
        renovation_budget = self.bd.get("renovation_budget", 0)
        seventy_max = self.bd.get("seventy_pct_max_price", 0)
        purchase = self.bd.get("purchase_price", 0)
        meets_rule = self.bd.get("meets_70_rule", False)

        r = self._section_header(ws, r, "70% RULE — MAXIMUM ALLOWABLE OFFER")
        rows = [
            ("After-Repair Value (ARV)", arv, _CUR),
            ("× 70%", arv * 0.70, _CUR),
            ("− Renovation Budget", renovation_budget, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Maximum Allowable Offer (MAO)").font = _TOTAL_FNT
        c = ws.cell(r, 2, seventy_max)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")
        r += 1

        ws.cell(r, 1, "Your Purchase Price").font = _BODY_FONT
        ws.cell(r, 2, purchase).number_format = _CUR
        r += 1

        ws.cell(r, 1, "Passes 70% Rule?").font = _SUB_FONT
        vc = ws.cell(r, 2, "YES" if meets_rule else "NO")
        vc.font = Font(bold=True, size=11, name="Calibri")
        vc.fill = _GOOD_FILL if meets_rule else _BAD_FILL
        r += 2

        r = self._section_header(ws, r, "HARD MONEY FINANCING")
        hm_loan = self.bd.get("hard_money_loan", 0)
        down = self.bd.get("down_payment", 0)
        closing = self.bd.get("closing_costs", 0)
        inspection = self.bd.get("inspection_costs", 0)
        total_acq = self.bd.get("total_acquisition_cash", 0)

        rows2 = [
            ("Purchase Price", purchase, _CUR),
            ("Hard Money Loan", hm_loan, _CUR),
            ("Down Payment", down, _CUR),
            ("Closing Costs", closing, _CUR),
            ("Inspection Costs", inspection, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Cash for Acquisition").font = _TOTAL_FNT
        c = ws.cell(r, 2, total_acq)
        c.number_format = _CUR
        c.font = _TOTAL_FNT
        r += 1

        self._auto_width(ws)

    # ── Tab 3: Renovation & Holding ───────────────────────────────────────

    def _tab_renovation_holding(self):
        ws = self.wb.create_sheet("Renovation & Holding")
        r = 1

        renovation_budget = self.bd.get("renovation_budget", 0)
        contingency = self.bd.get("contingency", 0)
        total_reno = self.bd.get("total_renovation", 0)

        r = self._section_header(ws, r, "RENOVATION BUDGET")
        rows = [
            ("Renovation Budget", renovation_budget, _CUR),
            ("Contingency", contingency, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Renovation Cost").font = _TOTAL_FNT
        c = ws.cell(r, 2, total_reno)
        c.number_format = _CUR
        c.font = _TOTAL_FNT
        r += 2

        r = self._section_header(ws, r, "MONTHLY HOLDING COSTS")
        hm_interest = self.bd.get("hard_money_interest", 0)
        taxes = self.bd.get("property_taxes", 0)
        insurance = self.bd.get("insurance", 0)
        utilities = self.bd.get("utilities", 0)
        security = self.bd.get("security_maintenance", 0)
        total_holding = self.bd.get("total_holding_costs", 0)

        rows2 = [
            ("Hard Money Interest", hm_interest, _CUR),
            ("Property Taxes (prorated)", taxes, _CUR),
            ("Insurance (prorated)", insurance, _CUR),
            ("Utilities", utilities, _CUR),
            ("Security & Maintenance", security, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Holding Costs").font = _TOTAL_FNT
        c = ws.cell(r, 2, total_holding)
        c.number_format = _CUR
        c.font = _TOTAL_FNT
        r += 2

        r = self._section_header(ws, r, "TOTAL INVESTMENT REQUIRED")
        total_acq = self.bd.get("total_acquisition_cash", 0)
        total_cash = self.bd.get("total_cash_required", 0)

        rows3 = [
            ("Acquisition Cash", total_acq, _CUR),
            ("Renovation Cost", total_reno, _CUR),
            ("Holding Costs", total_holding, _CUR),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Cash Required").font = _TOTAL_FNT
        c = ws.cell(r, 2, total_cash)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")

        self._auto_width(ws)

    # ── Tab 4: Sale & Profit ──────────────────────────────────────────────

    def _tab_sale_profit(self):
        ws = self.wb.create_sheet("Sale & Profit")
        r = 1

        arv = self.bd.get("arv", 0)
        commission = self.bd.get("realtor_commission", 0)
        seller_closing = self.bd.get("seller_closing_costs", 0)
        total_sell = self.bd.get("total_selling_costs", 0)
        net_proceeds = self.bd.get("net_sale_proceeds", 0)

        r = self._section_header(ws, r, "SALE PROCEEDS")
        rows = [
            ("After-Repair Value (Sale Price)", arv, _CUR),
            ("− Realtor Commission (6%)", commission, _CUR),
            ("− Seller Closing Costs (2%)", seller_closing, _CUR),
            ("= Total Selling Costs", total_sell, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Net Sale Proceeds").font = _TOTAL_FNT
        c = ws.cell(r, 2, net_proceeds)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "PROFIT ANALYSIS")
        total_project = self.bd.get("total_project_cost", 0)
        gross = self.bd.get("gross_profit", 0)
        hm_loan = self.bd.get("hard_money_loan", 0)
        total_cash = self.bd.get("total_cash_required", 0)
        net_before = self.bd.get("net_profit_before_tax", 0)
        cap_gains = self.bd.get("capital_gains_tax", 0)
        net_after = self.bd.get("net_profit_after_tax", 0)

        rows2 = [
            ("Total Project Cost", total_project, _CUR),
            ("Gross Profit (ARV − Project Cost)", gross, _CUR),
            ("Hard Money Loan Payoff", hm_loan, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Net Profit (Before Tax)").font = _TOTAL_FNT
        c = ws.cell(r, 2, net_before)
        c.number_format = _CUR
        c.font = Font(bold=True, size=12, color=_GREEN if net_before > 0 else _RED, name="Calibri")
        r += 1

        r = self._label_value(ws, r, "Capital Gains Tax (15%)", cap_gains, fmt=_CUR)

        ws.cell(r, 1, "= Net Profit (After Tax)").font = _TOTAL_FNT
        c = ws.cell(r, 2, net_after)
        c.number_format = _CUR
        c.font = Font(bold=True, size=12, color=_GREEN if net_after > 0 else _RED, name="Calibri")

        self._auto_width(ws)

    # ── Tab 5: Deal Metrics ───────────────────────────────────────────────

    def _tab_deal_metrics(self):
        ws = self.wb.create_sheet("Deal Metrics")
        r = 1

        roi = self.bd.get("roi", 0)
        ann_roi = self.bd.get("annualized_roi", 0)
        margin = self.bd.get("profit_margin", 0)
        total_cash = self.bd.get("total_cash_required", 0)
        breakeven = self.bd.get("minimum_sale_for_breakeven", 0)
        arv = self.bd.get("arv", 0)
        meets_rule = self.bd.get("meets_70_rule", False)

        r = self._section_header(ws, r, "RETURN METRICS")
        rows = [
            ("ROI (on Cash Invested)", roi, _PCT),
            ("Annualized ROI", ann_roi, _PCT),
            ("Profit Margin (Net / ARV)", margin, _PCT),
            ("Total Cash Required", total_cash, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        r += 1
        r = self._section_header(ws, r, "BREAK-EVEN ANALYSIS")
        cushion = arv - breakeven if arv and breakeven else 0
        rows2 = [
            ("Minimum Sale Price to Break Even", breakeven, _CUR),
            ("After-Repair Value (ARV)", arv, _CUR),
            ("Cushion Above Breakeven", cushion, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        r += 1
        r = self._section_header(ws, r, "70% RULE CHECK")
        ws.cell(r, 1, "Passes 70% Rule?").font = _SUB_FONT
        vc = ws.cell(r, 2, "YES — Purchase below MAO" if meets_rule else "NO — Purchase exceeds MAO")
        vc.font = Font(bold=True, size=11, name="Calibri")
        vc.fill = _GOOD_FILL if meets_rule else _BAD_FILL

        self._auto_width(ws)

    # ── Tab 6: Assumptions ────────────────────────────────────────────────

    def _tab_assumptions(self):
        ws = self.wb.create_sheet("Assumptions")
        r = 1

        r = self._section_header(ws, r, "FIX & FLIP ASSUMPTIONS")
        assumptions = [
            ("Purchase Discount", "20% off market value"),
            ("Hard Money LTV", "70%"),
            ("Hard Money Rate", "12%"),
            ("Holding Period", f"{self.bd.get('holding_period_months', 6)} months"),
            ("Selling Costs", "8% (6% commission + 2% closing)"),
            ("Capital Gains Tax Rate", "15%"),
            ("Contingency", "10% of renovation budget"),
        ]
        for label, val in assumptions:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "DATA SOURCES")
        src = self.d.sources
        sources = [
            ("Rent Estimate", src.rent_estimate_source),
            ("Property Value", src.property_value_source),
            ("Tax Data", src.tax_data_source),
            ("Market Data", src.market_data_source),
            ("Data Freshness", src.data_freshness),
        ]
        for label, val in sources:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "REPORT METADATA")
        r = self._label_value(ws, r, "Generated", datetime.now().strftime("%B %d, %Y at %I:%M %p"))
        r = self._label_value(ws, r, "Strategy", "Fix & Flip")
        r = self._label_value(ws, r, "Engine", "DealGapIQ StrategyIQ")

        self._auto_width(ws)

    # ── Helpers ───────────────────────────────────────────────────────────

    def _section_header(self, ws, row: int, title: str) -> int:
        c = ws.cell(row, 1, title)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="left")
        ws.cell(row, 2).fill = _HDR_FILL
        return row + 1

    def _label_value(
        self, ws, row: int, label: str, value: Any,
        *, fmt: str | None = None, highlight: bool = False,
    ) -> int:
        lc = ws.cell(row, 1, label)
        lc.font = _BODY_FONT
        lc.alignment = Alignment(horizontal="left")

        vc = ws.cell(row, 2, value)
        vc.font = _BODY_FONT
        vc.alignment = Alignment(horizontal="right")
        if fmt:
            vc.number_format = fmt
        if highlight:
            vc.fill = _GOOD_FILL
            vc.font = Font(bold=True, size=10, name="Calibri")
        return row + 1

    @staticmethod
    def _auto_width(ws, min_w: int = 12, max_w: int = 55):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)
