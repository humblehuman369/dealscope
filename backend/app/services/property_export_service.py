"""
Property data export service — Excel report with combined property data (RentCast + AXESSO) and Calculated sheets.
"""
import json
import logging
from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

logger = logging.getLogger(__name__)

SOURCE_RENTCAST = "RentCast"
SOURCE_AXESSO = "AXESSO"


def _flatten_to_rows(obj: Any, prefix: str = "") -> List[Tuple[str, Any]]:
    """Convert nested dict/list to (key_path, value) rows. Keys are dot-separated."""
    rows: List[Tuple[str, Any]] = []
    if obj is None:
        rows.append((prefix or "value", None))
        return rows
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else k
            if isinstance(v, (dict, list)) and not (isinstance(v, dict) and v and all(isinstance(x, (str, int, float, bool, type(None))) for x in v.values())):
                rows.extend(_flatten_to_rows(v, key))
            else:
                if isinstance(v, (dict, list)):
                    v = json.dumps(v) if v else ""
                rows.append((key, v))
        return rows
    if isinstance(obj, list):
        for i, item in enumerate(obj):
            key = f"{prefix}[{i}]"
            if isinstance(item, (dict, list)):
                rows.extend(_flatten_to_rows(item, key))
            else:
                rows.append((key, item))
        return rows
    rows.append((prefix or "value", obj))
    return rows


def _build_rentcast_rows(raw: Dict[str, Any], key_prefix: str = SOURCE_RENTCAST) -> List[Tuple[str, Any]]:
    """Build flat key/value rows for RentCast (skip internal keys). Keys prefixed with key_prefix when set."""
    out: List[Tuple[str, Any]] = []
    for k, v in raw.items():
        if k.startswith("_"):
            continue
        base = f"{key_prefix}.{k}" if key_prefix else k
        out.extend(_flatten_to_rows(v, base) if isinstance(v, (dict, list)) else [(base, v)])
    return out


def _build_axesso_rows(raw: Dict[str, Any], key_prefix: str = SOURCE_AXESSO) -> List[Tuple[str, Any]]:
    """Build flat key/value rows for AXESSO. Keys prefixed with key_prefix when set."""
    out: List[Tuple[str, Any]] = []
    for k, v in raw.items():
        base = f"{key_prefix}.{k}" if key_prefix else k
        if isinstance(v, (dict, list)):
            out.extend(_flatten_to_rows(v, base))
        else:
            out.append((base, v))
    return out


def _write_key_value_sheet(ws: Any, rows: List[Tuple[str, Any]]) -> None:
    """Write a single sheet with Key, Value columns."""
    ws.append(["Key", "Value"])
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)
    for key, value in rows:
        ws.append([key, value])
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 45


def generate_rentcast_only_excel(export_data: Dict[str, Any]) -> bytes:
    """
    Generate a single Excel workbook with RentCast data only (for auditing).
    One sheet: Key, Value with all RentCast API data (property, value_estimate, rent_estimate, market_statistics).
    """
    raw_rentcast = export_data.get("raw_rentcast") or {}
    rows = _build_rentcast_rows(raw_rentcast, key_prefix="")
    wb = Workbook()
    ws = wb.active
    ws.title = "RentCast Data"
    _write_key_value_sheet(ws, rows)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_axesso_only_excel(export_data: Dict[str, Any]) -> bytes:
    """
    Generate a single Excel workbook with AXESSO data only (for auditing).
    One sheet: Key, Value with all AXESSO/Zillow API data (search_by_address, property_details).
    """
    raw_axesso = export_data.get("raw_axesso") or {}
    rows = _build_axesso_rows(raw_axesso, key_prefix="")
    wb = Workbook()
    ws = wb.active
    ws.title = "AXESSO Data"
    _write_key_value_sheet(ws, rows)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_property_data_report_excel(export_data: Dict[str, Any]) -> bytes:
    """
    Generate one report file with two sheets: data received from RentCast and data received from AXESSO.
    Sheet 1 "RentCast" — all data we receive from RentCast for this property.
    Sheet 2 "AXESSO" — all data we receive from AXESSO/Zillow for this property.
    """
    raw_rentcast = export_data.get("raw_rentcast") or {}
    raw_axesso = export_data.get("raw_axesso") or {}
    rentcast_rows = _build_rentcast_rows(raw_rentcast, key_prefix="")
    axesso_rows = _build_axesso_rows(raw_axesso, key_prefix="")

    wb = Workbook()
    ws_rc = wb.active
    ws_rc.title = "RentCast"
    _write_key_value_sheet(ws_rc, rentcast_rows)

    ws_ax = wb.create_sheet("AXESSO", 1)
    _write_key_value_sheet(ws_ax, axesso_rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_property_data_sheet(ws: Any, rentcast_rows: List[Tuple[str, Any]], axesso_rows: List[Tuple[str, Any]]) -> None:
    """Write one sheet with all RentCast then all AXESSO data (Key, Value)."""
    ws.append(["Key", "Value"])
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)
    row_num = 2
    if rentcast_rows:
        ws.cell(row=row_num, column=1, value=f"--- {SOURCE_RENTCAST} (all data used for calculations) ---")
        ws.cell(row=row_num, column=1).font = Font(bold=True, italic=True)
        row_num += 1
        for key, value in rentcast_rows:
            ws.cell(row=row_num, column=1, value=key)
            ws.cell(row=row_num, column=2, value=value)
            row_num += 1
        row_num += 1
    if axesso_rows:
        ws.cell(row=row_num, column=1, value=f"--- {SOURCE_AXESSO} (all data used for calculations) ---")
        ws.cell(row=row_num, column=1).font = Font(bold=True, italic=True)
        row_num += 1
        for key, value in axesso_rows:
            ws.cell(row=row_num, column=1, value=key)
            ws.cell(row=row_num, column=2, value=value)
            row_num += 1
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 45


def _build_calculated_rows(verdict: Dict[str, Any], property_data: Dict[str, Any]) -> List[Tuple[str, str]]:
    """Build Calculated sheet: Verdict and Strategy metrics as Label, Value rows."""
    rows: List[Tuple[str, str]] = []
    v = verdict

    def fmt(x: Any) -> str:
        if x is None:
            return ""
        if isinstance(x, float):
            return f"{x:,.2f}" if abs(x) < 1e10 else str(x)
        return str(x)

    # Core verdict
    rows.append(("Deal Score", fmt(v.get("dealScore") or v.get("deal_score"))))
    rows.append(("Deal Verdict", v.get("dealVerdict") or v.get("deal_verdict") or ""))
    rows.append(("Verdict Description", v.get("verdictDescription") or v.get("verdict_description") or ""))
    rows.append(("List Price", fmt(v.get("listPrice") or v.get("list_price"))))
    rows.append(("Income Value", fmt(v.get("incomeValue") or v.get("income_value"))))
    rows.append(("Purchase Price (Target Buy)", fmt(v.get("purchasePrice") or v.get("purchase_price"))))
    rows.append(("Discount %", fmt(v.get("discountPercent") or v.get("discount_percent"))))
    rows.append(("Deal Gap %", fmt(v.get("dealGapPercent") or v.get("deal_gap_percent") or v.get("opportunity_factors", {}).get("deal_gap"))))
    rows.append(("Wholesale MAO", fmt(v.get("wholesaleMao") or v.get("wholesale_mao"))))

    # Opportunity factors
    of = v.get("opportunityFactors") or v.get("opportunity_factors") or {}
    rows.append(("Seller Motivation Score", fmt(of.get("motivation"))))
    rows.append(("Motivation Label", of.get("motivationLabel") or of.get("motivation_label") or ""))
    rows.append(("Buyer Market", of.get("buyerMarket") or of.get("buyer_market") or ""))

    # Return factors (primary strategy)
    rf = v.get("returnFactors") or v.get("return_factors") or {}
    rows.append(("Cap Rate %", fmt(rf.get("capRate") or rf.get("cap_rate"))))
    rows.append(("Cash on Cash %", fmt(rf.get("cashOnCash") or rf.get("cash_on_cash"))))
    rows.append(("DSCR", fmt(rf.get("dscr"))))

    # Strategies table
    strategies = v.get("strategies") or []
    rows.append(("", ""))
    rows.append(("Strategy", "Metric | Score | Cap Rate | CoC | DSCR | Cash Flow"))
    for s in strategies:
        name = s.get("name") or s.get("id") or ""
        metric = s.get("metric") or ""
        score = s.get("score")
        cap = s.get("cap_rate")
        coc = s.get("cash_on_cash")
        dscr = s.get("dscr")
        acf = s.get("annual_cash_flow") or s.get("annualCashFlow")
        mcf = s.get("monthly_cash_flow") or s.get("monthlyCashFlow")
        row_val = f"{metric} | {score} | {fmt(cap)} | {fmt(coc)} | {fmt(dscr)} | ACF: {fmt(acf)} MCF: {fmt(mcf)}"
        rows.append((name, row_val))

    return rows


def generate_property_data_excel(export_data: Dict[str, Any]) -> bytes:
    """
    Generate a 2-sheet Excel workbook from get_property_export_data() result.

    Sheets:
    1. Property Data (RentCast & AXESSO) — all RentCast and all AXESSO API data
       used to calculate the values in sheet 2 (property, value_estimate, rent_estimate,
       market_statistics, search_by_address, property_details).
    2. Calculated (Verdict & Strategy) — Verdict and Strategy metrics (deal score,
       list price, income value, target buy, strategies).
    """
    wb = Workbook()
    raw_rentcast = export_data.get("raw_rentcast") or {}
    raw_axesso = export_data.get("raw_axesso") or {}
    verdict = export_data.get("verdict") or {}
    property_data = export_data.get("property") or {}

    # Sheet 1: All property data from RentCast + AXESSO (source data for calculations)
    ws_prop = wb.active
    ws_prop.title = "Property Data (RentCast & AXESSO)"
    rentcast_rows = _build_rentcast_rows(raw_rentcast)
    axesso_rows = _build_axesso_rows(raw_axesso)
    _write_property_data_sheet(ws_prop, rentcast_rows, axesso_rows)

    # Sheet 2: Calculated (Verdict + Strategy)
    ws_calc = wb.create_sheet("Calculated (Verdict & Strategy)", 1)
    calc_rows = _build_calculated_rows(verdict, property_data)
    ws_calc.append(["Label", "Value"])
    ws_calc.cell(row=1, column=1).font = Font(bold=True)
    ws_calc.cell(row=1, column=2).font = Font(bold=True)
    for label, value in calc_rows:
        ws_calc.append([label, value])
    ws_calc.column_dimensions["A"].width = 35
    ws_calc.column_dimensions["B"].width = 55

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
