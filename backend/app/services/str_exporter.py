"""
Short-Term Rental (STR) Proforma — Excel Exporter  (v1.0)

Generates a professional proforma tailored for Airbnb / VRBO style
short-term rental investments:

  1. Property Summary     – property details, ADR, occupancy, overview
  2. Revenue Analysis     – gross revenue, platform fees, net revenue
  3. Operating Expenses   – full expense breakdown
  4. Cash Flow & Returns  – NOI, debt service, cash flow, key metrics
  5. Seasonality Analysis – peak / shoulder / off-season breakdown
  6. Assumptions          – STR-specific inputs + data sources
"""

from io import BytesIO
from datetime import datetime
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill,
)
from openpyxl.utils import get_column_letter

from app.schemas.proforma import FinancialProforma


# ── Style tokens ─────────────────────────────────────────────────────────────

_BRAND     = "0A84FF"
_HEADER_BG = "1F4E79"
_GREEN     = "22C55E"
_RED       = "EF4444"
_WHITE     = "FFFFFF"

_HDR_FILL  = PatternFill("solid", fgColor=_HEADER_BG)
_HDR_FONT  = Font(bold=True, color=_WHITE, size=11, name="Calibri")
_SUB_FILL  = PatternFill("solid", fgColor="D6DCE4")
_SUB_FONT  = Font(bold=True, size=10, name="Calibri")
_TOTAL_FNT = Font(bold=True, size=10, name="Calibri")
_BODY_FONT = Font(size=10, name="Calibri")
_GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAD_FILL  = PatternFill("solid", fgColor="FCE4EC")

_CUR = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
_PCT = "0.00%"
_INT = "#,##0"


class STRExcelExporter:
    """Generate a short-term rental proforma workbook."""

    def __init__(self, proforma: FinancialProforma):
        self.d = proforma
        self.wb = Workbook()
        self.bd: Dict[str, Any] = proforma.strategy_breakdown or {}

    # ── public ────────────────────────────────────────────────────────────

    def generate(self) -> BytesIO:
        self.wb.remove(self.wb.active)

        self._tab_property_summary()
        self._tab_revenue()
        self._tab_expenses()
        self._tab_cash_flow()
        self._tab_seasonality()
        self._tab_assumptions()

        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    # ── Tab 1: Property Summary ───────────────────────────────────────────

    def _tab_property_summary(self):
        ws = self.wb.create_sheet("Property Summary")
        r = 1

        r = self._section_header(ws, r, "SHORT-TERM RENTAL SUMMARY")

        prop = self.d.property
        rows = [
            ("Property Address", prop.address),
            ("City / State / Zip", f"{prop.city}, {prop.state} {prop.zip}"),
            ("Property Type", prop.property_type),
            ("Bedrooms / Bathrooms", f"{prop.bedrooms} bd / {prop.bathrooms} ba"),
            ("Square Feet", prop.square_feet),
            ("Year Built", prop.year_built),
            ("Lot Size (sqft)", prop.lot_size),
        ]
        for label, val in rows:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "STRATEGY OVERVIEW")
        rows2 = [
            ("Strategy", "Short-Term Rental (Airbnb / VRBO)"),
            ("Objective", "Maximize revenue through nightly rentals"),
            ("Revenue Model", "ADR × Occupancy × 365 nights"),
            ("Financing", "Conventional or DSCR loan"),
            ("Risk Profile", "Higher revenue potential, higher operating costs"),
        ]
        for label, val in rows2:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "KEY NUMBERS AT A GLANCE")
        glance = [
            ("Average Daily Rate (ADR)", self.bd.get("average_daily_rate", 0), _CUR),
            ("Occupancy Rate", self.bd.get("occupancy_rate", 0), _PCT),
            ("Annual Gross Revenue", self.bd.get("total_gross_revenue", 0), _CUR),
            ("Annual NOI", self.bd.get("noi", 0), _CUR),
            ("Monthly Cash Flow", self.bd.get("monthly_cash_flow", 0), _CUR),
            ("Cap Rate", self.bd.get("cap_rate", 0), _PCT),
            ("Cash-on-Cash Return", self.bd.get("cash_on_cash_return", 0), _PCT),
        ]
        for label, val, fmt in glance:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        self._auto_width(ws)

    # ── Tab 2: Revenue Analysis ───────────────────────────────────────────

    def _tab_revenue(self):
        ws = self.wb.create_sheet("Revenue Analysis")
        r = 1

        adr = self.bd.get("average_daily_rate", 0)
        occ = self.bd.get("occupancy_rate", 0)
        nights = self.bd.get("nights_occupied", 0)
        bookings = self.bd.get("num_bookings", 0)
        rental_rev = self.bd.get("rental_revenue", 0)
        clean_rev = self.bd.get("cleaning_fee_revenue", 0)
        gross = self.bd.get("total_gross_revenue", 0)
        rpan = self.bd.get("revenue_per_available_night", 0)

        r = self._section_header(ws, r, "BOOKING METRICS")
        rows = [
            ("Average Daily Rate (ADR)", adr, _CUR),
            ("Occupancy Rate", occ, _PCT),
            ("Nights Occupied / Year", nights, _INT),
            ("Est. Number of Bookings", round(bookings), _INT),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        r += 1
        r = self._section_header(ws, r, "REVENUE BREAKDOWN")
        rows2 = [
            ("Nightly Rental Revenue", rental_rev, _CUR),
            ("Cleaning Fee Revenue", clean_rev, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Gross Revenue").font = _TOTAL_FNT
        c = ws.cell(r, 2, gross)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "EFFICIENCY METRICS")
        monthly_rev = gross / 12 if gross else 0
        rows3 = [
            ("Revenue per Available Night (RevPAN)", rpan, _CUR),
            ("Average Monthly Revenue", monthly_rev, _CUR),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        self._auto_width(ws)

    # ── Tab 3: Operating Expenses ─────────────────────────────────────────

    def _tab_expenses(self):
        ws = self.wb.create_sheet("Operating Expenses")
        r = 1

        r = self._section_header(ws, r, "STR OPERATING EXPENSES (ANNUAL)")
        expenses = [
            ("Platform Fees (Airbnb/VRBO)", self.bd.get("platform_fees", 0), _CUR),
            ("STR Management", self.bd.get("str_management", 0), _CUR),
            ("Cleaning Costs", self.bd.get("cleaning_costs", 0), _CUR),
            ("Supplies", self.bd.get("supplies", 0), _CUR),
            ("Utilities", self.bd.get("utilities", 0), _CUR),
            ("Property Taxes", self.bd.get("property_taxes", 0), _CUR),
            ("Insurance (STR Policy)", self.bd.get("insurance", 0), _CUR),
            ("Maintenance", self.bd.get("maintenance", 0), _CUR),
            ("Landscaping", self.bd.get("landscaping", 0), _CUR),
            ("Pest Control", self.bd.get("pest_control", 0), _CUR),
        ]
        for label, val, fmt in expenses:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        total = self.bd.get("total_operating_expenses", 0)
        ws.cell(r, 1, "= Total Operating Expenses").font = _TOTAL_FNT
        c = ws.cell(r, 2, total)
        c.number_format = _CUR
        c.font = _TOTAL_FNT
        r += 2

        r = self._section_header(ws, r, "EXPENSE RATIO")
        gross = self.bd.get("total_gross_revenue", 0)
        ratio = total / gross if gross else 0
        rows2 = [
            ("Total Gross Revenue", gross, _CUR),
            ("Total Expenses", total, _CUR),
            ("Expense Ratio", ratio, _PCT),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        self._auto_width(ws)

    # ── Tab 4: Cash Flow & Returns ────────────────────────────────────────

    def _tab_cash_flow(self):
        ws = self.wb.create_sheet("Cash Flow & Returns")
        r = 1

        noi = self.bd.get("noi", 0)
        debt = self.bd.get("annual_debt_service", 0)
        annual_cf = self.bd.get("annual_cash_flow", 0)
        monthly_cf = self.bd.get("monthly_cash_flow", 0)

        r = self._section_header(ws, r, "CASH FLOW ANALYSIS")
        rows = [
            ("Net Operating Income (NOI)", noi, _CUR),
            ("Annual Debt Service", debt, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Annual Cash Flow").font = _TOTAL_FNT
        c = ws.cell(r, 2, annual_cf)
        c.number_format = _CUR
        c.font = Font(bold=True, size=12, color=_GREEN if annual_cf > 0 else _RED, name="Calibri")
        r += 1

        ws.cell(r, 1, "= Monthly Cash Flow").font = _TOTAL_FNT
        c = ws.cell(r, 2, monthly_cf)
        c.number_format = _CUR
        c.font = Font(bold=True, size=12, color=_GREEN if monthly_cf > 0 else _RED, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "RETURN METRICS")
        cap = self.bd.get("cap_rate", 0)
        coc = self.bd.get("cash_on_cash_return", 0)
        dscr = self.bd.get("dscr", 0)
        be_occ = self.bd.get("break_even_occupancy", 0)

        rows2 = [
            ("Cap Rate", cap, _PCT),
            ("Cash-on-Cash Return", coc, _PCT),
            ("DSCR", dscr, "0.00x" if isinstance(dscr, (int, float)) else None),
            ("Break-Even Occupancy", be_occ, _PCT),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        r += 1
        r = self._section_header(ws, r, "INVESTMENT SUMMARY")
        down = self.bd.get("down_payment", 0)
        closing = self.bd.get("closing_costs", 0)
        furniture = self.bd.get("furniture_setup", 0)
        total_cash = self.bd.get("total_cash_required", 0)

        rows3 = [
            ("Down Payment", down, _CUR),
            ("Closing Costs", closing, _CUR),
            ("Furniture & Setup", furniture, _CUR),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Cash Required").font = _TOTAL_FNT
        c = ws.cell(r, 2, total_cash)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")

        self._auto_width(ws)

    # ── Tab 5: Seasonality Analysis ───────────────────────────────────────

    def _tab_seasonality(self):
        ws = self.wb.create_sheet("Seasonality Analysis")
        r = 1

        r = self._section_header(ws, r, "SEASONAL REVENUE PROJECTIONS")
        seasons = self.bd.get("seasonality_analysis", [])

        if seasons:
            # Table header
            headers = ["Season", "Months", "Occupancy", "ADR", "Est. Revenue"]
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(r, col_idx, h)
                c.font = _SUB_FONT
                c.fill = _SUB_FILL
            r += 1

            total_rev = 0
            for s in seasons:
                ws.cell(r, 1, s.get("season", "")).font = _BODY_FONT
                ws.cell(r, 2, s.get("months", 0)).font = _BODY_FONT
                c_occ = ws.cell(r, 3, s.get("occupancy", 0))
                c_occ.number_format = _PCT
                c_occ.font = _BODY_FONT
                c_adr = ws.cell(r, 4, s.get("adr", 0))
                c_adr.number_format = _CUR
                c_adr.font = _BODY_FONT
                rev = s.get("revenue", 0)
                total_rev += rev
                c_rev = ws.cell(r, 5, rev)
                c_rev.number_format = _CUR
                c_rev.font = _BODY_FONT
                r += 1

            # Total row
            ws.cell(r, 1, "TOTAL").font = _TOTAL_FNT
            c = ws.cell(r, 5, total_rev)
            c.number_format = _CUR
            c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")
            r += 2
        else:
            r = self._label_value(ws, r, "Seasonality data", "Not available")
            r += 1

        r = self._section_header(ws, r, "OCCUPANCY SENSITIVITY")
        adr = self.bd.get("average_daily_rate", 0)
        total_expenses = self.bd.get("total_operating_expenses", 0)
        debt = self.bd.get("annual_debt_service", 0)

        headers2 = ["Occupancy", "Gross Revenue", "Est. Cash Flow"]
        for col_idx, h in enumerate(headers2, 1):
            c = ws.cell(r, col_idx, h)
            c.font = _SUB_FONT
            c.fill = _SUB_FILL
        r += 1

        for occ_pct in [0.50, 0.60, 0.70, 0.80, 0.90]:
            rev = adr * 365 * occ_pct
            est_cf = rev - total_expenses - debt
            ws.cell(r, 1, occ_pct).number_format = _PCT
            ws.cell(r, 2, rev).number_format = _CUR
            cf_cell = ws.cell(r, 3, est_cf)
            cf_cell.number_format = _CUR
            cf_cell.font = Font(size=10, color=_GREEN if est_cf > 0 else _RED, name="Calibri")
            r += 1

        self._auto_width(ws)

    # ── Tab 6: Assumptions ────────────────────────────────────────────────

    def _tab_assumptions(self):
        ws = self.wb.create_sheet("Assumptions")
        r = 1

        r = self._section_header(ws, r, "SHORT-TERM RENTAL ASSUMPTIONS")
        assumptions = [
            ("Average Daily Rate", f"${self.bd.get('average_daily_rate', 0):,.0f}"),
            ("Occupancy Rate", f"{self.bd.get('occupancy_rate', 0) * 100:.0f}%"),
            ("Average Stay Length", f"{self.bd.get('avg_length_of_stay', 3)} nights"),
            ("Platform Fee Rate", "15%"),
            ("STR Management Rate", "20%"),
            ("Furniture & Setup", f"${self.bd.get('furniture_setup', 5000):,.0f}"),
            ("STR Insurance Premium", "Higher than LTR (est. 0.75% of value)"),
            ("Down Payment", "20%"),
        ]
        for label, val in assumptions:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "DATA SOURCES")
        src = self.d.sources
        sources = [
            ("Rent Estimate", src.rent_estimate_source),
            ("Property Value", src.property_value_source),
            ("Tax Data", src.tax_data_source),
            ("Market Data", src.market_data_source),
            ("Data Freshness", src.data_freshness),
        ]
        for label, val in sources:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "REPORT METADATA")
        r = self._label_value(ws, r, "Generated", datetime.now().strftime("%B %d, %Y at %I:%M %p"))
        r = self._label_value(ws, r, "Strategy", "Short-Term Rental (Airbnb / VRBO)")
        r = self._label_value(ws, r, "Engine", "RealVestIQ StrategyIQ")

        self._auto_width(ws)

    # ── Helpers ───────────────────────────────────────────────────────────

    def _section_header(self, ws, row: int, title: str) -> int:
        c = ws.cell(row, 1, title)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="left")
        for col in range(2, 6):  # wider header for seasonality table
            ws.cell(row, col).fill = _HDR_FILL
        return row + 1

    def _label_value(
        self, ws, row: int, label: str, value: Any,
        *, fmt: str | None = None, highlight: bool = False,
    ) -> int:
        lc = ws.cell(row, 1, label)
        lc.font = _BODY_FONT
        lc.alignment = Alignment(horizontal="left")

        vc = ws.cell(row, 2, value)
        vc.font = _BODY_FONT
        vc.alignment = Alignment(horizontal="right")
        if fmt:
            vc.number_format = fmt
        if highlight:
            vc.fill = _GOOD_FILL
            vc.font = Font(bold=True, size=10, name="Calibri")
        return row + 1

    @staticmethod
    def _auto_width(ws, min_w: int = 12, max_w: int = 55):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)
