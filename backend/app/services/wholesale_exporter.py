"""
Wholesale Deal Proforma — Excel Exporter  (v1.0)

Generates a professional, single-transaction financial proforma tailored
for wholesale real-estate deals.  No multi-year projections, no loan
amortization, no depreciation — just the numbers a wholesaler cares about:

  1. Deal Summary        – property + strategy overview
  2. Deal Structure      – MAO, contract price, assignment fee, end-buyer price
  3. Costs & Profit      – earnest money, marketing, net profit, ROI
  4. Buyer Analysis      – rent estimate, AMV, rehab cost → buyer's numbers
  5. Deal Viability       – spread, viability grade, deals-to-income targets
  6. Assumptions          – all inputs + data sources
"""

from io import BytesIO
from datetime import datetime
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from app.schemas.proforma import FinancialProforma


# ── Style tokens ─────────────────────────────────────────────────────────────

_BRAND     = "0A84FF"   # DealGapIQ blue
_DARK_BG   = "111827"   # near-black
_CARD_BG   = "1A2332"   # dark card
_HEADER_BG = "1F4E79"
_GREEN     = "22C55E"
_RED       = "EF4444"
_AMBER     = "F59E0B"
_GRAY      = "9CA3AF"
_WHITE     = "FFFFFF"
_LIGHT_ROW = "F8FAFC"

_HDR_FILL  = PatternFill("solid", fgColor=_HEADER_BG)
_HDR_FONT  = Font(bold=True, color=_WHITE, size=11, name="Calibri")
_SUB_FILL  = PatternFill("solid", fgColor="D6DCE4")
_SUB_FONT  = Font(bold=True, size=10, name="Calibri")
_TOTAL_FNT = Font(bold=True, size=10, name="Calibri")
_BODY_FONT = Font(size=10, name="Calibri")
_GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAD_FILL  = PatternFill("solid", fgColor="FCE4EC")
_THIN      = Border(*(Side(style="thin"),) * 4)

_CUR   = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
_PCT   = "0.00%"
_NUM   = "#,##0.00"
_INT   = "#,##0"


class WholesaleExcelExporter:
    """Generate a wholesale-specific proforma workbook."""

    def __init__(
        self,
        proforma: FinancialProforma,
        *,
        rent_estimate: Optional[float] = None,
        amv: Optional[float] = None,
        wholesale_fee: Optional[float] = None,
    ):
        self.d = proforma
        self.wb = Workbook()
        # Wholesale breakdown dict (from strategy calculator)
        self.ws_data: Dict[str, Any] = proforma.strategy_breakdown or {}

        # User-supplied overrides take priority over calculated values
        self._rent = rent_estimate or proforma.income.monthly_rent
        self._amv = amv or proforma.acquisition.list_price or proforma.acquisition.purchase_price
        self._fee = wholesale_fee or self.ws_data.get("assignment_fee", 15_000)

    # ── public ────────────────────────────────────────────────────────────

    def generate(self) -> BytesIO:
        """Build workbook → return BytesIO buffer."""
        self.wb.remove(self.wb.active)

        self._tab_deal_summary()
        self._tab_deal_structure()
        self._tab_costs_profit()
        self._tab_buyer_analysis()
        self._tab_deal_viability()
        self._tab_assumptions()

        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf

    # ── Tab 1: Deal Summary ───────────────────────────────────────────────

    def _tab_deal_summary(self):
        ws = self.wb.create_sheet("Deal Summary")
        r = 1

        # Title
        r = self._section_header(ws, r, "WHOLESALE DEAL SUMMARY")

        prop = self.d.property
        rows = [
            ("Property Address", prop.address),
            ("City / State / Zip", f"{prop.city}, {prop.state} {prop.zip}"),
            ("Property Type", prop.property_type),
            ("Bedrooms / Bathrooms", f"{prop.bedrooms} bd / {prop.bathrooms} ba"),
            ("Square Feet", prop.square_feet),
            ("Year Built", prop.year_built),
            ("Lot Size (sqft)", prop.lot_size),
        ]
        for label, val in rows:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "STRATEGY OVERVIEW")
        rows2 = [
            ("Strategy", "Wholesale (Assignment)"),
            ("Objective", "Assign contract to end buyer for a fee"),
            ("Hold Period", "None — no ownership transfer"),
            ("Financing Required", "None"),
            ("Risk Profile", "Low capital at risk"),
        ]
        for label, val in rows2:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "KEY NUMBERS AT A GLANCE")
        glance = [
            ("Est. After-Market Value (AMV)", self._amv, _CUR),
            ("Est. Monthly Rent", self._rent, _CUR),
            ("Assignment / Wholesale Fee", self._fee, _CUR),
            ("Contract Price (MAO)", self.ws_data.get("contract_price", 0), _CUR),
            ("End-Buyer Price", self.ws_data.get("end_buyer_price", 0), _CUR),
            ("Net Profit", self.ws_data.get("net_profit", 0), _CUR),
            ("ROI on Cash at Risk", self.ws_data.get("roi", 0), _PCT),
        ]
        for label, val, fmt in glance:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        self._auto_width(ws)

    # ── Tab 2: Deal Structure ─────────────────────────────────────────────

    def _tab_deal_structure(self):
        ws = self.wb.create_sheet("Deal Structure")
        r = 1

        r = self._section_header(ws, r, "70% RULE — MAXIMUM ALLOWABLE OFFER")
        arv = self.ws_data.get("arv", self._amv)
        rehab = self.ws_data.get("estimated_rehab", 0)
        mao = self.ws_data.get("seventy_pct_max_offer", 0)

        rows = [
            ("After-Repair Value (ARV)", arv, _CUR),
            ("× 70%", arv * 0.70, _CUR),
            ("− Estimated Rehab Costs", rehab, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        # MAO total line
        ws.cell(r, 1, "= Maximum Allowable Offer (MAO)").font = _TOTAL_FNT
        c = ws.cell(r, 2, mao)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "TRANSACTION STRUCTURE")
        contract = self.ws_data.get("contract_price", mao)
        fee = self._fee
        end_buyer = self.ws_data.get("end_buyer_price", contract + fee)

        rows2 = [
            ("Your Contract Price (with Seller)", contract, _CUR),
            ("+ Your Assignment Fee", fee, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= End-Buyer Purchase Price").font = _TOTAL_FNT
        c = ws.cell(r, 2, end_buyer)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_BRAND, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "PRICE COMPARISON")
        discount = ((self._amv - contract) / self._amv * 100) if self._amv else 0
        buyer_disc = ((self._amv - end_buyer) / self._amv * 100) if self._amv else 0
        rows3 = [
            ("Est. Market Value (AMV)", self._amv, _CUR),
            ("Your Contract Price", contract, _CUR),
            ("Your Discount from AMV", f"{discount:.1f}%", None),
            ("End-Buyer Price", end_buyer, _CUR),
            ("Buyer's Discount from AMV", f"{buyer_disc:.1f}%", None),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        self._auto_width(ws)

    # ── Tab 3: Costs & Profit ─────────────────────────────────────────────

    def _tab_costs_profit(self):
        ws = self.wb.create_sheet("Costs & Profit")
        r = 1

        earnest = self.ws_data.get("earnest_money", 1000)
        marketing = self.ws_data.get("marketing_costs", 500)
        total_risk = self.ws_data.get("total_cash_at_risk", earnest + marketing)

        r = self._section_header(ws, r, "YOUR COSTS (CASH AT RISK)")
        rows = [
            ("Earnest Money Deposit", earnest, _CUR),
            ("Marketing & Lead Gen", marketing, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Total Cash at Risk").font = _TOTAL_FNT
        c = ws.cell(r, 2, total_risk)
        c.number_format = _CUR
        c.font = _TOTAL_FNT
        r += 2

        r = self._section_header(ws, r, "PROFIT ANALYSIS")
        gross = self.ws_data.get("gross_profit", self._fee)
        net = self.ws_data.get("net_profit", gross - marketing)
        roi = self.ws_data.get("roi", 0)
        ann_roi = self.ws_data.get("annualized_roi", 0)
        hourly = self.ws_data.get("effective_hourly_rate", 0)
        hours = self.ws_data.get("time_investment_hours", 50)

        rows2 = [
            ("Gross Profit (Assignment Fee)", gross, _CUR),
            ("− Marketing Costs", marketing, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Net Profit").font = _TOTAL_FNT
        c = ws.cell(r, 2, net)
        c.number_format = _CUR
        c.font = Font(bold=True, size=12, color=_GREEN, name="Calibri")
        r += 2

        r = self._section_header(ws, r, "RETURN METRICS")
        rows3 = [
            ("ROI on Cash at Risk", roi, _PCT),
            ("Annualized ROI", ann_roi, _PCT),
            ("Est. Time Investment (hours)", hours, _INT),
            ("Effective Hourly Rate", hourly, _CUR),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt, highlight=True)

        self._auto_width(ws)

    # ── Tab 4: Buyer Analysis ─────────────────────────────────────────────

    def _tab_buyer_analysis(self):
        """What the end buyer sees — helps you pitch the deal."""
        ws = self.wb.create_sheet("Buyer Analysis")
        r = 1

        end_buyer = self.ws_data.get("end_buyer_price", 0)
        rehab = self.ws_data.get("estimated_rehab", 0)
        arv = self.ws_data.get("arv", self._amv)

        r = self._section_header(ws, r, "END-BUYER ACQUISITION")
        rows = [
            ("End-Buyer Purchase Price", end_buyer, _CUR),
            ("Estimated Rehab Costs", rehab, _CUR),
            ("Total Buyer Investment", end_buyer + rehab, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)
        r += 1

        r = self._section_header(ws, r, "BUYER'S UPSIDE (ARV METHOD)")
        buyer_equity = arv - end_buyer - rehab
        buyer_roi = (buyer_equity / (end_buyer + rehab) * 100) if (end_buyer + rehab) else 0
        rows2 = [
            ("After-Repair Value (ARV)", arv, _CUR),
            ("− Buyer's Total Investment", end_buyer + rehab, _CUR),
        ]
        for label, val, fmt in rows2:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        ws.cell(r, 1, "= Buyer's Potential Equity").font = _TOTAL_FNT
        c = ws.cell(r, 2, buyer_equity)
        c.number_format = _CUR
        c.font = Font(bold=True, size=11, color=_GREEN if buyer_equity > 0 else _RED, name="Calibri")
        r += 1
        r = self._label_value(ws, r, "Buyer's ROI Potential", f"{buyer_roi:.1f}%")
        r += 1

        r = self._section_header(ws, r, "BUYER'S RENTAL INCOME POTENTIAL")
        annual_rent = self._rent * 12
        # Estimate buyer's expenses at 45% of rent (rule of thumb)
        est_expenses = annual_rent * 0.45
        est_noi = annual_rent - est_expenses
        # Estimate buyer's mortgage on end_buyer_price (80% LTV, 7%, 30yr)
        loan = end_buyer * 0.80
        monthly_pmt = (loan * (0.07 / 12)) / (1 - (1 + 0.07 / 12) ** -360) if loan > 0 else 0
        annual_debt = monthly_pmt * 12
        est_cash_flow = est_noi - annual_debt
        cap_rate = (est_noi / end_buyer * 100) if end_buyer else 0

        rows3 = [
            ("Est. Monthly Rent", self._rent, _CUR),
            ("Est. Annual Gross Rent", annual_rent, _CUR),
            ("Est. Operating Expenses (45%)", est_expenses, _CUR),
            ("Est. NOI", est_noi, _CUR),
            ("Est. Annual Debt Service (80% LTV, 7%)", annual_debt, _CUR),
            ("Est. Annual Cash Flow", est_cash_flow, _CUR),
            ("Est. Cap Rate", f"{cap_rate:.2f}%", None),
        ]
        for label, val, fmt in rows3:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        self._auto_width(ws)

    # ── Tab 5: Deal Viability ─────────────────────────────────────────────

    def _tab_deal_viability(self):
        ws = self.wb.create_sheet("Deal Viability")
        r = 1

        spread = self.ws_data.get("spread_available", 0)
        viability = self.ws_data.get("deal_viability", "Unknown")
        days = self.ws_data.get("timeline_days", 45)

        r = self._section_header(ws, r, "SPREAD & VIABILITY ANALYSIS")
        rows = [
            ("Total Spread Available", spread, _CUR),
            ("− Your Assignment Fee", self._fee, _CUR),
            ("= Buyer's Remaining Spread", spread - self._fee, _CUR),
        ]
        for label, val, fmt in rows:
            r = self._label_value(ws, r, label, val, fmt=fmt)
        r += 1

        # Viability badge
        ws.cell(r, 1, "Deal Viability Grade").font = _SUB_FONT
        vc = ws.cell(r, 2, viability)
        vc.font = Font(bold=True, size=12, name="Calibri")
        fill_map = {"Strong": _GOOD_FILL, "Moderate": _WARN_FILL, "Tight": _BAD_FILL, "Not Viable": _BAD_FILL}
        vc.fill = fill_map.get(viability, _WARN_FILL)
        r += 2

        r = self._section_header(ws, r, "TIMELINE")
        r = self._label_value(ws, r, "Est. Days to Close", days, fmt=_INT)
        r += 1

        r = self._section_header(ws, r, "INCOME SCALING — DEALS TO TARGET")
        net = self.ws_data.get("net_profit", 0)
        d50 = self.ws_data.get("deals_needed_50k", 0)
        d100 = self.ws_data.get("deals_needed_100k", 0)

        targets = [
            ("Net Profit per Deal", net, _CUR),
            ("Deals for $50,000 / year", round(d50) if d50 != float('inf') else "N/A", _INT if d50 != float('inf') else None),
            ("Deals for $100,000 / year", round(d100) if d100 != float('inf') else "N/A", _INT if d100 != float('inf') else None),
        ]
        for label, val, fmt in targets:
            r = self._label_value(ws, r, label, val, fmt=fmt)

        r += 1
        r = self._section_header(ws, r, "QUICK REFERENCE — VIABILITY CRITERIA")
        criteria = [
            ("Strong", "Buyer spread ≥ Fee + $20,000"),
            ("Moderate", "Buyer spread ≥ Fee + $10,000"),
            ("Tight", "Buyer spread ≥ Fee"),
            ("Not Viable", "Buyer spread < Fee"),
        ]
        for grade, desc in criteria:
            ws.cell(r, 1, grade).font = Font(bold=True, size=10, name="Calibri")
            ws.cell(r, 2, desc).font = _BODY_FONT
            r += 1

        self._auto_width(ws)

    # ── Tab 6: Assumptions ────────────────────────────────────────────────

    def _tab_assumptions(self):
        ws = self.wb.create_sheet("Assumptions")
        r = 1

        r = self._section_header(ws, r, "WHOLESALE DEAL ASSUMPTIONS")
        assumptions = [
            ("ARV Discount %", "30% (70% Rule)"),
            ("Assignment Fee", f"${self._fee:,.0f}"),
            ("Earnest Money Deposit", f"${self.ws_data.get('earnest_money', 1000):,.0f}"),
            ("Marketing Budget", f"${self.ws_data.get('marketing_costs', 500):,.0f}"),
            ("Days to Close", str(self.ws_data.get("timeline_days", 45))),
            ("Time Investment", f"{self.ws_data.get('time_investment_hours', 50)} hours"),
        ]
        for label, val in assumptions:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "BUYER ANALYSIS ASSUMPTIONS")
        buyer_assumptions = [
            ("Buyer LTV", "80%"),
            ("Buyer Interest Rate", "7.0%"),
            ("Buyer Loan Term", "30 years"),
            ("Operating Expense Ratio", "45% of gross rent"),
        ]
        for label, val in buyer_assumptions:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "DATA SOURCES")
        src = self.d.sources
        sources = [
            ("Rent Estimate", src.rent_estimate_source),
            ("Property Value", src.property_value_source),
            ("Tax Data", src.tax_data_source),
            ("Market Data", src.market_data_source),
            ("Data Freshness", src.data_freshness),
        ]
        for label, val in sources:
            r = self._label_value(ws, r, label, val)

        r += 1
        r = self._section_header(ws, r, "REPORT METADATA")
        r = self._label_value(ws, r, "Generated", datetime.now().strftime("%B %d, %Y at %I:%M %p"))
        r = self._label_value(ws, r, "Strategy", "Wholesale (Assignment)")
        r = self._label_value(ws, r, "Engine", "DealGapIQ StrategyIQ")

        self._auto_width(ws)

    # ── Helpers ───────────────────────────────────────────────────────────

    def _section_header(self, ws, row: int, title: str) -> int:
        """Write a full-width section header and return next row."""
        c = ws.cell(row, 1, title)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="left")
        ws.cell(row, 2).fill = _HDR_FILL
        return row + 1

    def _label_value(
        self, ws, row: int, label: str, value: Any,
        *, fmt: str | None = None, highlight: bool = False,
    ) -> int:
        """Write a label–value pair and return next row."""
        lc = ws.cell(row, 1, label)
        lc.font = _BODY_FONT
        lc.alignment = Alignment(horizontal="left")

        vc = ws.cell(row, 2, value)
        vc.font = _BODY_FONT
        vc.alignment = Alignment(horizontal="right")
        if fmt:
            vc.number_format = fmt
        if highlight:
            vc.fill = _GOOD_FILL
            vc.font = Font(bold=True, size=10, name="Calibri")
        return row + 1

    @staticmethod
    def _auto_width(ws, min_w: int = 12, max_w: int = 55):
        """Auto-fit column widths."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)
