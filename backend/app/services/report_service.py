"""
Report Generation Service for creating PDF and Excel property reports.

Generates comprehensive financial statements including:
- NOI (Net Operating Income) Cash Flow Statement
- DSCR (Debt Service Coverage Ratio) Qualification Analysis
- 10-Year Pro Forma Projections
- Amortization Schedule
- Sensitivity Analysis
- Strategy Comparison Matrix
"""

import logging
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from io import BytesIO
import json
import math

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

logger = logging.getLogger(__name__)


class ReportService:
    """Service for generating property analysis reports."""
    
    # Brand colors
    BRAND_BLUE = "0465F2"
    BRAND_TEAL = "00E5FF"
    NAVY = "07172E"
    LIGHT_GRAY = "F3F4F6"
    MEDIUM_GRAY = "E5E7EB"
    DARK_GRAY = "6B7280"
    GREEN = "22C55E"
    RED = "EF4444"
    YELLOW = "F59E0B"
    
    # Strategy colors
    STRATEGY_COLORS = {
        "ltr": "0465F2",      # Blue
        "str": "8B5CF6",      # Purple
        "brrrr": "F97316",    # Orange
        "flip": "EC4899",     # Pink
        "house_hack": "22C55E",  # Green
        "wholesale": "22D3EE",   # Cyan
    }
    
    # Common styles
    @staticmethod
    def _get_currency_format():
        return '"$"#,##0'
    
    @staticmethod
    def _get_percent_format():
        return '0.00%'
    
    @staticmethod
    def _get_number_format():
        return '#,##0.00'
    
    def generate_excel_report(
        self,
        property_data: Dict[str, Any],
        analytics_data: Dict[str, Any],
        assumptions: Optional[Dict[str, Any]] = None,
        include_sensitivity: bool = False,
        strategy: Optional[str] = None,
        price_target: Optional[str] = None,
    ) -> bytes:
        """
        Generate a comprehensive Excel report for a property.
        
        Args:
            property_data: Property details (address, beds, baths, etc.)
            analytics_data: Calculated analytics for all strategies
            assumptions: Assumptions used in calculations
            include_sensitivity: Whether to include sensitivity analysis
            strategy: Active strategy (ltr, str, brrrr, flip, house_hack, wholesale)
            price_target: Active price target (breakeven, targetBuy, wholesale)
        
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Determine which strategy data to use for primary analysis
        active_strategy = strategy or "ltr"
        active_price_target = price_target or "targetBuy"
        
        # Create sheets - Financial Statements First
        self._create_summary_sheet(wb, property_data, analytics_data, active_strategy)
        
        # Create appropriate financial statement based on strategy
        if active_strategy in ["ltr", "str"]:
            # Rental strategies get Cash Flow Statement
            self._create_cash_flow_statement_sheet(wb, property_data, analytics_data, assumptions)
            self._create_dscr_analysis_sheet(wb, property_data, analytics_data, assumptions)
        elif active_strategy in ["flip", "wholesale"]:
            # Exit strategies get Deal Analysis instead of Cash Flow
            self._create_deal_analysis_sheet(wb, property_data, analytics_data, assumptions, active_strategy)
        elif active_strategy == "brrrr":
            # BRRRR gets both Cash Flow and Refinance analysis
            self._create_cash_flow_statement_sheet(wb, property_data, analytics_data, assumptions)
            self._create_dscr_analysis_sheet(wb, property_data, analytics_data, assumptions)
        elif active_strategy == "house_hack":
            # House Hack gets Housing Cost Analysis
            self._create_cash_flow_statement_sheet(wb, property_data, analytics_data, assumptions)
        
        self._create_pro_forma_sheet(wb, property_data, analytics_data, assumptions)
        self._create_amortization_sheet(wb, property_data, analytics_data, assumptions)
        
        if include_sensitivity:
            self._create_sensitivity_sheet(wb, property_data, analytics_data, assumptions)
        
        # Strategy-specific sheets - put active strategy first
        strategy_order = [active_strategy] + [s for s in ["ltr", "str", "brrrr", "flip", "house_hack", "wholesale"] if s != active_strategy]
        
        for strat in strategy_order:
            if strat == "ltr":
                self._create_ltr_sheet(wb, property_data, analytics_data.get("ltr"))
            elif strat == "str":
                self._create_str_sheet(wb, property_data, analytics_data.get("str"))
            elif strat == "brrrr":
                self._create_brrrr_sheet(wb, property_data, analytics_data.get("brrrr"))
            elif strat == "flip":
                self._create_flip_sheet(wb, property_data, analytics_data.get("flip"))
            elif strat == "house_hack":
                self._create_house_hack_sheet(wb, property_data, analytics_data.get("house_hack"))
            elif strat == "wholesale":
                self._create_wholesale_sheet(wb, property_data, analytics_data.get("wholesale"))
        
        if assumptions:
            self._create_assumptions_sheet(wb, assumptions)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _create_deal_analysis_sheet(
        self, 
        wb: Workbook, 
        property_data: Dict, 
        analytics: Dict,
        assumptions: Optional[Dict] = None,
        strategy: str = "flip"
    ):
        """Create Deal Analysis sheet for flip/wholesale strategies."""
        ws = wb.create_sheet("Deal Analysis")
        
        # Styles
        title_font = Font(bold=True, size=16, color=self.BRAND_BLUE)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color=self.NAVY, end_color=self.NAVY, fill_type="solid")
        
        # Get strategy data
        data = analytics.get(strategy, {})
        
        # Title
        ws.merge_cells('A1:D1')
        strategy_name = "Fix & Flip Analysis" if strategy == "flip" else "Wholesale Analysis"
        ws['A1'] = strategy_name.upper()
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Property Address
        address = property_data.get("address", {})
        ws.merge_cells('A2:D2')
        ws['A2'] = address.get("full_address", "Property Address")
        ws['A2'].font = Font(italic=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        row = 4
        if strategy == "flip":
            # Flip analysis
            ws[f'A{row}'] = "ACQUISITION"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            metrics = [
                ("Purchase Price", data.get("purchase_price", 0), "currency"),
                ("Closing Costs", data.get("closing_costs", 0), "currency"),
                ("Rehab Budget", data.get("rehab_budget", 0), "currency"),
            ]
            for label, value, fmt in metrics:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = f"${value:,.0f}" if fmt == "currency" else value
                row += 1
            
            row += 1
            ws[f'A{row}'] = "SALE"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            metrics = [
                ("ARV", data.get("arv", 0), "currency"),
                ("Selling Costs", data.get("selling_costs", 0), "currency"),
                ("Net Profit", data.get("net_profit_before_tax", 0), "currency"),
                ("ROI", data.get("roi", 0), "percent"),
            ]
            for label, value, fmt in metrics:
                ws[f'A{row}'] = label
                if fmt == "currency":
                    ws[f'B{row}'] = f"${value:,.0f}"
                else:
                    ws[f'B{row}'] = f"{value*100:.1f}%"
                row += 1
        else:
            # Wholesale analysis
            ws[f'A{row}'] = "PROPERTY ANALYSIS"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            metrics = [
                ("ARV", data.get("arv", 0), "currency"),
                ("Estimated Repairs", data.get("estimated_repairs", 0), "currency"),
                ("Max Allowable Offer", data.get("max_allowable_offer", 0), "currency"),
            ]
            for label, value, fmt in metrics:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = f"${value:,.0f}"
                row += 1
            
            row += 1
            ws[f'A{row}'] = "ASSIGNMENT"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            
            metrics = [
                ("Contract Price", data.get("your_offer_price", 0), "currency"),
                ("Assignment Fee", data.get("assignment_fee", 0), "currency"),
                ("Net Profit", data.get("net_profit", 0), "currency"),
                ("ROI", data.get("roi", 0), "percent"),
            ]
            for label, value, fmt in metrics:
                ws[f'A{row}'] = label
                if fmt == "currency":
                    ws[f'B{row}'] = f"${value:,.0f}"
                else:
                    ws[f'B{row}'] = f"{value*100:.1f}%"
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def generate_financial_statements_report(
        self,
        property_data: Dict[str, Any],
        analytics_data: Dict[str, Any],
        assumptions: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        Generate a focused financial statements report (NOI, DSCR, Pro Forma).
        
        This is a streamlined report specifically for lender/investor presentations.
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Financial statement sheets only
        self._create_executive_summary_sheet(wb, property_data, analytics_data, assumptions)
        self._create_cash_flow_statement_sheet(wb, property_data, analytics_data, assumptions)
        self._create_dscr_analysis_sheet(wb, property_data, analytics_data, assumptions)
        self._create_pro_forma_sheet(wb, property_data, analytics_data, assumptions)
        self._create_amortization_sheet(wb, property_data, analytics_data, assumptions)
        self._create_sensitivity_sheet(wb, property_data, analytics_data, assumptions)
        
        if assumptions:
            self._create_assumptions_sheet(wb, assumptions)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    # =========================================================================
    # EXECUTIVE SUMMARY SHEET
    # =========================================================================
    
    def _create_executive_summary_sheet(
        self, 
        wb: Workbook, 
        property_data: Dict, 
        analytics: Dict,
        assumptions: Optional[Dict] = None
    ):
        """Create executive summary for lender/investor presentation."""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Styles
        title_font = Font(bold=True, size=20, color=self.BRAND_BLUE)
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color=self.BRAND_BLUE, end_color=self.BRAND_BLUE, fill_type="solid")
        section_font = Font(bold=True, size=12, color=self.NAVY)
        value_font = Font(bold=True, size=11)
        
        thin_border = Border(
            left=Side(style='thin', color=self.MEDIUM_GRAY),
            right=Side(style='thin', color=self.MEDIUM_GRAY),
            top=Side(style='thin', color=self.MEDIUM_GRAY),
            bottom=Side(style='thin', color=self.MEDIUM_GRAY)
        )
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "INVESTMENT PROPERTY FINANCIAL ANALYSIS"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Subtitle
        ws.merge_cells('A2:F2')
        address = property_data.get("address", {})
        ws['A2'] = address.get("full_address", "Property Address")
        ws['A2'].font = Font(size=14, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Report date
        ws['A3'] = f"Report Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A3'].font = Font(size=10, color=self.DARK_GRAY)
        
        # Get LTR data for primary analysis
        ltr = analytics.get("ltr", {})
        
        # Property Overview Section
        row = 5
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "PROPERTY OVERVIEW"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        details = property_data.get("details", {})
        valuations = property_data.get("valuations", {})
        
        overview_data = [
            ("Property Type", details.get("property_type", "Single Family")),
            ("Bedrooms / Bathrooms", f"{details.get('bedrooms', 'N/A')} / {details.get('bathrooms', 'N/A')}"),
            ("Square Footage", f"{details.get('square_footage', 0):,}" if details.get('square_footage') else "N/A"),
            ("Year Built", details.get("year_built", "N/A")),
            ("Current Value (AVM)", f"${valuations.get('current_value_avm', 0):,.0f}" if valuations.get('current_value_avm') else "N/A"),
        ]
        
        for label, value in overview_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].border = thin_border
            row += 1
        
        # Key Investment Metrics Section
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "KEY INVESTMENT METRICS"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        
        # Calculate key metrics
        purchase_price = assumptions.get("purchase_price", 400000) if assumptions else 400000
        noi = ltr.get("noi", 0)
        cap_rate = ltr.get("cap_rate", 0)
        coc = ltr.get("cash_on_cash_return", 0)
        dscr = ltr.get("dscr", 0)
        monthly_cf = ltr.get("monthly_cash_flow", 0)
        
        metrics = [
            ("Purchase Price", f"${purchase_price:,.0f}", ""),
            ("Net Operating Income (NOI)", f"${noi:,.0f}", "/year"),
            ("Cap Rate", f"{cap_rate*100:.2f}%", ""),
            ("Cash-on-Cash Return", f"{coc*100:.2f}%", ""),
            ("DSCR", f"{dscr:.2f}x", ""),
            ("Monthly Cash Flow", f"${monthly_cf:,.0f}", "/month"),
        ]
        
        for label, value, suffix in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value + suffix
            ws[f'B{row}'].font = value_font
            
            # Color code based on metric
            if "DSCR" in label:
                if dscr >= 1.25:
                    ws[f'B{row}'].font = Font(bold=True, color=self.GREEN)
                elif dscr >= 1.0:
                    ws[f'B{row}'].font = Font(bold=True, color=self.YELLOW)
                else:
                    ws[f'B{row}'].font = Font(bold=True, color=self.RED)
            elif "Cash Flow" in label:
                if monthly_cf > 0:
                    ws[f'B{row}'].font = Font(bold=True, color=self.GREEN)
                else:
                    ws[f'B{row}'].font = Font(bold=True, color=self.RED)
            
            ws[f'B{row}'].border = thin_border
            row += 1
        
        # Investment Summary Box
        row += 2
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "INVESTMENT VERDICT"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        
        # Determine verdict
        if dscr >= 1.25 and monthly_cf > 200:
            verdict = "STRONG INVESTMENT"
            verdict_color = self.GREEN
            verdict_detail = "Property meets lender DSCR requirements with positive cash flow."
        elif dscr >= 1.0 and monthly_cf > 0:
            verdict = "MODERATE INVESTMENT"
            verdict_color = self.YELLOW
            verdict_detail = "Positive cash flow but DSCR below typical lender threshold of 1.25x."
        else:
            verdict = "NEEDS ANALYSIS"
            verdict_color = self.RED
            verdict_detail = "Consider negotiating price or increasing rent to improve returns."
        
        ws[f'A{row}'] = verdict
        ws[f'A{row}'].font = Font(bold=True, size=16, color=verdict_color)
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = verdict_detail
        ws[f'A{row}'].font = Font(italic=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    # =========================================================================
    # CASH FLOW STATEMENT (NOI) SHEET
    # =========================================================================
    
    def _create_cash_flow_statement_sheet(
        self, 
        wb: Workbook, 
        property_data: Dict, 
        analytics: Dict,
        assumptions: Optional[Dict] = None
    ):
        """Create professional NOI/Cash Flow Statement sheet."""
        ws = wb.create_sheet("Cash Flow Statement")
        
        # Styles
        title_font = Font(bold=True, size=16, color=self.BRAND_BLUE)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color=self.NAVY, end_color=self.NAVY, fill_type="solid")
        section_fill = PatternFill(start_color=self.LIGHT_GRAY, end_color=self.LIGHT_GRAY, fill_type="solid")
        subtotal_fill = PatternFill(start_color=self.MEDIUM_GRAY, end_color=self.MEDIUM_GRAY, fill_type="solid")
        noi_fill = PatternFill(start_color=self.BRAND_BLUE, end_color=self.BRAND_BLUE, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_bottom = Border(bottom=Side(style='medium'))
        
        # Get LTR data
        ltr = analytics.get("ltr", {})
        
        # Extract or calculate values
        monthly_rent = ltr.get("monthly_rent", 2500)
        annual_gross_rent = ltr.get("annual_gross_rent", monthly_rent * 12)
        other_income = 0
        vacancy_rate = 0.05
        vacancy_loss = ltr.get("vacancy_loss", annual_gross_rent * vacancy_rate)
        
        # Operating Expenses
        property_taxes = ltr.get("property_taxes", 4800)
        insurance = ltr.get("insurance", 1800)
        property_management = ltr.get("property_management", annual_gross_rent * 0.10)
        maintenance = ltr.get("maintenance", annual_gross_rent * 0.05)
        utilities = ltr.get("utilities", 900)
        landscaping = ltr.get("landscaping", 500)
        pest_control = ltr.get("pest_control", 200)
        hoa_fees = ltr.get("hoa_fees", 0)
        
        total_operating = ltr.get("total_operating_expenses", 
            property_taxes + insurance + property_management + maintenance + 
            utilities + landscaping + pest_control + hoa_fees)
        
        noi = ltr.get("noi", annual_gross_rent - vacancy_loss - total_operating)
        annual_debt_service = ltr.get("annual_debt_service", 0)
        cash_flow = ltr.get("annual_cash_flow", noi - annual_debt_service)
        
        # Title
        row = 1
        ws.merge_cells('A1:D1')
        ws['A1'] = "PROPERTY CASH FLOW STATEMENT"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Property Address
        row = 2
        address = property_data.get("address", {})
        ws.merge_cells('A2:D2')
        ws['A2'] = address.get("full_address", "Property Address")
        ws['A2'].font = Font(italic=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Column Headers
        row = 4
        ws['A4'] = "Description"
        ws['B4'] = "Monthly"
        ws['C4'] = "Annual"
        ws['D4'] = "% of Gross"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}4'].font = header_font
            ws[f'{col}4'].fill = header_fill
            ws[f'{col}4'].alignment = Alignment(horizontal='center')
            ws[f'{col}4'].border = thin_border
        
        # === INCOME SECTION ===
        row = 5
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "GROSS POTENTIAL INCOME"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = section_fill
        
        row += 1
        income_items = [
            ("  Scheduled Rent", monthly_rent, annual_gross_rent),
            ("  Other Income (parking, storage, etc.)", other_income / 12, other_income),
        ]
        
        for label, monthly, annual in income_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = monthly
            ws[f'B{row}'].number_format = self._get_currency_format()
            ws[f'C{row}'] = annual
            ws[f'C{row}'].number_format = self._get_currency_format()
            ws[f'D{row}'] = annual / annual_gross_rent if annual_gross_rent > 0 else 0
            ws[f'D{row}'].number_format = self._get_percent_format()
            row += 1
        
        # Gross Potential Income Total
        gpi = annual_gross_rent + other_income
        ws[f'A{row}'] = "Gross Potential Income"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = gpi / 12
        ws[f'B{row}'].number_format = self._get_currency_format()
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = gpi
        ws[f'C{row}'].number_format = self._get_currency_format()
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = 1.0
        ws[f'D{row}'].number_format = self._get_percent_format()
        ws[f'A{row}'].fill = subtotal_fill
        ws[f'B{row}'].fill = subtotal_fill
        ws[f'C{row}'].fill = subtotal_fill
        ws[f'D{row}'].fill = subtotal_fill
        row += 1
        
        # === VACANCY & CREDIT LOSS ===
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "VACANCY & CREDIT LOSS"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = section_fill
        
        row += 1
        ws[f'A{row}'] = f"  Vacancy Loss ({vacancy_rate*100:.0f}%)"
        ws[f'B{row}'] = -vacancy_loss / 12
        ws[f'B{row}'].number_format = self._get_currency_format()
        ws[f'B{row}'].font = Font(color=self.RED)
        ws[f'C{row}'] = -vacancy_loss
        ws[f'C{row}'].number_format = self._get_currency_format()
        ws[f'C{row}'].font = Font(color=self.RED)
        ws[f'D{row}'] = -vacancy_rate
        ws[f'D{row}'].number_format = self._get_percent_format()
        row += 1
        
        # Effective Gross Income
        egi = gpi - vacancy_loss
        ws[f'A{row}'] = "Effective Gross Income (EGI)"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = egi / 12
        ws[f'B{row}'].number_format = self._get_currency_format()
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = egi
        ws[f'C{row}'].number_format = self._get_currency_format()
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = egi / gpi if gpi > 0 else 0
        ws[f'D{row}'].number_format = self._get_percent_format()
        ws[f'A{row}'].fill = subtotal_fill
        ws[f'B{row}'].fill = subtotal_fill
        ws[f'C{row}'].fill = subtotal_fill
        ws[f'D{row}'].fill = subtotal_fill
        row += 1
        
        # === OPERATING EXPENSES ===
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "OPERATING EXPENSES"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = section_fill
        
        row += 1
        expenses = [
            ("  Property Taxes", property_taxes),
            ("  Insurance", insurance),
            ("  Property Management", property_management),
            ("  Repairs & Maintenance", maintenance),
            ("  Utilities", utilities),
            ("  Landscaping", landscaping),
            ("  Pest Control", pest_control),
            ("  HOA Fees", hoa_fees),
        ]
        
        for label, annual in expenses:
            if annual > 0:  # Only show non-zero expenses
                ws[f'A{row}'] = label
                ws[f'B{row}'] = annual / 12
                ws[f'B{row}'].number_format = self._get_currency_format()
                ws[f'C{row}'] = annual
                ws[f'C{row}'].number_format = self._get_currency_format()
                ws[f'D{row}'] = annual / gpi if gpi > 0 else 0
                ws[f'D{row}'].number_format = self._get_percent_format()
                row += 1
        
        # Total Operating Expenses
        ws[f'A{row}'] = "Total Operating Expenses"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_operating / 12
        ws[f'B{row}'].number_format = self._get_currency_format()
        ws[f'B{row}'].font = Font(bold=True, color=self.RED)
        ws[f'C{row}'] = total_operating
        ws[f'C{row}'].number_format = self._get_currency_format()
        ws[f'C{row}'].font = Font(bold=True, color=self.RED)
        ws[f'D{row}'] = total_operating / gpi if gpi > 0 else 0
        ws[f'D{row}'].number_format = self._get_percent_format()
        ws[f'A{row}'].fill = subtotal_fill
        ws[f'B{row}'].fill = subtotal_fill
        ws[f'C{row}'].fill = subtotal_fill
        ws[f'D{row}'].fill = subtotal_fill
        row += 1
        
        # === NET OPERATING INCOME (NOI) ===
        row += 1
        ws[f'A{row}'] = "NET OPERATING INCOME (NOI)"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = noi_fill
        ws[f'B{row}'] = noi / 12
        ws[f'B{row}'].number_format = self._get_currency_format()
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'B{row}'].fill = noi_fill
        ws[f'C{row}'] = noi
        ws[f'C{row}'].number_format = self._get_currency_format()
        ws[f'C{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'C{row}'].fill = noi_fill
        ws[f'D{row}'] = noi / gpi if gpi > 0 else 0
        ws[f'D{row}'].number_format = self._get_percent_format()
        ws[f'D{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'D{row}'].fill = noi_fill
        row += 1
        
        # === DEBT SERVICE ===
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "DEBT SERVICE"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = section_fill
        
        row += 1
        ws[f'A{row}'] = "  Annual Mortgage Payment (P&I)"
        ws[f'B{row}'] = annual_debt_service / 12
        ws[f'B{row}'].number_format = self._get_currency_format()
        ws[f'C{row}'] = annual_debt_service
        ws[f'C{row}'].number_format = self._get_currency_format()
        ws[f'D{row}'] = annual_debt_service / gpi if gpi > 0 else 0
        ws[f'D{row}'].number_format = self._get_percent_format()
        row += 1
        
        # === CASH FLOW BEFORE TAX ===
        row += 1
        cf_color = self.GREEN if cash_flow >= 0 else self.RED
        ws[f'A{row}'] = "CASH FLOW BEFORE TAX"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=cf_color, end_color=cf_color, fill_type="solid")
        ws[f'B{row}'] = cash_flow / 12
        ws[f'B{row}'].number_format = self._get_currency_format()
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'B{row}'].fill = PatternFill(start_color=cf_color, end_color=cf_color, fill_type="solid")
        ws[f'C{row}'] = cash_flow
        ws[f'C{row}'].number_format = self._get_currency_format()
        ws[f'C{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'C{row}'].fill = PatternFill(start_color=cf_color, end_color=cf_color, fill_type="solid")
        ws[f'D{row}'] = cash_flow / gpi if gpi > 0 else 0
        ws[f'D{row}'].number_format = self._get_percent_format()
        ws[f'D{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'D{row}'].fill = PatternFill(start_color=cf_color, end_color=cf_color, fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
    
    # =========================================================================
    # DSCR QUALIFICATION ANALYSIS SHEET
    # =========================================================================
    
    def _create_dscr_analysis_sheet(
        self, 
        wb: Workbook, 
        property_data: Dict, 
        analytics: Dict,
        assumptions: Optional[Dict] = None
    ):
        """Create DSCR qualification analysis for lender presentation."""
        ws = wb.create_sheet("DSCR Analysis")
        
        # Styles
        title_font = Font(bold=True, size=16, color=self.BRAND_BLUE)
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color=self.NAVY, end_color=self.NAVY, fill_type="solid")
        section_fill = PatternFill(start_color=self.LIGHT_GRAY, end_color=self.LIGHT_GRAY, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get LTR data
        ltr = analytics.get("ltr", {})
        
        noi = ltr.get("noi", 16000)
        annual_debt_service = ltr.get("annual_debt_service", 20000)
        dscr = ltr.get("dscr", noi / annual_debt_service if annual_debt_service > 0 else 0)
        monthly_rent = ltr.get("monthly_rent", 2500)
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "DEBT SERVICE COVERAGE RATIO (DSCR) ANALYSIS"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Property Address
        address = property_data.get("address", {})
        ws.merge_cells('A2:E2')
        ws['A2'] = address.get("full_address", "Property Address")
        ws['A2'].font = Font(italic=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # === DSCR CALCULATION ===
        row = 4
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "DSCR CALCULATION"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        calc_items = [
            ("Net Operating Income (NOI)", f"${noi:,.0f}", "Annual"),
            ("Annual Debt Service", f"${annual_debt_service:,.0f}", "Annual"),
            ("DSCR (NOI ÷ Debt Service)", f"{dscr:.2f}x", ""),
        ]
        
        for label, value, note in calc_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True) if "DSCR" in label else Font()
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            if "DSCR" in label:
                ws[f'B{row}'].font = Font(bold=True, size=14)
                if dscr >= 1.25:
                    ws[f'B{row}'].font = Font(bold=True, size=14, color=self.GREEN)
                elif dscr >= 1.0:
                    ws[f'B{row}'].font = Font(bold=True, size=14, color=self.YELLOW)
                else:
                    ws[f'B{row}'].font = Font(bold=True, size=14, color=self.RED)
            ws[f'C{row}'] = note
            row += 1
        
        # === LENDER REQUIREMENTS ===
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "LENDER REQUIREMENTS"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        lender_requirements = [
            ("Conventional Lenders", "1.20x - 1.25x", "Standard requirement"),
            ("DSCR Loan Programs", "1.00x - 1.25x", "Investment property loans"),
            ("Commercial Lenders", "1.25x - 1.50x", "Commercial/multifamily"),
            ("SBA Loans", "1.15x - 1.25x", "Small business loans"),
        ]
        
        for lender, requirement, note in lender_requirements:
            ws[f'A{row}'] = lender
            ws[f'B{row}'] = requirement
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'] = note
            ws[f'C{row}'].font = Font(italic=True, color=self.DARK_GRAY)
            row += 1
        
        # === QUALIFICATION STATUS ===
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "QUALIFICATION STATUS"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        qualifications = [
            ("Conventional (1.25x)", dscr >= 1.25),
            ("DSCR Loan (1.00x)", dscr >= 1.00),
            ("Commercial (1.25x)", dscr >= 1.25),
        ]
        
        for loan_type, qualifies in qualifications:
            ws[f'A{row}'] = loan_type
            ws[f'B{row}'] = "✓ QUALIFIES" if qualifies else "✗ DOES NOT QUALIFY"
            ws[f'B{row}'].font = Font(bold=True, color=self.GREEN if qualifies else self.RED)
            row += 1
        
        # === IMPROVEMENT SCENARIOS ===
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "SCENARIOS TO ACHIEVE 1.25x DSCR"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        
        # Calculate rent needed for 1.25 DSCR
        required_noi_125 = annual_debt_service * 1.25
        current_expense_ratio = 1 - (noi / (monthly_rent * 12)) if monthly_rent > 0 else 0.35
        required_rent_125 = (required_noi_125 / (1 - current_expense_ratio)) / 12 if current_expense_ratio < 1 else 0
        rent_increase_needed = required_rent_125 - monthly_rent
        
        # Calculate max purchase price for 1.25 DSCR at current rent
        # This is more complex, but we can estimate
        
        ws[f'A{row}'] = "Current Monthly Rent"
        ws[f'B{row}'] = f"${monthly_rent:,.0f}"
        row += 1
        
        ws[f'A{row}'] = "Rent Needed for 1.25x DSCR"
        ws[f'B{row}'] = f"${required_rent_125:,.0f}"
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Rent Increase Required"
        if rent_increase_needed > 0:
            ws[f'B{row}'] = f"+${rent_increase_needed:,.0f}/month"
            ws[f'B{row}'].font = Font(color=self.RED)
        else:
            ws[f'B{row}'] = "Already qualified"
            ws[f'B{row}'].font = Font(color=self.GREEN)
        row += 1
        
        ws[f'A{row}'] = "Rent Increase Percentage"
        rent_pct_increase = (rent_increase_needed / monthly_rent * 100) if monthly_rent > 0 else 0
        ws[f'B{row}'] = f"+{rent_pct_increase:.1f}%" if rent_increase_needed > 0 else "0%"
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
    
    # =========================================================================
    # 10-YEAR PRO FORMA PROJECTION SHEET
    # =========================================================================
    
    def _create_pro_forma_sheet(
        self, 
        wb: Workbook, 
        property_data: Dict, 
        analytics: Dict,
        assumptions: Optional[Dict] = None
    ):
        """Create 10-year pro forma projection sheet."""
        ws = wb.create_sheet("10-Year Pro Forma")
        
        # Styles
        title_font = Font(bold=True, size=16, color=self.BRAND_BLUE)
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color=self.NAVY, end_color=self.NAVY, fill_type="solid")
        year_fill = PatternFill(start_color=self.BRAND_BLUE, end_color=self.BRAND_BLUE, fill_type="solid")
        total_fill = PatternFill(start_color=self.LIGHT_GRAY, end_color=self.LIGHT_GRAY, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get LTR data and projections
        ltr = analytics.get("ltr", {})
        projections = ltr.get("ten_year_projection", [])
        
        # Default values if no projections
        if not projections:
            monthly_rent = ltr.get("monthly_rent", 2500)
            annual_rent = monthly_rent * 12
            total_expenses = ltr.get("total_operating_expenses", 12000)
            debt_service = ltr.get("annual_debt_service", 20000)
            
            # Generate projections
            rent_growth = 0.03
            expense_growth = 0.02
            appreciation = 0.03
            purchase_price = assumptions.get("purchase_price", 400000) if assumptions else 400000
            
            projections = []
            for year in range(1, 11):
                year_rent = annual_rent * ((1 + rent_growth) ** (year - 1))
                year_expenses = total_expenses * ((1 + expense_growth) ** (year - 1))
                year_noi = year_rent * 0.95 - year_expenses  # 5% vacancy
                year_cf = year_noi - debt_service
                property_value = purchase_price * ((1 + appreciation) ** year)
                
                projections.append({
                    "year": year,
                    "gross_rent": year_rent,
                    "operating_expenses": year_expenses,
                    "noi": year_noi,
                    "debt_service": debt_service,
                    "cash_flow": year_cf,
                    "property_value": property_value,
                })
        
        # Title
        ws.merge_cells('A1:L1')
        ws['A1'] = "10-YEAR PRO FORMA PROJECTION"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Property Address
        address = property_data.get("address", {})
        ws.merge_cells('A2:L2')
        ws['A2'] = address.get("full_address", "Property Address")
        ws['A2'].font = Font(italic=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Column Headers
        row = 4
        headers = ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", 
                   "Year 6", "Year 7", "Year 8", "Year 9", "Year 10", "Total/Avg"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Row labels and data
        row = 5
        metrics = [
            ("Gross Rental Income", "gross_rent", "sum"),
            ("Operating Expenses", "operating_expenses", "sum"),
            ("Net Operating Income", "noi", "sum"),
            ("Debt Service", "debt_service", "sum"),
            ("Cash Flow", "cash_flow", "sum"),
            ("Property Value", "property_value", "last"),
        ]
        
        for label, key, agg_type in metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=1).border = thin_border
            
            total = 0
            for i, proj in enumerate(projections[:10], 2):
                value = proj.get(key, 0)
                cell = ws.cell(row=row, column=i, value=value)
                cell.number_format = self._get_currency_format()
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
                
                # Color code cash flow
                if key == "cash_flow":
                    if value >= 0:
                        cell.font = Font(color=self.GREEN)
                    else:
                        cell.font = Font(color=self.RED)
                
                if agg_type == "sum":
                    total += value
                else:
                    total = value  # Last value for property value
            
            # Total/Average column
            total_cell = ws.cell(row=row, column=12, value=total)
            total_cell.number_format = self._get_currency_format()
            total_cell.font = Font(bold=True)
            total_cell.fill = total_fill
            total_cell.border = thin_border
            
            row += 1
        
        # Add cumulative cash flow row
        row += 1
        ws.cell(row=row, column=1, value="Cumulative Cash Flow").font = Font(bold=True)
        cumulative = 0
        for i, proj in enumerate(projections[:10], 2):
            cumulative += proj.get("cash_flow", 0)
            cell = ws.cell(row=row, column=i, value=cumulative)
            cell.number_format = self._get_currency_format()
            cell.border = thin_border
            if cumulative >= 0:
                cell.font = Font(color=self.GREEN)
            else:
                cell.font = Font(color=self.RED)
        
        ws.cell(row=row, column=12, value=cumulative)
        ws.cell(row=row, column=12).number_format = self._get_currency_format()
        ws.cell(row=row, column=12).font = Font(bold=True)
        ws.cell(row=row, column=12).fill = total_fill
        
        # Add equity row
        row += 1
        ws.cell(row=row, column=1, value="Estimated Equity").font = Font(bold=True)
        for i, proj in enumerate(projections[:10], 2):
            equity = proj.get("equity", proj.get("property_value", 0) * 0.25)  # Estimate if not provided
            cell = ws.cell(row=row, column=i, value=equity)
            cell.number_format = self._get_currency_format()
            cell.border = thin_border
            cell.font = Font(color=self.BRAND_BLUE)
        
        # Total Return Analysis
        row += 3
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "10-YEAR INVESTMENT SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        total_cash_flow = sum(p.get("cash_flow", 0) for p in projections[:10])
        final_property_value = projections[-1].get("property_value", 0) if projections else 0
        initial_investment = assumptions.get("total_cash_required", 92000) if assumptions else 92000
        
        summary_items = [
            ("Total Cash Flow (10 Years)", f"${total_cash_flow:,.0f}"),
            ("Final Property Value", f"${final_property_value:,.0f}"),
            ("Initial Investment", f"${initial_investment:,.0f}"),
            ("Total Return", f"${total_cash_flow + final_property_value - initial_investment:,.0f}"),
        ]
        
        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 22
        for col in range(2, 13):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    # =========================================================================
    # AMORTIZATION SCHEDULE SHEET
    # =========================================================================
    
    def _create_amortization_sheet(
        self, 
        wb: Workbook, 
        property_data: Dict, 
        analytics: Dict,
        assumptions: Optional[Dict] = None
    ):
        """Create loan amortization schedule sheet."""
        ws = wb.create_sheet("Amortization Schedule")
        
        # Styles
        title_font = Font(bold=True, size=16, color=self.BRAND_BLUE)
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color=self.NAVY, end_color=self.NAVY, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get loan parameters
        ltr = analytics.get("ltr", {})
        loan_amount = ltr.get("loan_amount", 320000)
        interest_rate = assumptions.get("interest_rate", 0.075) if assumptions else 0.075
        loan_term = assumptions.get("loan_term_years", 30) if assumptions else 30
        monthly_payment = ltr.get("monthly_pi", 0)
        
        if monthly_payment == 0:
            # Calculate monthly payment
            monthly_rate = interest_rate / 12
            num_payments = loan_term * 12
            if monthly_rate > 0:
                monthly_payment = loan_amount * (monthly_rate * (1 + monthly_rate) ** num_payments) / \
                                  ((1 + monthly_rate) ** num_payments - 1)
            else:
                monthly_payment = loan_amount / num_payments
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "LOAN AMORTIZATION SCHEDULE"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Loan Summary
        row = 3
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "LOAN SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        loan_info = [
            ("Original Loan Amount", f"${loan_amount:,.0f}"),
            ("Interest Rate", f"{interest_rate*100:.3f}%"),
            ("Loan Term", f"{loan_term} years"),
            ("Monthly Payment (P&I)", f"${monthly_payment:,.2f}"),
            ("Total Payments", f"${monthly_payment * loan_term * 12:,.0f}"),
            ("Total Interest", f"${monthly_payment * loan_term * 12 - loan_amount:,.0f}"),
        ]
        
        for label, value in loan_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        # Annual Amortization Table
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "ANNUAL AMORTIZATION (YEAR-BY-YEAR)"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        headers = ["Year", "Beginning Balance", "Principal Paid", "Interest Paid", 
                   "Total Payment", "Ending Balance", "Equity Built"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=self.BRAND_BLUE, end_color=self.BRAND_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Calculate amortization
        row += 1
        balance = loan_amount
        monthly_rate = interest_rate / 12
        total_principal = 0
        total_interest = 0
        
        for year in range(1, min(loan_term + 1, 31)):  # Show up to 30 years
            year_principal = 0
            year_interest = 0
            start_balance = balance
            
            for month in range(12):
                if balance <= 0:
                    break
                interest_payment = balance * monthly_rate
                principal_payment = monthly_payment - interest_payment
                year_principal += principal_payment
                year_interest += interest_payment
                balance -= principal_payment
            
            if balance < 0:
                balance = 0
            
            total_principal += year_principal
            total_interest += year_interest
            
            # Write row
            ws.cell(row=row, column=1, value=year).border = thin_border
            ws.cell(row=row, column=2, value=start_balance)
            ws.cell(row=row, column=2).number_format = self._get_currency_format()
            ws.cell(row=row, column=2).border = thin_border
            
            ws.cell(row=row, column=3, value=year_principal)
            ws.cell(row=row, column=3).number_format = self._get_currency_format()
            ws.cell(row=row, column=3).font = Font(color=self.GREEN)
            ws.cell(row=row, column=3).border = thin_border
            
            ws.cell(row=row, column=4, value=year_interest)
            ws.cell(row=row, column=4).number_format = self._get_currency_format()
            ws.cell(row=row, column=4).border = thin_border
            
            ws.cell(row=row, column=5, value=monthly_payment * 12)
            ws.cell(row=row, column=5).number_format = self._get_currency_format()
            ws.cell(row=row, column=5).border = thin_border
            
            ws.cell(row=row, column=6, value=balance)
            ws.cell(row=row, column=6).number_format = self._get_currency_format()
            ws.cell(row=row, column=6).border = thin_border
            
            ws.cell(row=row, column=7, value=total_principal)
            ws.cell(row=row, column=7).number_format = self._get_currency_format()
            ws.cell(row=row, column=7).font = Font(bold=True, color=self.BRAND_BLUE)
            ws.cell(row=row, column=7).border = thin_border
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
    
    # =========================================================================
    # SENSITIVITY ANALYSIS SHEET
    # =========================================================================
    
    def _create_sensitivity_sheet(
        self, 
        wb: Workbook, 
        property_data: Dict, 
        analytics: Dict,
        assumptions: Optional[Dict] = None
    ):
        """Create sensitivity analysis sheet showing what-if scenarios."""
        ws = wb.create_sheet("Sensitivity Analysis")
        
        # Styles
        title_font = Font(bold=True, size=16, color=self.BRAND_BLUE)
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color=self.NAVY, end_color=self.NAVY, fill_type="solid")
        scenario_fill = PatternFill(start_color=self.LIGHT_GRAY, end_color=self.LIGHT_GRAY, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get base values
        ltr = analytics.get("ltr", {})
        base_rent = ltr.get("monthly_rent", 2500)
        base_noi = ltr.get("noi", 16000)
        base_cf = ltr.get("annual_cash_flow", 8000)
        base_coc = ltr.get("cash_on_cash_return", 0.08)
        base_dscr = ltr.get("dscr", 1.2)
        
        purchase_price = assumptions.get("purchase_price", 400000) if assumptions else 400000
        vacancy_rate = assumptions.get("vacancy_rate", 0.05) if assumptions else 0.05
        interest_rate = assumptions.get("interest_rate", 0.075) if assumptions else 0.075
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "SENSITIVITY ANALYSIS - WHAT-IF SCENARIOS"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # === RENT SENSITIVITY ===
        row = 4
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "RENT SENSITIVITY ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        rent_headers = ["Monthly Rent", "Change", "Annual Income", "Est. Cash Flow", "Cash-on-Cash", "DSCR", "Status"]
        for col, header in enumerate(rent_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=self.BRAND_BLUE, end_color=self.BRAND_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row += 1
        rent_scenarios = [-20, -10, -5, 0, 5, 10, 20]
        
        for pct in rent_scenarios:
            rent = base_rent * (1 + pct/100)
            annual_income = rent * 12
            # Estimate new metrics (simplified)
            income_change = (rent - base_rent) * 12 * 0.95  # After vacancy
            est_cf = base_cf + income_change
            est_noi = base_noi + income_change
            est_dscr = est_noi / (base_noi / base_dscr) if base_dscr > 0 else 0
            est_coc = est_cf / (purchase_price * 0.23) if purchase_price > 0 else 0  # Approx cash required
            
            ws.cell(row=row, column=1, value=rent)
            ws.cell(row=row, column=1).number_format = self._get_currency_format()
            
            ws.cell(row=row, column=2, value=f"{pct:+d}%")
            if pct == 0:
                ws.cell(row=row, column=2).fill = scenario_fill
                ws.cell(row=row, column=2).font = Font(bold=True)
            
            ws.cell(row=row, column=3, value=annual_income)
            ws.cell(row=row, column=3).number_format = self._get_currency_format()
            
            ws.cell(row=row, column=4, value=est_cf)
            ws.cell(row=row, column=4).number_format = self._get_currency_format()
            ws.cell(row=row, column=4).font = Font(color=self.GREEN if est_cf > 0 else self.RED)
            
            ws.cell(row=row, column=5, value=est_coc)
            ws.cell(row=row, column=5).number_format = self._get_percent_format()
            
            ws.cell(row=row, column=6, value=est_dscr)
            ws.cell(row=row, column=6).number_format = '0.00"x"'
            
            status = "✓ Strong" if est_dscr >= 1.25 and est_cf > 0 else "⚠ Marginal" if est_dscr >= 1.0 else "✗ Weak"
            ws.cell(row=row, column=7, value=status)
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            
            if pct == 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = scenario_fill
            
            row += 1
        
        # === VACANCY SENSITIVITY ===
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "VACANCY RATE SENSITIVITY"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        vac_headers = ["Vacancy Rate", "Vacancy Loss", "Effective Income", "Est. Cash Flow", "Cash-on-Cash", "Cash Flow+", "Risk Level"]
        for col, header in enumerate(vac_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=self.BRAND_BLUE, end_color=self.BRAND_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row += 1
        vacancy_scenarios = [0, 0.03, 0.05, 0.08, 0.10, 0.15, 0.20]
        
        for vac in vacancy_scenarios:
            gross = base_rent * 12
            vacancy_loss = gross * vac
            egi = gross - vacancy_loss
            # Estimate new cash flow
            base_vacancy_loss = gross * vacancy_rate
            cf_change = base_vacancy_loss - vacancy_loss
            est_cf = base_cf + cf_change
            est_coc = est_cf / (purchase_price * 0.23) if purchase_price > 0 else 0
            
            ws.cell(row=row, column=1, value=vac)
            ws.cell(row=row, column=1).number_format = self._get_percent_format()
            
            ws.cell(row=row, column=2, value=vacancy_loss)
            ws.cell(row=row, column=2).number_format = self._get_currency_format()
            ws.cell(row=row, column=2).font = Font(color=self.RED)
            
            ws.cell(row=row, column=3, value=egi)
            ws.cell(row=row, column=3).number_format = self._get_currency_format()
            
            ws.cell(row=row, column=4, value=est_cf)
            ws.cell(row=row, column=4).number_format = self._get_currency_format()
            ws.cell(row=row, column=4).font = Font(color=self.GREEN if est_cf > 0 else self.RED)
            
            ws.cell(row=row, column=5, value=est_coc)
            ws.cell(row=row, column=5).number_format = self._get_percent_format()
            
            cash_flow_positive = "Yes" if est_cf > 0 else "No"
            ws.cell(row=row, column=6, value=cash_flow_positive)
            ws.cell(row=row, column=6).font = Font(color=self.GREEN if est_cf > 0 else self.RED)
            
            risk = "Low" if vac <= 0.05 else "Medium" if vac <= 0.10 else "High"
            ws.cell(row=row, column=7, value=risk)
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
            
            if vac == vacancy_rate:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = scenario_fill
            
            row += 1
        
        # === INTEREST RATE SENSITIVITY ===
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "INTEREST RATE SENSITIVITY"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        rate_headers = ["Interest Rate", "Monthly P&I", "Annual Debt", "Est. Cash Flow", "DSCR", "Impact"]
        for col, header in enumerate(rate_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color=self.BRAND_BLUE, end_color=self.BRAND_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row += 1
        rate_scenarios = [0.05, 0.055, 0.06, 0.065, 0.07, 0.075, 0.08, 0.085, 0.09]
        
        loan_amount = ltr.get("loan_amount", purchase_price * 0.80)
        base_debt_service = ltr.get("annual_debt_service", 20000)
        
        for rate in rate_scenarios:
            # Calculate new payment
            monthly_rate = rate / 12
            num_payments = 360
            if monthly_rate > 0:
                monthly_pi = loan_amount * (monthly_rate * (1 + monthly_rate) ** num_payments) / \
                            ((1 + monthly_rate) ** num_payments - 1)
            else:
                monthly_pi = loan_amount / num_payments
            
            annual_debt = monthly_pi * 12
            debt_change = annual_debt - base_debt_service
            est_cf = base_cf - debt_change
            est_dscr = base_noi / annual_debt if annual_debt > 0 else 0
            
            ws.cell(row=row, column=1, value=rate)
            ws.cell(row=row, column=1).number_format = self._get_percent_format()
            
            ws.cell(row=row, column=2, value=monthly_pi)
            ws.cell(row=row, column=2).number_format = self._get_currency_format()
            
            ws.cell(row=row, column=3, value=annual_debt)
            ws.cell(row=row, column=3).number_format = self._get_currency_format()
            
            ws.cell(row=row, column=4, value=est_cf)
            ws.cell(row=row, column=4).number_format = self._get_currency_format()
            ws.cell(row=row, column=4).font = Font(color=self.GREEN if est_cf > 0 else self.RED)
            
            ws.cell(row=row, column=5, value=est_dscr)
            ws.cell(row=row, column=5).number_format = '0.00"x"'
            
            impact = f"${debt_change:+,.0f}/yr"
            ws.cell(row=row, column=6, value=impact)
            ws.cell(row=row, column=6).font = Font(color=self.RED if debt_change > 0 else self.GREEN)
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
            
            if abs(rate - interest_rate) < 0.001:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = scenario_fill
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
    
    def _create_summary_sheet(
        self, 
        wb: Workbook, 
        property_data: Dict, 
        analytics: Dict,
        active_strategy: str = "ltr"
    ):
        """Create the summary comparison sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color=self.BRAND_BLUE, end_color=self.BRAND_BLUE, fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        money_font = Font(bold=True, size=12)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "DealGapIQ Property Analysis Report"
        ws['A1'].font = Font(bold=True, size=18, color=self.BRAND_BLUE)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Property Info
        ws['A3'] = "Property Address:"
        ws['A3'].font = Font(bold=True)
        address = property_data.get("address", {})
        ws['B3'] = address.get("full_address", "N/A")
        
        ws['A4'] = "Generated:"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        
        # Property Details
        details = property_data.get("details", {})
        ws['A6'] = "Property Details"
        ws['A6'].font = Font(bold=True, size=12)
        ws.merge_cells('A6:B6')
        
        row = 7
        for label, key in [
            ("Property Type", "property_type"),
            ("Bedrooms", "bedrooms"),
            ("Bathrooms", "bathrooms"),
            ("Square Footage", "square_footage"),
            ("Year Built", "year_built"),
            ("Lot Size (sqft)", "lot_size_sqft"),
        ]:
            ws[f'A{row}'] = label
            value = details.get(key, "N/A")
            if key == "square_footage" and value != "N/A":
                value = f"{value:,}"
            ws[f'B{row}'] = value
            row += 1
        
        # Valuation
        valuations = property_data.get("valuations", {})
        ws['A14'] = "Valuations"
        ws['A14'].font = Font(bold=True, size=12)
        ws.merge_cells('A14:B14')
        
        row = 15
        for label, key in [
            ("Current Value (AVM)", "current_value_avm"),
            ("Zestimate", "zestimate"),
            ("Tax Assessment", "tax_assessment"),
        ]:
            ws[f'A{row}'] = label
            value = valuations.get(key)
            ws[f'B{row}'] = f"${value:,.0f}" if value else "N/A"
            row += 1
        
        # Strategy Comparison Header
        ws['A20'] = "Strategy Comparison"
        ws['A20'].font = Font(bold=True, size=14)
        ws.merge_cells('A20:G20')
        
        # Column headers
        headers = ["Strategy", "Cash Required", "Monthly Cash Flow", "Annual Cash Flow", "CoC Return", "Key Metric", "Recommendation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=21, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Strategy data
        strategies = [
            ("Long-Term Rental", "ltr", "cap_rate", "Cap Rate"),
            ("Short-Term Rental", "str", "break_even_occupancy", "Break-Even Occ."),
            ("BRRRR", "brrrr", "cash_left_in_deal", "Cash Left"),
            ("Fix & Flip", "flip", "net_profit_before_tax", "Net Profit"),
            ("House Hack", "house_hack", "savings_vs_renting_a", "Monthly Savings"),
            ("Wholesale", "wholesale", "assignment_fee", "Assignment Fee"),
        ]
        
        row = 22
        for name, key, metric_key, metric_label in strategies:
            data = analytics.get(key, {})
            if not data:
                continue
                
            ws.cell(row=row, column=1, value=name).border = thin_border
            
            # Cash Required
            cash_req = data.get("total_cash_required") or data.get("total_cash_invested") or data.get("total_cash_at_risk", 0)
            ws.cell(row=row, column=2, value=f"${cash_req:,.0f}" if cash_req else "N/A").border = thin_border
            
            # Monthly Cash Flow
            monthly_cf = data.get("monthly_cash_flow", 0)
            cell = ws.cell(row=row, column=3, value=f"${monthly_cf:,.0f}" if monthly_cf else "N/A")
            cell.border = thin_border
            if monthly_cf and monthly_cf > 0:
                cell.font = Font(color="22C55E", bold=True)
            elif monthly_cf and monthly_cf < 0:
                cell.font = Font(color="EF4444", bold=True)
            
            # Annual Cash Flow
            annual_cf = data.get("annual_cash_flow", 0)
            ws.cell(row=row, column=4, value=f"${annual_cf:,.0f}" if annual_cf else "N/A").border = thin_border
            
            # CoC Return
            coc = data.get("cash_on_cash_return", 0)
            ws.cell(row=row, column=5, value=f"{coc*100:.1f}%" if coc else "N/A").border = thin_border
            
            # Key Metric
            metric_value = data.get(metric_key)
            if metric_key in ["cap_rate", "break_even_occupancy"]:
                display = f"{metric_value*100:.1f}%" if metric_value else "N/A"
            elif metric_key in ["net_profit_before_tax", "cash_left_in_deal", "assignment_fee", "savings_vs_renting_a"]:
                display = f"${metric_value:,.0f}" if metric_value else "N/A"
            else:
                display = str(metric_value) if metric_value else "N/A"
            ws.cell(row=row, column=6, value=display).border = thin_border
            
            # Recommendation
            rec = self._get_recommendation(key, data)
            ws.cell(row=row, column=7, value=rec).border = thin_border
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 20
    
    def _get_recommendation(self, strategy: str, data: Dict) -> str:
        """Get recommendation emoji and text for a strategy."""
        if strategy == "ltr":
            coc = data.get("cash_on_cash_return", 0)
            if coc >= 0.08:
                return "✅ Strong"
            elif coc >= 0.05:
                return "⚠️ Moderate"
            else:
                return "❌ Weak"
        elif strategy == "str":
            occ = data.get("break_even_occupancy", 1)
            if occ <= 0.5:
                return "✅ Strong"
            elif occ <= 0.7:
                return "⚠️ Moderate"
            else:
                return "❌ Risky"
        elif strategy == "brrrr":
            infinite = data.get("infinite_roi_achieved", False)
            if infinite:
                return "✅ Infinite ROI"
            cash_left = data.get("cash_left_in_deal", 0)
            if cash_left <= 10000:
                return "⚠️ Low Cash"
            else:
                return "❌ High Cash"
        elif strategy == "flip":
            meets_70 = data.get("meets_70_rule", False)
            if meets_70:
                return "✅ Meets 70%"
            else:
                return "⚠️ Above 70%"
        elif strategy == "house_hack":
            savings = data.get("savings_vs_renting_a", 0)
            if savings >= 500:
                return "✅ Great Savings"
            elif savings >= 0:
                return "⚠️ Some Savings"
            else:
                return "❌ Negative"
        elif strategy == "wholesale":
            viability = data.get("deal_viability", "")
            if "good" in viability.lower():
                return "✅ Good Deal"
            elif "marginal" in viability.lower():
                return "⚠️ Marginal"
            else:
                return "❌ Poor"
        return "—"
    
    def _create_ltr_sheet(self, wb: Workbook, property_data: Dict, ltr_data: Optional[Dict]):
        """Create Long-Term Rental analysis sheet."""
        if not ltr_data:
            return
            
        ws = wb.create_sheet("Long-Term Rental")
        self._create_strategy_sheet(ws, "Long-Term Rental", self.STRATEGY_COLORS["ltr"], ltr_data, [
            ("Monthly Gross Rent", "monthly_gross_rent", "currency"),
            ("Vacancy Loss", "vacancy_loss", "currency"),
            ("Net Operating Income", "noi", "currency"),
            ("Monthly Mortgage (P&I)", "monthly_mortgage", "currency"),
            ("Monthly Cash Flow", "monthly_cash_flow", "currency"),
            ("Annual Cash Flow", "annual_cash_flow", "currency"),
            ("Cash-on-Cash Return", "cash_on_cash_return", "percent"),
            ("Cap Rate", "cap_rate", "percent"),
            ("DSCR", "dscr", "number"),
            ("Total Cash Required", "total_cash_required", "currency"),
            ("1% Rule", "one_percent_rule", "percent"),
        ])
    
    def _create_str_sheet(self, wb: Workbook, property_data: Dict, str_data: Optional[Dict]):
        """Create Short-Term Rental analysis sheet."""
        if not str_data:
            return
            
        ws = wb.create_sheet("Short-Term Rental")
        self._create_strategy_sheet(ws, "Short-Term Rental", self.STRATEGY_COLORS["str"], str_data, [
            ("Average Daily Rate", "average_daily_rate", "currency"),
            ("Occupancy Rate", "occupancy_rate", "percent"),
            ("Gross Annual Revenue", "gross_annual_revenue", "currency"),
            ("Net Operating Income", "noi", "currency"),
            ("Monthly Cash Flow", "monthly_cash_flow", "currency"),
            ("Annual Cash Flow", "annual_cash_flow", "currency"),
            ("Cash-on-Cash Return", "cash_on_cash_return", "percent"),
            ("Break-Even Occupancy", "break_even_occupancy", "percent"),
            ("Total Cash Required", "total_cash_required", "currency"),
        ])
    
    def _create_brrrr_sheet(self, wb: Workbook, property_data: Dict, brrrr_data: Optional[Dict]):
        """Create BRRRR analysis sheet."""
        if not brrrr_data:
            return
            
        ws = wb.create_sheet("BRRRR")
        self._create_strategy_sheet(ws, "BRRRR", self.STRATEGY_COLORS["brrrr"], brrrr_data, [
            ("Purchase Price", "purchase_price", "currency"),
            ("Rehab Budget", "rehab_budget", "currency"),
            ("After Repair Value (ARV)", "arv", "currency"),
            ("Total Cash Invested", "total_cash_invested", "currency"),
            ("Refinance Amount", "refinance_amount", "currency"),
            ("Cash Returned", "cash_returned", "currency"),
            ("Cash Left in Deal", "cash_left_in_deal", "currency"),
            ("Post-Refi Cash Flow", "post_refi_monthly_cash_flow", "currency"),
            ("Post-Refi CoC Return", "post_refi_cash_on_cash", "percent"),
            ("Infinite ROI Achieved", "infinite_roi_achieved", "boolean"),
            ("Equity Position", "equity_position", "currency"),
        ])
    
    def _create_flip_sheet(self, wb: Workbook, property_data: Dict, flip_data: Optional[Dict]):
        """Create Fix & Flip analysis sheet."""
        if not flip_data:
            return
            
        ws = wb.create_sheet("Fix & Flip")
        self._create_strategy_sheet(ws, "Fix & Flip", self.STRATEGY_COLORS["flip"], flip_data, [
            ("Purchase Price", "purchase_price", "currency"),
            ("Rehab Budget", "rehab_budget", "currency"),
            ("After Repair Value (ARV)", "arv", "currency"),
            ("Holding Costs", "holding_costs", "currency"),
            ("Selling Costs", "selling_costs", "currency"),
            ("Total Project Cost", "total_project_cost", "currency"),
            ("Net Profit (Before Tax)", "net_profit_before_tax", "currency"),
            ("ROI", "roi", "percent"),
            ("Annualized ROI", "annualized_roi", "percent"),
            ("Meets 70% Rule", "meets_70_rule", "boolean"),
            ("Max Allowable Offer", "max_allowable_offer", "currency"),
        ])
    
    def _create_house_hack_sheet(self, wb: Workbook, property_data: Dict, hh_data: Optional[Dict]):
        """Create House Hack analysis sheet."""
        if not hh_data:
            return
            
        ws = wb.create_sheet("House Hack")
        self._create_strategy_sheet(ws, "House Hack", self.STRATEGY_COLORS["house_hack"], hh_data, [
            ("Total Rental Income", "total_rental_income", "currency"),
            ("Net Housing Cost (A)", "net_housing_cost_scenario_a", "currency"),
            ("Net Housing Cost (B)", "net_housing_cost_scenario_b", "currency"),
            ("Savings vs Renting (A)", "savings_vs_renting_a", "currency"),
            ("Savings vs Renting (B)", "savings_vs_renting_b", "currency"),
            ("ROI on Savings", "roi_on_savings", "percent"),
            ("Total Cash Required", "total_cash_required", "currency"),
            ("Live for Free", "live_for_free", "boolean"),
        ])
    
    def _create_wholesale_sheet(self, wb: Workbook, property_data: Dict, ws_data: Optional[Dict]):
        """Create Wholesale analysis sheet."""
        if not ws_data:
            return
            
        ws = wb.create_sheet("Wholesale")
        self._create_strategy_sheet(ws, "Wholesale", self.STRATEGY_COLORS["wholesale"], ws_data, [
            ("After Repair Value (ARV)", "arv", "currency"),
            ("Max Allowable Offer", "max_allowable_offer", "currency"),
            ("Your Offer Price", "your_offer_price", "currency"),
            ("Assignment Fee", "assignment_fee", "currency"),
            ("Total Cash at Risk", "total_cash_at_risk", "currency"),
            ("Net Profit", "net_profit", "currency"),
            ("ROI", "roi", "percent"),
            ("Deal Viability", "deal_viability", "text"),
        ])
    
    def _create_strategy_sheet(
        self, 
        ws, 
        title: str, 
        color: str, 
        data: Dict, 
        metrics: List[tuple]
    ):
        """Create a standardized strategy analysis sheet."""
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, size=14, color="FFFFFF")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"{title} Analysis"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        
        # Metrics
        row = 4
        for label, key, fmt in metrics:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            
            value = data.get(key)
            if value is None:
                display = "N/A"
            elif fmt == "currency":
                display = f"${value:,.0f}"
            elif fmt == "percent":
                display = f"{value*100:.1f}%"
            elif fmt == "boolean":
                display = "✅ Yes" if value else "❌ No"
            elif fmt == "number":
                display = f"{value:.2f}"
            else:
                display = str(value)
            
            cell = ws.cell(row=row, column=2, value=display)
            
            # Color positive/negative values
            if fmt == "currency" and isinstance(value, (int, float)):
                if value > 0:
                    cell.font = Font(color="22C55E", bold=True)
                elif value < 0:
                    cell.font = Font(color="EF4444", bold=True)
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_assumptions_sheet(self, wb: Workbook, assumptions: Dict):
        """Create assumptions reference sheet."""
        ws = wb.create_sheet("Assumptions")
        
        ws['A1'] = "Assumptions Used in Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        row = 3
        
        def write_section(title: str, data: Dict, row: int) -> int:
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
            row += 1
            for key, value in data.items():
                if isinstance(value, dict):
                    continue
                ws.cell(row=row, column=1, value=key.replace("_", " ").title())
                if isinstance(value, float) and value < 1:
                    ws.cell(row=row, column=2, value=f"{value*100:.1f}%")
                elif isinstance(value, (int, float)):
                    ws.cell(row=row, column=2, value=f"{value:,.2f}")
                else:
                    ws.cell(row=row, column=2, value=str(value))
                row += 1
            return row + 1
        
        for section_name, section_data in assumptions.items():
            if isinstance(section_data, dict):
                row = write_section(section_name.replace("_", " ").title(), section_data, row)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15


# Singleton instance
report_service = ReportService()

