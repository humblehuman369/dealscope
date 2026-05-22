"""Tests for comprehensive Strategy Excel export."""

from io import BytesIO
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from app.core.defaults import GROWTH
from app.schemas.analytics import IQVerdictInput
from app.schemas.property import AllAssumptions
from app.services.calculators import calculate_ltr
from app.services.comprehensive_excel_exporter import ComprehensiveExcelExporter
from app.services.verdict_assumptions import apply_verdict_input_to_assumptions, resolve_purchase_price


class TestVerdictAssumptions:
    def test_apply_purchase_price_override(self):
        a = AllAssumptions()
        inp = IQVerdictInput(list_price=300000, purchase_price=240000, monthly_rent=2000)
        keys = apply_verdict_input_to_assumptions(a, inp)
        assert a.financing.purchase_price == 240000
        assert "purchase_price" in keys
        assert resolve_purchase_price(inp, a) == 240000


class TestComprehensiveExcelExporter:
    def test_generates_summary_and_strategy_sheets(self):
        ltr = calculate_ltr(
            purchase_price=250000,
            monthly_rent=2200,
            property_taxes_annual=3600,
            down_payment_pct=0.20,
            interest_rate=0.065,
            loan_term_years=30,
            closing_costs_pct=0.03,
            vacancy_rate=0.05,
            property_management_pct=0.10,
            maintenance_pct=0.05,
            insurance_annual=2500,
            utilities_monthly=100,
            landscaping_annual=0,
            pest_control_annual=0,
            appreciation_rate=GROWTH.appreciation_rate,
            rent_growth_rate=GROWTH.rent_growth_rate,
            expense_growth_rate=GROWTH.expense_growth_rate,
        )

        analytics = {"ltr": ltr}
        property_data = {
            "address": {"full_address": "123 Main St, Austin, TX", "street": "123 Main St"},
            "details": {"bedrooms": 3, "bathrooms": 2, "square_footage": 1500},
            "valuations": {"current_value_avm": 275000, "zestimate": 270000},
        }

        mock_proforma = MagicMock()

        with patch(
            "app.services.comprehensive_excel_exporter.ProformaExcelExporter"
        ) as mock_exporter_cls:
            mock_exporter_cls.return_value.add_tabs_to_workbook = MagicMock()
            exporter = ComprehensiveExcelExporter()
            raw = exporter.generate(
                property_data=property_data,
                analytics_data=analytics,
                proforma=mock_proforma,
                assumptions={"financing": {"purchase_price": 250000}},
                active_strategy="ltr",
                include_sensitivity=False,
            )

        wb = load_workbook(BytesIO(raw))
        names = wb.sheetnames
        assert "Summary" in names
        assert "Long-Term Rental" in names
        assert "Cash Flow Statement" in names

    def test_custom_purchase_price_in_assumptions(self):
        inp = IQVerdictInput(list_price=300000, purchase_price=199000, monthly_rent=1800)
        a = AllAssumptions()
        apply_verdict_input_to_assumptions(a, inp)
        assert resolve_purchase_price(inp, a) == 199000
